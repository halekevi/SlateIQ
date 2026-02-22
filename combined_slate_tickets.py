#!/usr/bin/env python3
"""
combined_slate_tickets.py

Combined NBA + CBB Slate & Ticket Generator
Merges NBA (step8_all_direction_clean.xlsx) and CBB (step6_ranked_cbb.xlsx ELIGIBLE)
Outputs:
  - combined_slate_tickets_YYYY-MM-DD.xlsx
  - tickets_latest.json / tickets_latest.html (web-friendly, static)
  - docs/tickets_latest.json / docs/tickets_latest.html (for GitHub Pages /docs)

Sheets: SUMMARY, Full Slate, NBA Slate, CBB Slate,
        NBA 3/4/5/6-Leg tickets (Goblin/Standard/Demon/Mix),
        CBB 3/4/5/6-Leg tickets, Combined 3/4/5/6-Leg tickets,
        Cross-sport Standard Mix, Cross-sport Goblin Mix

NEW (Web):
- Adds player headshot thumbnails when an ID is available:
    NBA: uses nba_player_id (if present) -> cdn.nba.com headshot
    CBB: uses espn_player_id (if present) -> espncdn headshot
  If no ID exists, it falls back to a simple initials avatar.
- JSON includes image_url per leg.
- More helpful file-path resolution (tries script dir + recursive search if file not found)

HOTFIX:
- Fixes crash when CBB "direction" becomes a DataFrame due to duplicate columns.
  We de-duplicate columns BEFORE touching df["direction"].str.upper().
"""

from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime, timezone
from typing import Optional, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Color palette ─────────────────────────────────────────────────────────────
C = {
    "hdr": "1C1C1C",
    "hdr_nba": "1A5276",
    "hdr_cbb": "1E8449",
    "hdr_mix": "6C3483",
    "hdr_sum": "117A65",
    "hit": "27AE60",
    "miss": "E74C3C",
    "push": "F39C12",
    "tier_a": "D5F5E3",
    "tier_b": "D6EAF8",
    "tier_c": "FEF9E7",
    "tier_d": "FDEDEC",
    "goblin": "E8D5F5",
    "demon": "FDEDEC",
    "standard": "F2F3F4",
    "over": "D6EAF8",
    "under": "FDEBD0",
    "alt": "F2F3F4",
    "white": "FFFFFF",
    "nba": "EBF5FB",
    "cbb": "EAFAF1",
    "mix": "F5EEF8",
    "gold": "F9E79F",
}

PAYOUT = {
    2: {"power": 3.0, "flex": 3.0},
    3: {"power": 4.37, "flex": 1.73},
    4: {"power": 10.0, "flex": 6.0},
    5: {"power": 20.0, "flex": 10.0},
    6: {"power": 40.0, "flex": 16.0},
}


# ── Excel style helpers ───────────────────────────────────────────────────────
def side(color: str = "CCCCCC") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def hc(ws, r, c, v, bg=None, fc="FFFFFF", bold=True, sz=9, align="center"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=fc, name="Arial", size=sz)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = side()
    return cell


def dc(ws, r, c, v, bg=None, bold=False, sz=9, align="center", fc="000000", fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, name="Arial", size=sz, color=fc)
    cell.fill = PatternFill("solid", start_color=bg or C["white"])
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = side()
    if fmt:
        cell.number_format = fmt
    return cell


def sw(ws, widths: List[int]):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def tier_bg(t) -> str:
    return {"A": C["tier_a"], "B": C["tier_b"], "C": C["tier_c"], "D": C["tier_d"]}.get(
        str(t).upper(), C["white"]
    )


def pt_bg(pt) -> str:
    return {"Goblin": C["goblin"], "Demon": C["demon"], "Standard": C["standard"]}.get(pt, C["white"])


def hr_bg(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "DDDDDD"
    if v >= 0.65:
        return C["hit"]
    if v >= 0.50:
        return C["push"]
    return C["miss"]


def pct_cell(ws, r, c, val):
    nan = val is None or (isinstance(val, float) and np.isnan(val))
    bg = hr_bg(val) if not nan else "DDDDDD"
    cell = dc(ws, r, c, val if not nan else "", bg=bg, bold=True)
    if not nan:
        cell.number_format = "0%"
        cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
    return cell


def win_prob(hit_rates, _n_legs: int) -> float:
    vals = []
    for h in hit_rates:
        try:
            if h is None:
                continue
            if isinstance(h, float) and np.isnan(h):
                continue
            vals.append(float(h))
        except Exception:
            continue
    if not vals:
        return 0.0
    return float(np.prod(vals))


# ──────────────────────────────────────────────────────────────────────────────
# Path resolution helpers (fixes the “file not found” headaches)
# ──────────────────────────────────────────────────────────────────────────────
def _norm_path(p: str) -> str:
    p = (p or "").strip().strip('"').strip("'")
    p = os.path.expanduser(p)
    return os.path.abspath(p)


def _find_first_by_filename(root_dir: str, filename: str) -> Optional[str]:
    try:
        for base, _dirs, files in os.walk(root_dir):
            for f in files:
                if f.lower() == filename.lower():
                    return os.path.join(base, f)
    except Exception:
        return None
    return None


def resolve_input_path(path: str, fallback_filename: Optional[str] = None) -> str:
    """
    Tries:
    1) exact path as provided
    2) relative to script directory
    3) recursive search from script directory by filename (case-insensitive)
    """
    if not path:
        raise FileNotFoundError("Empty input path.")

    raw = path.strip().strip('"').strip("'")
    p = _norm_path(raw)
    if os.path.exists(p):
        return p

    script_dir = os.path.dirname(os.path.abspath(__file__))
    p2 = os.path.abspath(os.path.join(script_dir, raw))
    if os.path.exists(p2):
        return p2

    filename = fallback_filename or os.path.basename(raw)
    found = _find_first_by_filename(script_dir, filename)
    if found and os.path.exists(found):
        return os.path.abspath(found)

    raise FileNotFoundError(
        f"Could not find file: {path}\nTried:\n- {p}\n- {p2}\n- recursive search for: {filename}"
    )


# ──────────────────────────────────────────────────────────────────────────────
# Web outputs (static HTML + JSON) + player images
# ──────────────────────────────────────────────────────────────────────────────
def _safe_float(x):
    try:
        if x is None:
            return None
        if isinstance(x, float) and np.isnan(x):
            return None
        return float(x)
    except Exception:
        return None


def _clean_id(x) -> str:
    """Return a clean integer-like string for IDs, or ''."""
    if x is None:
        return ""
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return ""
    # handle 1628368.0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    if re.fullmatch(r"\d+", s):
        return s
    return ""


def player_initials(name: str) -> str:
    s = (name or "").strip()
    if not s:
        return "?"
    parts = [p for p in re.split(r"\s+", s) if p]
    if len(parts) == 1:
        return parts[0][:1].upper()
    return (parts[0][:1] + parts[-1][:1]).upper()


def compute_image_url(leg: dict) -> Optional[str]:
    """
    NBA:
      needs nba_player_id -> https://cdn.nba.com/headshots/nba/latest/1040x760/<id>.png
    CBB:
      needs espn_player_id -> https://a.espncdn.com/i/headshots/mens-college-basketball/players/full/<id>.png
    """
    sport = (leg.get("sport") or "").upper()
    if sport == "NBA":
        pid = _clean_id(leg.get("nba_player_id"))
        if pid:
            return f"https://cdn.nba.com/headshots/nba/latest/1040x760/{pid}.png"
        return None
    if sport == "CBB":
        eid = _clean_id(leg.get("espn_player_id"))
        if eid:
            return f"https://a.espncdn.com/i/headshots/mens-college-basketball/players/full/{eid}.png"
        return None
    return None


def ticket_groups_to_payload(all_ticket_groups, date_str, thresholds):
    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "date": date_str,
        "filters": thresholds,
        "groups": [],
    }

    for group_name, tickets, _bg in all_ticket_groups:
        if not tickets:
            continue

        group = {
            "group_name": str(group_name),
            "n_legs": int(tickets[0].get("n_legs", 0) or 0),
            "power_payout": _safe_float(tickets[0].get("power_payout")),
            "flex_payout": _safe_float(tickets[0].get("flex_payout")),
            "tickets": [],
        }

        for ti, t in enumerate(tickets, start=1):
            rows = t.get("rows", [])
            slip = {
                "ticket_no": ti,
                "avg_hit_rate": _safe_float(t.get("avg_hit_rate")),
                "avg_rank_score": _safe_float(t.get("avg_rank_score")),
                "est_win_prob": _safe_float(t.get("est_win_prob")),
                "legs": [],
            }

            for row in rows:

                def gv(field):
                    return row.get(field, "") if isinstance(row, dict) else getattr(row, field, "")

                leg = {
                    "sport": str(gv("sport") or ""),
                    "player": str(gv("player") or ""),
                    "team": str(gv("team") or ""),
                    "opp": str(gv("opp") or ""),
                    "prop_type": str(gv("prop_type") or ""),
                    "pick_type": str(gv("pick_type") or ""),
                    "direction": str(gv("direction") or ""),
                    "line": _safe_float(gv("line")),
                    "edge": _safe_float(gv("edge")),
                    "hit_rate": _safe_float(gv("hit_rate")),
                    "rank_score": _safe_float(gv("rank_score")),
                    "game_time": str(gv("game_time") or ""),
                    "nba_player_id": gv("nba_player_id"),
                    "espn_player_id": gv("espn_player_id"),
                }
                leg["image_url"] = compute_image_url(leg)
                leg["initials"] = player_initials(leg.get("player", ""))

                slip["legs"].append(leg)

            group["tickets"].append(slip)

        payload["groups"].append(group)

    return payload


def write_web_outputs(payload, outdir: str):
    os.makedirs(outdir, exist_ok=True)
    json_path = os.path.join(outdir, "tickets_latest.json")
    html_path = os.path.join(outdir, "tickets_latest.html")

    def fmt_pct(x) -> str:
        try:
            if x is None:
                return ""
            return f"{float(x) * 100:.2f}%"
        except Exception:
            return ""

    def fmt_2(x) -> str:
        try:
            if x is None:
                return ""
            return f"{float(x):.2f}"
        except Exception:
            return ""

    def fmt_line(x) -> str:
        # keep lines readable (avoid 5.5000000003)
        try:
            if x is None:
                return ""
            xf = float(x)
            if abs(xf - round(xf)) < 1e-9:
                return str(int(round(xf)))
            return f"{xf:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(x) if x is not None else ""

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    html = []
    html.append("<!doctype html><html><head><meta charset='utf-8'/>")
    html.append("<meta name='viewport' content='width=device-width, initial-scale=1'/>")
    html.append("<title>Latest Tickets</title>")
    html.append("<style>")
    html.append("body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial;margin:18px;}")
    html.append("a{color:#1a5276;text-decoration:none} a:hover{text-decoration:underline}")
    html.append(".top{display:flex;gap:12px;flex-wrap:wrap;align-items:center;}")
    html.append(".pill{background:#f2f3f4;border:1px solid #ddd;border-radius:999px;padding:6px 10px;font-size:13px;}")
    html.append(".group{border:1px solid #ddd;border-radius:12px;padding:14px;margin:14px 0;}")
    html.append(".ticket{border:1px solid #eee;border-radius:10px;padding:12px;margin:10px 0;background:#fff;}")
    html.append("h1{margin:4px 0 10px 0;font-size:22px}")
    html.append("h2{margin:0 0 6px 0;font-size:18px}")
    html.append("h3{margin:0 0 8px 0;font-size:15px}")
    html.append(".muted{color:#666;font-size:13px}")
    html.append("table{width:100%;border-collapse:collapse;margin-top:8px;}")
    html.append("th,td{padding:8px;border-bottom:1px solid #eee;text-align:left;font-size:13px;vertical-align:middle;}")
    html.append("th{background:#1c1c1c;color:#fff;}")
    html.append(".dir-over{background:#d6eaf8;padding:2px 6px;border-radius:6px;font-weight:600}")
    html.append(".dir-under{background:#fdebd0;padding:2px 6px;border-radius:6px;font-weight:600}")
    html.append(".pwrap{display:flex;gap:10px;align-items:center;}")
    html.append(".avatar{width:34px;height:34px;border-radius:999px;overflow:hidden;border:1px solid #ddd;flex:0 0 auto;background:#f2f3f4;display:flex;align-items:center;justify-content:center;font-weight:700;color:#333;}")
    html.append(".avatar img{width:100%;height:100%;object-fit:cover;display:block;}")
    html.append("</style></head><body>")

    html.append("<div class='top'>")
    html.append("<h1>🎟️ Latest Generated Tickets</h1>")
    html.append("</div>")

    html.append(f"<div class='muted'>Generated: {payload.get('generated_at','')} | Date: {payload.get('date','')}</div>")
    html.append("<div class='top' style='margin-top:10px'>")
    html.append("<span class='pill'>Outputs are static (GitHub Pages friendly)</span>")
    html.append("<a class='pill' href='tickets_latest.json'>Download JSON</a>")
    html.append("<a class='pill' href='index.html'>Home</a>")
    html.append("</div>")

    filters = payload.get("filters", {})
    html.append("<div style='margin-top:10px' class='pill'>")
    html.append(
        f"Filters → tiers: {filters.get('tiers','ALL')} | min_hit_rate: {filters.get('min_hit_rate',0)} | "
        f"min_edge: {filters.get('min_edge',0)} | min_rank: {filters.get('min_rank','None')} | "
        f"pick_types: {filters.get('pick_types','ALL')}"
    )
    html.append("</div>")

    for g in payload.get("groups", []):
        html.append("<div class='group'>")
        html.append(f"<h2>{g.get('group_name','Group')}</h2>")
        html.append(
            f"<div class='muted'>Legs: {g.get('n_legs','')} | Power: {g.get('power_payout','')}x | Flex: {g.get('flex_payout','')}x</div>"
        )

        for t in g.get("tickets", []):
            avg_hr = t.get("avg_hit_rate")
            avg_rs = t.get("avg_rank_score")
            wp = t.get("est_win_prob")
            html.append("<div class='ticket'>")
            html.append(f"<h3>Ticket #{t.get('ticket_no','')}</h3>")
            html.append(
                f"<div class='muted'>Avg hit rate: {fmt_pct(avg_hr)} | "
                f"Est win prob: {fmt_pct(wp)} | "
                f"Avg rank: {fmt_2(avg_rs)}</div>"
            )

            html.append(
                "<table><thead><tr>"
                "<th>#</th><th>Sport</th><th>Player</th><th>Prop</th><th>Line</th>"
                "<th>Pick</th><th>Dir</th><th>HitRate</th><th>Edge</th><th>Rank</th>"
                "</tr></thead><tbody>"
            )

            for i, leg in enumerate(t.get("legs", []), start=1):
                dirv = (leg.get("direction") or "").upper()
                dir_span = (
                    "<span class='dir-over'>OVER</span>"
                    if dirv == "OVER"
                    else f"<span class='dir-under'>{dirv or ''}</span>"
                )

                img = leg.get("image_url")
                initials = leg.get("initials") or "?"
                if img:
                    avatar = f"<div class='avatar'><img src='{img}' alt='{initials}' onerror=\"this.remove();\"></div>"
                else:
                    avatar = f"<div class='avatar'>{initials}</div>"

                player_cell = (
                    "<div class='pwrap'>"
                    f"{avatar}"
                    f"<div>{leg.get('player','')}</div>"
                    "</div>"
                )

                html.append(
                    "<tr>"
                    f"<td>{i}</td>"
                    f"<td>{leg.get('sport','')}</td>"
                    f"<td>{player_cell}</td>"
                    f"<td>{leg.get('prop_type','')}</td>"
                    f"<td>{fmt_line(leg.get('line'))}</td>"
                    f"<td>{leg.get('pick_type','')}</td>"
                    f"<td>{dir_span}</td>"
                    f"<td>{fmt_pct(leg.get('hit_rate')) if leg.get('hit_rate') is not None else ''}</td>"
                    f"<td>{fmt_2(leg.get('edge')) if leg.get('edge') is not None else ''}</td>"
                    f"<td>{fmt_2(leg.get('rank_score')) if leg.get('rank_score') is not None else ''}</td>"
                    "</tr>"
                )

            html.append("</tbody></table>")
            html.append("</div>")  # ticket

        html.append("</div>")  # group

    html.append("</body></html>")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))

    print(f"✅ Web JSON  -> {json_path}")
    print(f"✅ Web HTML  -> {html_path}")


# ── Load & normalize NBA ───────────────────────────────────────────────────────
def load_nba(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_all_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(
        columns={
            "Tier": "tier",
            "Rank Score": "rank_score",
            "Player": "player",
            "Pos": "pos",
            "Team": "team",
            "Opp": "opp",
            "Game Time": "game_time",
            "Prop": "prop_type",
            "Pick Type": "pick_type",
            "Line": "line",
            "Direction": "direction",
            "Edge": "edge",
            "Projection": "projection",
            "Hit Rate (5g)": "hit_rate",
            "Last 5 Avg": "l5_avg",
            "Season Avg": "season_avg",
            "L5 Over": "l5_over",
            "L5 Under": "l5_under",
            "Def Rank": "def_rank",
            "Def Tier": "def_tier",
            "Min Tier": "min_tier",
            "Shot Role": "shot_role",
            "Usage Role": "usage_role",
            "Void Reason": "void_reason",
            # OPTIONAL if your NBA file has it:
            "nba_player_id": "nba_player_id",
            "NBA Player ID": "nba_player_id",
            "player_id": "nba_player_id",
            "Player ID": "nba_player_id",
        }
    )

    # ✅ IMPORTANT: de-dupe before using any column as Series
    df = df.loc[:, ~df.columns.duplicated()].copy()

    df["sport"] = "NBA"

    if "direction" in df.columns:
        if isinstance(df["direction"], pd.DataFrame):
            df["direction"] = df["direction"].iloc[:, 0]
        df["direction"] = df["direction"].astype(str).str.upper()

    if "tier" in df.columns:
        if isinstance(df["tier"], pd.DataFrame):
            df["tier"] = df["tier"].iloc[:, 0]
        df["tier"] = df["tier"].astype(str).str.upper()

    # Drop voids if present
    if "void_reason" in df.columns:
        if isinstance(df["void_reason"], pd.DataFrame):
            df["void_reason"] = df["void_reason"].iloc[:, 0]
        df = df[df["void_reason"].isna() | (df["void_reason"].astype(str).str.strip() == "")]

    # Clean ID if present
    if "nba_player_id" in df.columns:
        df["nba_player_id"] = df["nba_player_id"].apply(_clean_id)

    return df


# ── Load & normalize CBB ───────────────────────────────────────────────────────
def load_cbb(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step6_ranked_cbb.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = (
        "ELIGIBLE"
        if "ELIGIBLE" in xl.sheet_names
        else ("ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    )
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(
        columns={
            "final_bet_direction": "direction",
            "bet_direction": "direction",
            "opp_team_abbr": "opp",
            "start_time": "game_time",
            "line_hit_rate": "hit_rate",
            "stat_last5_avg": "l5_avg",
            "stat_season_avg": "season_avg",
            "line_hits_over_5": "l5_over",
            "line_hits_under_5": "l5_under",
            # OPTIONAL IDs
            "espn_player_id": "espn_player_id",
            "ESPN Player ID": "espn_player_id",
            "player_id": "espn_player_id",
        }
    )

    # ✅ CRITICAL HOTFIX: de-duplicate columns BEFORE df["direction"].str.upper()
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # ✅ If direction is still a DataFrame for any reason, take the first column.
    if "direction" in df.columns and isinstance(df["direction"], pd.DataFrame):
        df["direction"] = df["direction"].iloc[:, 0]

    df["sport"] = "CBB"

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()

    if "tier" in df.columns:
        if isinstance(df["tier"], pd.DataFrame):
            df["tier"] = df["tier"].iloc[:, 0]
        df["tier"] = df["tier"].astype(str).str.upper()

    if "void_reason" in df.columns:
        if isinstance(df["void_reason"], pd.DataFrame):
            df["void_reason"] = df["void_reason"].iloc[:, 0]
        df = df[df["void_reason"].isna() | (df["void_reason"].astype(str).str.strip() == "")]

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    return df


# ── Merge to full slate ────────────────────────────────────────────────────────
def build_combined_slate(nba: pd.DataFrame, cbb: pd.DataFrame) -> pd.DataFrame:
    keep = [
        "sport",
        "tier",
        "rank_score",
        "player",
        "team",
        "opp",
        "game_time",
        "prop_type",
        "pick_type",
        "line",
        "direction",
        "edge",
        "projection",
        "hit_rate",
        "l5_avg",
        "season_avg",
        "l5_over",
        "l5_under",
        "def_tier",
        "nba_player_id",
        "espn_player_id",
    ]

    def safe_keep(df, cols):
        df = df.loc[:, ~df.columns.duplicated()].copy()
        return df[[c for c in cols if c in df.columns]].copy()

    combined = pd.concat([safe_keep(nba, keep), safe_keep(cbb, keep)], ignore_index=True)

    if "rank_score" in combined.columns:
        combined["rank_score"] = pd.to_numeric(combined["rank_score"], errors="coerce")
    if "hit_rate" in combined.columns:
        combined["hit_rate"] = pd.to_numeric(combined["hit_rate"], errors="coerce")
    if "edge" in combined.columns:
        combined["edge"] = pd.to_numeric(combined["edge"], errors="coerce")

    combined = combined.sort_values("rank_score", ascending=False, na_position="last").reset_index(drop=True)
    return combined


# ── Filter eligible props for tickets ─────────────────────────────────────────
def filter_eligible(df: pd.DataFrame, min_hit_rate=0.0, min_edge=0.0, min_rank=None, tiers=None, pick_types=None):
    mask = pd.Series([True] * len(df), index=df.index)
    if min_hit_rate > 0 and "hit_rate" in df.columns:
        mask &= df["hit_rate"].fillna(0) >= min_hit_rate
    if min_edge > 0 and "edge" in df.columns:
        mask &= df["edge"].fillna(0) >= min_edge
    if min_rank is not None and "rank_score" in df.columns:
        mask &= df["rank_score"].fillna(-99) >= min_rank
    if tiers and "tier" in df.columns:
        mask &= df["tier"].isin([t.upper() for t in tiers])
    if pick_types and "pick_type" in df.columns:
        mask &= df["pick_type"].isin(pick_types)
    return df[mask].copy()


# ── Build tickets ──────────────────────────────────────────────────────────────
def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets=20, require_mix=False) -> list:
    """
    Build top tickets of n_legs from pool, sorted by avg rank score.
    If require_mix=True, each ticket must contain at least 1 NBA and 1 CBB leg.
    """
    pool = pool.copy().reset_index(drop=True)
    tickets = []

    has_sport_col = "sport" in pool.columns
    sports_available = pool["sport"].dropna().unique().tolist() if has_sport_col else []
    can_mix = require_mix and has_sport_col and len(sports_available) >= 2

    eligible = pool.sort_values("rank_score", ascending=False, na_position="last").reset_index(drop=True)

    for _ in range(max_tickets * 5):
        if len(tickets) >= max_tickets:
            break

        ticket_rows = []
        ticket_players = set()
        sports_in_ticket = set()

        if can_mix:
            for sport in sports_available:
                sport_pool = eligible[eligible["sport"] == sport]
                for _, row in sport_pool.iterrows():
                    player = str(row.get("player", "")).strip().lower()
                    if player and player not in ticket_players:
                        ticket_rows.append(row)
                        ticket_players.add(player)
                        sports_in_ticket.add(sport)
                        break

            for _, row in eligible.iterrows():
                if len(ticket_rows) == n_legs:
                    break
                player = str(row.get("player", "")).strip().lower()
                if player and player not in ticket_players:
                    ticket_rows.append(row)
                    ticket_players.add(player)
                    sports_in_ticket.add(row.get("sport", ""))
        else:
            for _, row in eligible.iterrows():
                if len(ticket_rows) == n_legs:
                    break
                player = str(row.get("player", "")).strip().lower()
                if player and player not in ticket_players:
                    ticket_rows.append(row)
                    ticket_players.add(player)

        if len(ticket_rows) == n_legs:
            if can_mix and len(sports_in_ticket) < 2:
                if len(eligible) > 1:
                    eligible = eligible.iloc[1:].reset_index(drop=True)
                continue

            if can_mix:
                ticket_rows = sorted(
                    ticket_rows,
                    key=lambda r: (str(r.get("sport", "")), -float(r.get("rank_score", 0) or 0)),
                )

            key = frozenset(
                (str(r.get("player", "")) + "|" + str(r.get("prop_type", ""))).strip() for r in ticket_rows
            )

            if key not in [t["key"] for t in tickets]:
                hrs = []
                rss = []
                for r in ticket_rows:
                    hrs.append(float(r.get("hit_rate", 0.5) or 0.5))
                    rss.append(float(r.get("rank_score", 0) or 0))
                avg_hr = float(np.mean(hrs)) if hrs else 0.0
                avg_rs = float(np.mean(rss)) if rss else 0.0
                ep = win_prob(hrs, n_legs)
                pout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})
                tickets.append(
                    {
                        "key": key,
                        "rows": ticket_rows,
                        "avg_hit_rate": avg_hr,
                        "avg_rank_score": avg_rs,
                        "est_win_prob": ep,
                        "power_payout": pout["power"],
                        "flex_payout": pout["flex"],
                        "n_legs": n_legs,
                    }
                )

        if len(eligible) > n_legs:
            eligible = eligible.iloc[1:].reset_index(drop=True)
        else:
            break

    tickets.sort(key=lambda x: (-x["avg_rank_score"], -x["avg_hit_rate"]))
    return tickets[:max_tickets]


# ──────────────────────────────────────────────────────────────────────────────
# FINAL web groups (ONLY the ticket sets you want) + ENFORCED Std/Gob mix
# ──────────────────────────────────────────────────────────────────────────────
def build_mixed_picktype_tickets(pool_df: pd.DataFrame, n_legs: int, max_tickets: int, min_standard: int) -> list:
    """
    Deterministic ticket builder that enforces a minimum number of Standard legs,
    while allowing remaining legs from Standard+Goblin pool.

    - Avoids duplicate players
    - Uses rank_score descending
    - Generates variety by sliding a start offset window
    """
    pool_df = pool_df.copy()
    if "rank_score" not in pool_df.columns or "pick_type" not in pool_df.columns:
        return []

    std = pool_df[pool_df["pick_type"] == "Standard"].sort_values("rank_score", ascending=False, na_position="last")
    gob = pool_df[pool_df["pick_type"] == "Goblin"].sort_values("rank_score", ascending=False, na_position="last")

    if len(std) < min_standard:
        return []

    tickets = []
    std_start = 0
    gob_start = 0
    attempts = 0
    max_attempts = max_tickets * 50

    while len(tickets) < max_tickets and attempts < max_attempts:
        attempts += 1
        legs = []
        used_players = set()

        # 1) Required Standards first
        for _, r in std.iloc[std_start:].iterrows():
            if sum(1 for x in legs if str(x.get("pick_type", "")) == "Standard") >= min_standard:
                break
            p = str(r.get("player", "")).strip().lower()
            if p and p not in used_players:
                legs.append(r)
                used_players.add(p)

        # 2) Fill remaining legs by best rank_score from (gob slice + std slice)
        combined_ranked = pd.concat([gob.iloc[gob_start:], std.iloc[std_start:]], ignore_index=True)
        combined_ranked = combined_ranked.sort_values("rank_score", ascending=False, na_position="last")

        for _, r in combined_ranked.iterrows():
            if len(legs) >= n_legs:
                break
            p = str(r.get("player", "")).strip().lower()
            if p and p not in used_players:
                legs.append(r)
                used_players.add(p)

        if len(legs) == n_legs:
            std_count = sum(1 for x in legs if str(x.get("pick_type", "")) == "Standard")
            if std_count >= min_standard:
                hrs = [float(x.get("hit_rate", 0.5) or 0.5) for x in legs]
                rss = [float(x.get("rank_score", 0) or 0) for x in legs]
                avg_hr = float(np.mean(hrs)) if hrs else 0.0
                avg_rs = float(np.mean(rss)) if rss else 0.0
                ep = win_prob(hrs, n_legs)
                pout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})

                key = frozenset((str(x.get("player", "")) + "|" + str(x.get("prop_type", ""))).strip() for x in legs)
                if key not in [t["key"] for t in tickets]:
                    tickets.append(
                        {
                            "key": key,
                            "rows": legs,
                            "avg_hit_rate": avg_hr,
                            "avg_rank_score": avg_rs,
                            "est_win_prob": ep,
                            "power_payout": pout["power"],
                            "flex_payout": pout["flex"],
                            "n_legs": n_legs,
                        }
                    )

        # Slide window to create different combos
        if len(std) > 0:
            std_start = min(std_start + 1, max(len(std) - 1, 0))
        if len(gob) > 0:
            gob_start = min(gob_start + 1, max(len(gob) - 1, 0))

    tickets.sort(key=lambda x: (-x["avg_rank_score"], -x["avg_hit_rate"]))
    return tickets[:max_tickets]


def build_final_web_ticket_groups(nba_pool: pd.DataFrame):
    mix_pool = nba_pool[nba_pool["pick_type"].isin(["Standard", "Goblin"])].copy()
    std_pool = nba_pool[nba_pool["pick_type"].isin(["Standard"])].copy()

    groups = []

    # Enforced mix rules:
    # 6-leg: >=2 Standard
    # 5-leg: >=2 Standard
    # 3-leg mix: >=1 Standard
    if len(mix_pool) >= 6:
        t6 = build_mixed_picktype_tickets(mix_pool, 6, max_tickets=1, min_standard=2)
        if t6:
            groups.append(("FINAL 6-Leg (NBA Std+Gob)", t6, None))

    if len(mix_pool) >= 5:
        t5 = build_mixed_picktype_tickets(mix_pool, 5, max_tickets=1, min_standard=2)
        if t5:
            groups.append(("FINAL 5-Leg (NBA Std+Gob)", t5, None))

    if len(mix_pool) >= 3:
        t3 = build_mixed_picktype_tickets(mix_pool, 3, max_tickets=2, min_standard=1)
        if t3:
            groups.append(("FINAL 3-Leg MIX x2 (NBA Std+Gob)", t3, None))

    if len(std_pool) >= 3:
        groups.append(("FINAL 3-Leg STANDARD ONLY (NBA)", build_tickets(std_pool, 3, max_tickets=1), None))

    return groups


# ── Write slate sheet ──────────────────────────────────────────────────────────
SLATE_COLS = [
    "sport",
    "tier",
    "rank_score",
    "player",
    "team",
    "opp",
    "prop_type",
    "pick_type",
    "line",
    "direction",
    "edge",
    "projection",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "def_tier",
    "game_time",
]
SLATE_WIDTHS = [6, 5, 10, 20, 6, 6, 18, 10, 6, 8, 7, 10, 10, 8, 10, 7, 7, 10, 16]
SLATE_HDRS = [
    "Sport",
    "Tier",
    "Rank Score",
    "Player",
    "Team",
    "Opp",
    "Prop",
    "Pick Type",
    "Line",
    "Dir",
    "Edge",
    "Proj",
    "Hit Rate",
    "L5 Avg",
    "Szn Avg",
    "L5 Over",
    "L5 Under",
    "Def Tier",
    "Game Time",
]


def write_slate_sheet(wb, df, sheet_name, bg_hdr, sport_label=""):
    ws = wb.create_sheet(sheet_name)
    cols = [c for c in SLATE_COLS if c in df.columns]
    hdrs = [SLATE_HDRS[SLATE_COLS.index(c)] for c in cols]
    widths = [SLATE_WIDTHS[SLATE_COLS.index(c)] for c in cols]
    sw(ws, widths)
    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(hdrs, 1):
        hc(ws, 1, ci, h, bg=bg_hdr)
    ws.freeze_panes = "A2"

    for ri, row in enumerate(df[cols].itertuples(index=False), 2):
        bg = C["alt"] if ri % 2 == 0 else C["white"]
        sp = getattr(row, "sport", "")
        if sp == "NBA":
            bg_row = C["nba"] if ri % 2 == 0 else C["white"]
        elif sp == "CBB":
            bg_row = C["cbb"] if ri % 2 == 0 else C["white"]
        else:
            bg_row = bg

        for ci, col in enumerate(cols, 1):
            val = getattr(row, col, "")
            if val is None or (isinstance(val, float) and np.isnan(val)):
                val = ""
            if col == "tier":
                dc(ws, ri, ci, val, bg=tier_bg(val), bold=True, align="center")
            elif col == "pick_type":
                dc(ws, ri, ci, val, bg=pt_bg(val), align="center")
            elif col == "hit_rate":
                pct_cell(ws, ri, ci, val if val != "" else np.nan)
                continue
            elif col == "rank_score":
                dc(ws, ri, ci, round(val, 2) if val != "" else "", bg=bg_row, bold=True, fmt="0.00")
            elif col == "direction":
                dbg = C["over"] if str(val).upper() == "OVER" else C["under"]
                dc(ws, ri, ci, val, bg=dbg, bold=True)
            elif col == "sport":
                sbg = C["hdr_nba"] if val == "NBA" else C["hdr_cbb"]
                dc(ws, ri, ci, val, bg=sbg, bold=True, fc="FFFFFF")
            elif col == "player":
                dc(ws, ri, ci, val, bg=bg_row, align="left", bold=True)
            elif col == "game_time":
                try:
                    if val and val != "":
                        dt = pd.to_datetime(val)
                        dc(ws, ri, ci, dt.strftime("%m/%d %I:%M%p"), bg=bg_row, align="center")
                    else:
                        dc(ws, ri, ci, "", bg=bg_row)
                except Exception:
                    dc(ws, ri, ci, str(val)[:16], bg=bg_row)
                continue
            else:
                dc(ws, ri, ci, val, bg=bg_row, align="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── Write ticket sheet ─────────────────────────────────────────────────────────
TICKET_COLS = [
    "#",
    "player",
    "team",
    "opp",
    "prop_type",
    "pick_type",
    "line",
    "direction",
    "edge",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "rank_score",
    "def_tier",
    "sport",
]
TICKET_HDRS = [
    "#",
    "Player",
    "Team",
    "Opp",
    "Prop",
    "Pick Type",
    "Line",
    "Dir",
    "Edge",
    "Hit Rate",
    "L5 Avg",
    "Szn Avg",
    "L5 Over",
    "L5 Under",
    "Rank Score",
    "Def Tier",
    "Sport",
]
TICKET_W = [4, 20, 6, 6, 18, 10, 6, 6, 7, 9, 8, 9, 7, 8, 11, 10, 6]


def write_ticket_sheet(wb, tickets, sheet_name, bg_hdr, label=""):
    if not tickets:
        return
    ws = wb.create_sheet(sheet_name)
    sw(ws, TICKET_W)
    ws.freeze_panes = "A2"

    ri = 1
    for ti, ticket in enumerate(tickets, 1):
        n = ticket["n_legs"]
        pout = ticket["power_payout"]
        fout = ticket["flex_payout"]
        cost = round(100 / pout, 0) if pout else 0
        avg_hr = ticket["avg_hit_rate"]
        ep = ticket["est_win_prob"]
        avg_rs = ticket["avg_rank_score"]

        banner = (
            f"  Ticket #{ti}  ·  {n}-Leg {label}  ·  "
            f"Power: {pout}x (${cost:.0f} to win $100)  ·  Flex: {fout}x  ·  "
            f"Avg Hit Rate: {avg_hr:.0%}  ·  Est Win Prob: {ep:.0%}  ·  "
            f"Avg Rank Score: {avg_rs:.2f}"
        )
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(TICKET_COLS))
        hc(ws, ri, 1, banner, bg=bg_hdr, sz=9, align="left")
        ws.row_dimensions[ri].height = 16
        ri += 1

        for ci, h in enumerate(TICKET_HDRS, 1):
            hc(ws, ri, ci, h, bg=C["hdr"], sz=8)
        ws.row_dimensions[ri].height = 14
        ri += 1

        for leg_i, row in enumerate(ticket["rows"], 1):
            bg = C["alt"] if leg_i % 2 == 0 else C["white"]
            sp = row.get("sport", "")
            if sp == "NBA":
                bg = C["nba"]
            elif sp == "CBB":
                bg = C["cbb"]

            def gv(field):
                return row.get(field, "")

            dc(ws, ri, 1, leg_i, bg=bg, bold=True, align="center")
            dc(ws, ri, 2, gv("player"), bg=bg, align="left", bold=True)
            dc(ws, ri, 3, gv("team"), bg=bg)
            dc(ws, ri, 4, gv("opp"), bg=bg)
            dc(ws, ri, 5, gv("prop_type"), bg=bg, align="left")
            ptv = gv("pick_type")
            dc(ws, ri, 6, ptv, bg=pt_bg(str(ptv)), align="center")
            dc(ws, ri, 7, gv("line"), bg=bg)
            dirv = str(gv("direction")).upper()
            dc(ws, ri, 8, dirv, bg=C["over"] if dirv == "OVER" else C["under"], bold=True)
            dc(ws, ri, 9, gv("edge"), bg=bg)
            pct_cell(ws, ri, 10, gv("hit_rate") if gv("hit_rate") != "" else np.nan)
            dc(ws, ri, 11, gv("l5_avg"), bg=bg)
            dc(ws, ri, 12, gv("season_avg"), bg=bg)
            dc(ws, ri, 13, gv("l5_over"), bg=bg)
            dc(ws, ri, 14, gv("l5_under"), bg=bg)
            rs = gv("rank_score")
            try:
                rs_out = round(float(rs), 2) if rs != "" and rs is not None else ""
            except Exception:
                rs_out = ""
            dc(ws, ri, 15, rs_out, bg=bg, bold=True)
            dc(ws, ri, 16, gv("def_tier"), bg=bg)
            sv = gv("sport")
            sbg = C["hdr_nba"] if sv == "NBA" else (C["hdr_cbb"] if sv == "CBB" else C["hdr"])
            dc(ws, ri, 17, sv, bg=sbg, bold=True, fc="FFFFFF")
            ws.row_dimensions[ri].height = 14
            ri += 1

        ws.row_dimensions[ri].height = 6
        ri += 1


# ── Write SUMMARY sheet ───────────────────────────────────────────────────────
def write_summary(wb, nba, cbb, combined, all_ticket_groups, date_str, thresholds):
    ws = wb.create_sheet("SUMMARY", 0)
    sw(ws, [28, 14, 10, 10, 10, 10, 10, 12, 18])

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"COMBINED NBA + CBB SLATE  |  {date_str}  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c2 = ws["A2"]
    c2.value = (
        f"Filters: Tier {thresholds.get('tiers','ALL')} | "
        f"Min Hit Rate: {thresholds.get('min_hit_rate',0):.0%} | "
        f"Min Edge: {thresholds.get('min_edge',0)} | "
        f"Min Rank Score: {thresholds.get('min_rank','None')} | "
        f"Pick Types: {thresholds.get('pick_types','ALL')}"
    )
    c2.font = Font(bold=False, name="Arial", size=9, color="000000")
    c2.fill = PatternFill("solid", start_color=C["gold"])
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4

    def sec(r, label, bg):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        hc(ws, r, 1, label, bg=bg, sz=10, align="left")
        ws.row_dimensions[r].height = 20
        return r + 1

    def stat_row(r, label, total, elig, bg=None):
        bg = bg or (C["alt"] if r % 2 == 0 else C["white"])
        dc(ws, r, 1, label, bg=bg, align="left", bold=True)
        dc(ws, r, 2, total, bg=bg)
        dc(ws, r, 3, elig, bg=bg)
        for ci in range(4, 10):
            dc(ws, r, ci, "", bg=bg)
        return r + 1

    row = sec(row, "📊 SLATE OVERVIEW", C["hdr_sum"])
    for ci, h in enumerate(["Category", "Total Props", "Eligible", "", "", "", "", "", ""], 1):
        hc(ws, row, ci, h, bg=C["hdr"], sz=8)
    ws.row_dimensions[row].height = 14
    row += 1

    elig_nba = len(nba[nba.get("tier", "").isin(["A", "B"])]) if "tier" in nba.columns else 0
    elig_cbb = len(cbb[cbb.get("tier", "").isin(["A", "B"])]) if "tier" in cbb.columns else 0
    elig_all = len(combined[combined.get("tier", "").isin(["A", "B"])]) if "tier" in combined.columns else 0
    row = stat_row(row, "NBA Props", len(nba), elig_nba, C["nba"])
    row = stat_row(row, "CBB Props", len(cbb), elig_cbb, C["cbb"])
    row = stat_row(row, "Combined Slate", len(combined), elig_all)
    row += 1

    row = sec(row, "🎟️ TICKET SUMMARY", C["hdr_mix"])
    for ci, h in enumerate(
        ["Sheet", "Legs", "Type", "# Tickets", "Avg Hit Rate", "Avg Win Prob", "Avg Rank Score", "Power Payout", "Players"],
        1,
    ):
        hc(ws, row, ci, h, bg=C["hdr"], sz=8)
    ws.row_dimensions[row].height = 14
    row += 1

    for group_name, tickets, bg_row in all_ticket_groups:
        if not tickets:
            continue
        avg_hr = np.mean([t["avg_hit_rate"] for t in tickets])
        avg_wp = np.mean([t["est_win_prob"] for t in tickets])
        avg_rs = np.mean([t["avg_rank_score"] for t in tickets])
        n = tickets[0]["n_legs"]
        pout = tickets[0]["power_payout"]
        bg = bg_row if bg_row else (C["alt"] if row % 2 == 0 else C["white"])
        dc(ws, row, 1, group_name, bg=bg, align="left", bold=True)
        dc(ws, row, 2, n, bg=bg)
        lbl = group_name.split(" ")[0] if group_name else ""
        dc(ws, row, 3, lbl, bg=bg)
        dc(ws, row, 4, len(tickets), bg=bg)
        pct_cell(ws, row, 5, avg_hr)
        pct_cell(ws, row, 6, avg_wp)
        dc(ws, row, 7, round(avg_rs, 2), bg=bg)
        dc(ws, row, 8, f"{pout}x", bg=bg)
        sample = " | ".join(f"{r.get('player','')}" for r in tickets[0]["rows"][:3]) + ("..." if n > 3 else "")
        dc(ws, row, 9, sample, bg=bg, align="left", sz=8)
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--nba", required=True, help="NBA step8_all_direction_clean.xlsx")
    ap.add_argument("--cbb", required=True, help="CBB step6_ranked_cbb.xlsx")
    ap.add_argument("--output", default="")
    ap.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    ap.add_argument("--tiers", default="A,B", help="Comma-separated tiers e.g. A,B")
    ap.add_argument("--min-hit-rate", type=float, default=0.0, dest="min_hit_rate")
    ap.add_argument("--min-edge", type=float, default=0.0, dest="min_edge")
    ap.add_argument("--min-rank", type=float, default=None, dest="min_rank")
    ap.add_argument("--pick-types", default="Goblin,Standard,Demon", dest="pick_types")
    ap.add_argument("--max-tickets", type=int, default=20, dest="max_tickets")

    # Web outputs
    ap.add_argument("--write-web", action="store_true", help="Write tickets_latest.html/json for GitHub Pages")
    ap.add_argument("--web-outdir", default="docs", help="Folder to write tickets_latest.html/json (default: docs)")
    ap.add_argument("--also-root", action="store_true", help="Also write tickets_latest.* in repo root")

    args = ap.parse_args()

    if not args.output:
        args.output = f"combined_slate_tickets_{args.date}.xlsx"

    tiers = [t.strip() for t in args.tiers.split(",") if t.strip()]
    pick_types = [p.strip() for p in args.pick_types.split(",") if p.strip()]
    thresholds = {
        "tiers": args.tiers,
        "min_hit_rate": args.min_hit_rate,
        "min_edge": args.min_edge,
        "min_rank": args.min_rank,
        "pick_types": args.pick_types,
    }

    print(f"Loading NBA slate from {args.nba}...")
    nba = load_nba(args.nba)
    print(f"  {len(nba)} NBA props loaded")

    print(f"Loading CBB slate from {args.cbb}...")
    cbb = load_cbb(args.cbb)
    print(f"  {len(cbb)} CBB props loaded")

    print("Building combined slate...")
    combined = build_combined_slate(nba, cbb)
    print(f"  {len(combined)} total props")

    def pool(df, pt=None):
        return filter_eligible(
            df,
            args.min_hit_rate,
            args.min_edge,
            args.min_rank,
            tiers if tiers else None,
            pt if pt is not None else pick_types,
        )

    nba_pool = pool(nba)
    cbb_pool = pool(cbb)
    combo_pool = pool(combined)
    print(f"  NBA eligible: {len(nba_pool)} | CBB eligible: {len(cbb_pool)} | Combined: {len(combo_pool)}")

    print("Generating tickets + workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    all_ticket_groups = []
    leg_sizes = [3, 4, 5, 6]

    def gen_tickets(pool_df, sport_label, bg_hdr, sport_prefix, pick_type_filter=None):
        rows_out = []
        for n in leg_sizes:
            sub_pool = pool_df if pick_type_filter is None else pool_df[pool_df["pick_type"].isin([pick_type_filter])]
            tickets = build_tickets(sub_pool, n, args.max_tickets)
            if tickets:
                pt_label = pick_type_filter or "Mix"
                sheet_name = f"{sport_prefix} {pt_label} {n}-Leg"[:31]
                write_ticket_sheet(wb, tickets, sheet_name, bg_hdr, label=f"{sport_label} {pt_label}")
                rows_out.append((sheet_name, tickets, None))
                print(f"  {sheet_name}: {len(tickets)} tickets")
        return rows_out

    # NBA tickets by pick type
    for pt in ["Goblin", "Standard", "Demon"]:
        pt_pool = pool(nba, [pt])
        if len(pt_pool) >= 3:
            all_ticket_groups += gen_tickets(pt_pool, "NBA", C["hdr_nba"], "NBA", pt)

    # NBA Mix
    if len(nba_pool) >= 3:
        all_ticket_groups += gen_tickets(nba_pool, "NBA", C["hdr_nba"], "NBA Mix")

    # CBB tickets by pick type
    for pt in ["Goblin", "Standard", "Demon"]:
        pt_pool = pool(cbb, [pt])
        if len(pt_pool) >= 3:
            all_ticket_groups += gen_tickets(pt_pool, "CBB", C["hdr_cbb"], "CBB", pt)

    # CBB Mix
    if len(cbb_pool) >= 3:
        all_ticket_groups += gen_tickets(cbb_pool, "CBB", C["hdr_cbb"], "CBB Mix")

    # Combined NBA+CBB tickets (all pick types mixed)
    if len(combo_pool) >= 3:
        all_ticket_groups += gen_tickets(combo_pool, "COMBO", C["hdr_mix"], "COMBO")

    # Cross-sport Standard Mix (enforce mix)
    nba_std = pool(nba, ["Standard"])
    cbb_std = pool(cbb, ["Standard"])
    std_mix_pool = pd.concat([nba_std, cbb_std], ignore_index=True).sort_values("rank_score", ascending=False)
    if len(std_mix_pool) >= 3:
        print("Generating cross-sport Standard Mix tickets...")
        for n in leg_sizes:
            tickets = build_tickets(std_mix_pool, n, args.max_tickets, require_mix=True)
            if tickets:
                sheet_name = f"MIX Standard {n}-Leg"[:31]
                write_ticket_sheet(wb, tickets, sheet_name, C["hdr_mix"], label="NBA+CBB Standard")
                all_ticket_groups.append((sheet_name, tickets, C["mix"]))
                print(f"  {sheet_name}: {len(tickets)} tickets")

    # Cross-sport Goblin Mix (enforce mix)
    nba_gob = pool(nba, ["Goblin"])
    cbb_gob = pool(cbb, ["Goblin"])
    gob_mix_pool = pd.concat([nba_gob, cbb_gob], ignore_index=True).sort_values("rank_score", ascending=False)
    if len(gob_mix_pool) >= 3:
        print("Generating cross-sport Goblin Mix tickets...")
        for n in leg_sizes:
            tickets = build_tickets(gob_mix_pool, n, args.max_tickets, require_mix=True)
            if tickets:
                sheet_name = f"MIX Goblin {n}-Leg"[:31]
                write_ticket_sheet(wb, tickets, sheet_name, C["goblin"], label="NBA+CBB Goblin")
                all_ticket_groups.append((sheet_name, tickets, C["goblin"]))
                print(f"  {sheet_name}: {len(tickets)} tickets")

    print("Writing slate sheets...")
    write_slate_sheet(wb, combined, "Full Slate", C["hdr"], "ALL")
    write_slate_sheet(wb, nba, "NBA Slate", C["hdr_nba"], "NBA")
    write_slate_sheet(wb, cbb, "CBB Slate", C["hdr_cbb"], "CBB")

    write_summary(wb, nba, cbb, combined, all_ticket_groups, args.date, thresholds)

    # Reorder
    summary_sheet = wb["SUMMARY"]
    wb.move_sheet(summary_sheet, offset=-len(wb.sheetnames))
    for sname in ["Full Slate", "NBA Slate", "CBB Slate"]:
        if sname in wb.sheetnames:
            wb.move_sheet(wb[sname], offset=-(len(wb.sheetnames) - 1))

    wb.save(args.output)
    print(f"\n✅ Saved -> {args.output}")
    print(f"   Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Web output (FINAL only)
    if args.write_web:
        print("\nWriting GitHub Pages web outputs (FINAL tickets only)...")
        final_groups = build_final_web_ticket_groups(nba_pool)
        payload = ticket_groups_to_payload(final_groups, args.date, thresholds)
        write_web_outputs(payload, args.web_outdir)
        if args.also_root:
            write_web_outputs(payload, outdir=".")
        print("✅ Web outputs complete (FINAL only).")


if __name__ == "__main__":
    main()