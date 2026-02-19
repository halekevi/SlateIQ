#!/usr/bin/env python3
"""
pull_and_grade_last_night.py

Grades props using stats.nba.com via nba_api only:
- scoreboardv2 -> list GAME_IDs for a date
- boxscoretraditionalv3 -> player traditional stats (includes FGA/FGM/FG3A/FG3M/FTA/FTM/etc.)

Supports:
- --ranked <ranked_csv_or_xlsx>
- --top with --overs <overs_csv_or_xlsx> --unders <unders_csv_or_xlsx>
- BOTH in one run (no mutual exclusivity)
- --all convenience (grades whatever inputs you pass)

Outputs:
- graded_ranked_<YYYY-MM-DD>_graded.csv and .xlsx (if --ranked used)
- graded_top_<YYYY-MM-DD>_graded.csv and .xlsx (if --top used)

FIXES:
- Grades shooting props: 3PTA/3PTM, FGA/FGM, FTA/FTM, 2PA/2PM (derived)
- Supports "Blks+Stls" / "Blocked Shots" naming variants
- Removes "(Combo)" so those props map correctly
- Handles Goblin/Demon OVER-only constraint fairly
- Avoids duplicate-column collisions by prefixing all graded outputs with `graded_`

NEW:
- Writes the same multi-tab workbook layout you showed:
  Box Raw, By Tier, By Prop Type, By Bet Dir, By Pick Type, By Abs Edge, By Dir OK,
  By Allowed Dir, By Platform OK, By Minutes, By Pick+Minutes, By Dir+Minutes,
  By Opp Def, By Min+Def, Void Reasons
"""

from __future__ import annotations

import argparse
import os
import random
import re
import time
import unicodedata
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, List

import numpy as np
import pandas as pd

from nba_api.stats.endpoints import scoreboardv2, boxscoretraditionalv3
from nba_api.stats.library.parameters import LeagueID

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# -----------------------------
# Prop normalization + mapping
# -----------------------------

def _norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = s.encode("ascii", "ignore").decode("ascii")
    s = s.strip().lower()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\(combo\)", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_prop_type(x: str) -> str:
    s = _norm_text(x)
    s = s.replace("score", "").strip()
    s = s.replace(" + ", "+").replace(" ", "")

    if s in ("pts+rebs+asts", "pra"):
        return "pts+rebs+asts"
    if s in ("pts+rebs", "points+rebounds"):
        return "pts+rebs"
    if s in ("pts+asts", "points+assists"):
        return "pts+asts"
    if s in ("rebs+asts", "rebounds+assists"):
        return "rebs+asts"

    s2 = _norm_text(x).replace("score", "").strip()

    aliases = {
        "points": "points",
        "pts": "points",
        "rebounds": "rebounds",
        "rebs": "rebounds",
        "assists": "assists",
        "asts": "assists",
        "steals": "steals",
        "stl": "steals",
        "blocked shots": "blocked shots",
        "blocks": "blocked shots",
        "blk": "blocked shots",
        "turnovers": "turnovers",
        "tov": "turnovers",
        "fantasy": "fantasy score",
        "fantasy score": "fantasy score",

        "blks+stls": "blks+stls",
        "blocks+steals": "blks+stls",
        "stls+blks": "blks+stls",

        "3-pt attempted": "3-pt attempted",
        "3pt attempted": "3-pt attempted",
        "3 point attempted": "3-pt attempted",
        "3 pointers attempted": "3-pt attempted",
        "3-point attempted": "3-pt attempted",

        "3-pt made": "3-pt made",
        "3pt made": "3-pt made",
        "3 point made": "3-pt made",
        "3 pointers made": "3-pt made",
        "3-point made": "3-pt made",

        "fg attempted": "fg attempted",
        "field goals attempted": "fg attempted",
        "field goal attempted": "fg attempted",

        "fg made": "fg made",
        "field goals made": "fg made",
        "field goal made": "fg made",

        "free throws attempted": "free throws attempted",
        "ft attempted": "free throws attempted",

        "free throws made": "free throws made",
        "ft made": "free throws made",

        "two pointers attempted": "two pointers attempted",
        "2pt attempted": "two pointers attempted",
        "2-pt attempted": "two pointers attempted",

        "two pointers made": "two pointers made",
        "2pt made": "two pointers made",
        "2-pt made": "two pointers made",

        # not available in boxscoretraditionalv3
        "personal fouls": "personal fouls",
        "dunks": "dunks",
    }
    return aliases.get(s2, s2)


PROP_TO_PREFIX = {
    "points": "pts",
    "rebounds": "reb",
    "assists": "ast",
    "steals": "stl",
    "blocked shots": "blk",
    "turnovers": "tov",
    "fantasy score": "fantasy",
    "blks+stls": "stocks",

    "pts+rebs": "pr",
    "pts+asts": "pa",
    "rebs+asts": "ra",
    "pts+rebs+asts": "pra",

    "3-pt attempted": "fg3a",
    "3-pt made": "fg3m",
    "fg attempted": "fga",
    "fg made": "fgm",
    "free throws attempted": "fta",
    "free throws made": "ftm",
    "two pointers attempted": "fg2a",
    "two pointers made": "fg2m",

    "personal fouls": "pf",
    "dunks": "dunks",
}

COMBO_PARTS = {
    "pr": ["pts", "reb"],
    "pa": ["pts", "ast"],
    "ra": ["reb", "ast"],
    "pra": ["pts", "reb", "ast"],
    "stocks": ["stl", "blk"],
}


# -----------------------------
# NBA API helpers
# -----------------------------

def _safe_sleep(min_s: float = 0.35, max_s: float = 0.85) -> None:
    time.sleep(random.uniform(min_s, max_s))


def get_game_ids(date_str: str) -> List[str]:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    game_date = dt.strftime("%m/%d/%Y")
    sb = scoreboardv2.ScoreboardV2(game_date=game_date, league_id=LeagueID.default)
    df = sb.game_header.get_data_frame()
    if df.empty:
        return []
    return df["GAME_ID"].astype(str).tolist()


def fetch_boxscore_player_stats(game_id: str) -> pd.DataFrame:
    bs = boxscoretraditionalv3.BoxScoreTraditionalV3(game_id=game_id)
    data = bs.get_dict()

    players = []
    try:
        home = data["boxScoreTraditional"]["homeTeam"]["players"]
        away = data["boxScoreTraditional"]["awayTeam"]["players"]
        players = home + away
    except Exception:
        players = []

    if not players:
        return pd.DataFrame()

    rows = []
    for p in players:
        stats = p.get("statistics") or {}
        rows.append({
            "player_id": p.get("personId"),
            "player_name": f"{p.get('firstName','').strip()} {p.get('familyName','').strip()}".strip(),
            "pts": stats.get("points"),
            "reb": stats.get("reboundsTotal"),
            "ast": stats.get("assists"),
            "stl": stats.get("steals"),
            "blk": stats.get("blocks"),
            "tov": stats.get("turnovers"),
            "fgm": stats.get("fieldGoalsMade"),
            "fga": stats.get("fieldGoalsAttempted"),
            "fg3m": stats.get("threePointersMade"),
            "fg3a": stats.get("threePointersAttempted"),
            "ftm": stats.get("freeThrowsMade"),
            "fta": stats.get("freeThrowsAttempted"),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    for c in ["player_id", "pts", "reb", "ast", "stl", "blk", "tov", "fgm", "fga", "fg3m", "fg3a", "ftm", "fta"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    return df


def build_player_actuals_for_date(date_str: str) -> Tuple[Dict[int, Dict[str, float]], Dict[str, int]]:
    game_ids = get_game_ids(date_str)
    actuals_by_id: Dict[int, Dict[str, float]] = {}
    name_to_id: Dict[str, int] = {}

    for gid in game_ids:
        _safe_sleep()
        gdf = fetch_boxscore_player_stats(gid)
        if gdf.empty:
            continue

        for _, r in gdf.iterrows():
            pid = r.get("player_id")
            if pd.isna(pid):
                continue
            pid = int(pid)

            base = {
                "pts": float(r.get("pts") or 0),
                "reb": float(r.get("reb") or 0),
                "ast": float(r.get("ast") or 0),
                "stl": float(r.get("stl") or 0),
                "blk": float(r.get("blk") or 0),
                "tov": float(r.get("tov") or 0),
                "fgm": float(r.get("fgm") or 0),
                "fga": float(r.get("fga") or 0),
                "fg3m": float(r.get("fg3m") or 0),
                "fg3a": float(r.get("fg3a") or 0),
                "ftm": float(r.get("ftm") or 0),
                "fta": float(r.get("fta") or 0),
            }

            base["fg2a"] = base["fga"] - base["fg3a"]
            base["fg2m"] = base["fgm"] - base["fg3m"]
            base["stocks"] = base["stl"] + base["blk"]
            base["pr"] = base["pts"] + base["reb"]
            base["pa"] = base["pts"] + base["ast"]
            base["ra"] = base["reb"] + base["ast"]
            base["pra"] = base["pts"] + base["reb"] + base["ast"]

            base["fantasy"] = (
                base["pts"] + 1.2 * base["reb"] + 1.5 * base["ast"] + 3 * base["stl"] + 3 * base["blk"] - base["tov"]
            )

            actuals_by_id[pid] = base

            nm = _norm_text(r.get("player_name") or "")
            if nm and nm not in name_to_id:
                name_to_id[nm] = pid

    return actuals_by_id, name_to_id


# -----------------------------
# IO helpers
# -----------------------------

def read_any(path: str) -> pd.DataFrame:
    if path.lower().endswith(".xlsx"):
        return pd.read_excel(path)
    return pd.read_csv(path)


def _ws_write_df(ws, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = max(12, min(34, len(str(col)) + 4))


def _first_existing(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _abs_edge_bucket(x) -> str:
    try:
        v = float(x)
    except Exception:
        return "NA"
    v = abs(v)
    if v >= 5:
        return "5+"
    if v >= 4:
        return "4-4.99"
    if v >= 3:
        return "3-3.99"
    if v >= 2:
        return "2-2.99"
    if v >= 1:
        return "1-1.99"
    return "0-0.99"


def _minutes_bucket(x) -> str:
    try:
        v = float(x)
    except Exception:
        return "NA"
    if v >= 36:
        return "36+"
    if v >= 32:
        return "32-35.9"
    if v >= 28:
        return "28-31.9"
    if v >= 24:
        return "24-27.9"
    if v >= 20:
        return "20-23.9"
    return "<20"


def _standard_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    g = df.copy()
    res = g["graded_result"].fillna("VOID")

    out = (
        g.assign(
            _is_hit=(res == "HIT").astype(int),
            _is_miss=(res == "MISS").astype(int),
            _is_void=(res == "VOID").astype(int),
            _decided=res.isin(["HIT", "MISS"]).astype(int),
        )
        .groupby(group_col, dropna=False)
        .agg(
            total=("graded_result", "size"),
            hit=("_is_hit", "sum"),
            miss=("_is_miss", "sum"),
            void=("_is_void", "sum"),
            decided=("_decided", "sum"),
        )
        .reset_index()
    )
    out["hit_rate"] = np.where(out["decided"] > 0, out["hit"] / out["decided"], np.nan)
    return out


def write_xlsx_with_tabs(df: pd.DataFrame, out_path: str) -> None:
    """
    Writes a multi-tab workbook like your screenshot:
    Box Raw, By Tier, By Prop Type, By Bet Dir, By Pick Type, By Abs Edge, By Dir OK,
    By Allowed Dir, By Platform OK, By Minutes, By Pick+Minutes, By Dir+Minutes,
    By Opp Def, By Min+Def, Void Reasons
    """
    wb = Workbook()
    wb.remove(wb.active)

    # --- derive flexible fields used in summaries ---
    out = df.copy()

    # Prefer graded fields for direction/constraints
    if "graded_bet_dir" in out.columns:
        out["bet_dir_used"] = out["graded_bet_dir"]
    else:
        c = _first_existing(out, ["bet_dir", "bet_direction", "direction"])
        out["bet_dir_used"] = out[c] if c else "NA"

    out["pick_type_used"] = out[_first_existing(out, ["pick_type", "picktype"])] if _first_existing(out, ["pick_type", "picktype"]) else "standard"

    # Tier (if you already have it)
    tier_col = _first_existing(out, ["tier", "Tier", "TIER"])
    if tier_col:
        out["tier_used"] = out[tier_col]
    else:
        out["tier_used"] = "NA"

    # Abs edge (prefer abs_edge; else edge)
    edge_col = _first_existing(out, ["abs_edge", "absEdge", "ABS_EDGE", "edge", "EDGE"])
    if edge_col:
        out["abs_edge_used"] = pd.to_numeric(out[edge_col], errors="coerce").abs()
    else:
        out["abs_edge_used"] = np.nan
    out["abs_edge_bucket"] = out["abs_edge_used"].apply(_abs_edge_bucket)

    # Minutes (prefer minutes_proj then min)
    min_col = _first_existing(out, ["minutes_proj", "minutes", "min", "MIN"])
    if min_col:
        out["minutes_used"] = pd.to_numeric(out[min_col], errors="coerce")
    else:
        out["minutes_used"] = np.nan
    out["minutes_bucket"] = out["minutes_used"].apply(_minutes_bucket)

    # Opp def bucket (your pipeline often has _opp_def_bucket)
    def_col = _first_existing(out, ["_opp_def_bucket", "opp_def_bucket", "OVERALL_DEF_RANK", "overall_def_rank"])
    if def_col == "OVERALL_DEF_RANK" or def_col == "overall_def_rank":
        # bucket numeric ranks if present
        ranks = pd.to_numeric(out[def_col], errors="coerce")
        def_bucket = np.where(ranks.between(1, 10), "1-10 Strong",
                     np.where(ranks.between(11, 20), "11-20 Average",
                     np.where(ranks.between(21, 30), "21-30 Weak", "NA")))
        out["opp_def_used"] = def_bucket
    elif def_col:
        out["opp_def_used"] = out[def_col]
    else:
        out["opp_def_used"] = "NA"

    # Combine buckets
    out["pick_plus_minutes"] = out["pick_type_used"].astype(str).str.upper() + " | " + out["minutes_bucket"].astype(str)
    out["dir_plus_minutes"] = out["bet_dir_used"].astype(str).str.upper() + " | " + out["minutes_bucket"].astype(str)
    out["min_plus_def"] = out["minutes_bucket"].astype(str) + " | " + out["opp_def_used"].astype(str)

    # --- Box Raw ---
    ws = wb.create_sheet("Box Raw")
    _ws_write_df(ws, out)

    # --- summaries (only if graded_result exists) ---
    if "graded_result" not in out.columns:
        # still save Box Raw
        wb.save(out_path)
        return

    def _add_summary(sheet: str, col: str) -> None:
        if col not in out.columns:
            # still create the sheet so your layout matches, but show a message row
            wsx = wb.create_sheet(sheet)
            wsx.append([f"Column '{col}' not found in data."])
            return
        wsx = wb.create_sheet(sheet)
        summ = _standard_summary(out, col)
        _ws_write_df(wsx, summ)

    _add_summary("By Tier", "tier_used")
    _add_summary("By Prop Type", "graded_prop_type_norm" if "graded_prop_type_norm" in out.columns else "prop_type_norm" if "prop_type_norm" in out.columns else "prop_type")
    _add_summary("By Bet Dir", "bet_dir_used")
    _add_summary("By Pick Type", "pick_type_used")
    _add_summary("By Abs Edge", "abs_edge_bucket")
    _add_summary("By Dir OK", "graded_dir_ok")
    _add_summary("By Allowed Dir", "graded_allowed_dir")
    _add_summary("By Platform OK", "graded_platform_ok")
    _add_summary("By Minutes", "minutes_bucket")
    _add_summary("By Pick+Minutes", "pick_plus_minutes")
    _add_summary("By Dir+Minutes", "dir_plus_minutes")
    _add_summary("By Opp Def", "opp_def_used")
    _add_summary("By Min+Def", "min_plus_def")
    _add_summary("Void Reasons", "graded_void_reason")

    wb.save(out_path)


# -----------------------------
# Grading
# -----------------------------

def infer_bet_dir(row: pd.Series) -> str:
    for c in ["bet_dir", "bet_direction", "direction"]:
        if c in row.index and pd.notna(row[c]):
            v = str(row[c]).strip().upper()
            if v in ("OVER", "UNDER"):
                return v
    if "edge" in row.index:
        try:
            return "OVER" if float(row["edge"]) >= 0 else "UNDER"
        except Exception:
            pass
    return "OVER"


def infer_pick_type(row: pd.Series) -> str:
    for c in ["pick_type", "picktype"]:
        if c in row.index and pd.notna(row[c]):
            return str(row[c]).strip().lower()
    return "standard"


def infer_player_id(row: pd.Series, name_to_id: Dict[str, int]) -> Optional[int]:
    for c in ["nba_player_id", "player_id", "nba_playerid"]:
        if c in row.index and pd.notna(row[c]):
            try:
                return int(float(row[c]))
            except Exception:
                pass
    for c in ["player", "name"]:
        if c in row.index and pd.notna(row[c]):
            nm = _norm_text(row[c])
            return name_to_id.get(nm)
    return None


def infer_prop_type(row: pd.Series) -> str:
    for c in ["prop_type", "prop_type_norm", "prop", "stat"]:
        if c in row.index and pd.notna(row[c]):
            return str(row[c])
    return ""


def infer_line(row: pd.Series) -> Optional[float]:
    for c in ["line", "prop_line", "value"]:
        if c in row.index and pd.notna(row[c]):
            try:
                return float(row[c])
            except Exception:
                return None
    return None


def compute_actual_for_prop(prefix: str, actuals: Dict[str, float]) -> Tuple[Optional[float], str]:
    if prefix in ("pf", "dunks"):
        return None, f"UNSUPPORTED_STAT_{prefix.upper()}"
    if prefix in actuals:
        return float(actuals[prefix]), ""
    if prefix in COMBO_PARTS:
        parts = COMBO_PARTS[prefix]
        if all(p in actuals for p in parts):
            return float(sum(actuals[p] for p in parts)), ""
    return None, "NO_ACTUAL_FOUND"


def grade_row(
    row: pd.Series,
    actuals_by_id: Dict[int, Dict[str, float]],
    name_to_id: Dict[str, int],
) -> Dict[str, object]:
    prop_raw = infer_prop_type(row)
    prop_norm = norm_prop_type(prop_raw)
    prefix = PROP_TO_PREFIX.get(prop_norm, "")

    line = infer_line(row)
    bet_dir_model = infer_bet_dir(row)
    pick_type = infer_pick_type(row)

    allowed_dir = "BOTH"
    forced_dir = bet_dir_model
    if pick_type in ("goblin", "demon"):
        allowed_dir = "OVER_ONLY"
        forced_dir = "OVER"

    pid = infer_player_id(row, name_to_id)

    platform_ok = int((allowed_dir == "BOTH") or (bet_dir_model == "OVER"))
    dir_ok = int((allowed_dir == "BOTH") or (bet_dir_model == "OVER"))

    if pid is None or pid not in actuals_by_id:
        return {
            "prop_type_norm": prop_norm,
            "stat_prefix": prefix,
            "allowed_dir": allowed_dir,
            "bet_dir": forced_dir,
            "platform_ok": platform_ok,
            "dir_ok": np.nan,
            "actual": np.nan,
            "result": "VOID",
            "void_reason": "PLAYER_NOT_FOUND",
            "result_sign": 0,
        }

    if line is None or pd.isna(line):
        return {
            "prop_type_norm": prop_norm,
            "stat_prefix": prefix,
            "allowed_dir": allowed_dir,
            "bet_dir": forced_dir,
            "platform_ok": platform_ok,
            "dir_ok": np.nan,
            "actual": np.nan,
            "result": "VOID",
            "void_reason": "MISSING_LINE",
            "result_sign": 0,
        }

    if not prefix:
        return {
            "prop_type_norm": prop_norm,
            "stat_prefix": "",
            "allowed_dir": allowed_dir,
            "bet_dir": forced_dir,
            "platform_ok": platform_ok,
            "dir_ok": np.nan,
            "actual": np.nan,
            "result": "VOID",
            "void_reason": "UNMAPPED_PROP_TYPE",
            "result_sign": 0,
        }

    actual, void_reason = compute_actual_for_prop(prefix, actuals_by_id[pid])

    if actual is None or pd.isna(actual):
        return {
            "prop_type_norm": prop_norm,
            "stat_prefix": prefix,
            "allowed_dir": allowed_dir,
            "bet_dir": forced_dir,
            "platform_ok": platform_ok,
            "dir_ok": dir_ok,
            "actual": np.nan,
            "result": "VOID",
            "void_reason": void_reason,
            "result_sign": 0,
        }

    if forced_dir == "OVER":
        if actual > line:
            res, sign = "HIT", 1
        elif actual < line:
            res, sign = "MISS", -1
        else:
            res, sign = "VOID", 0
    else:
        if actual < line:
            res, sign = "HIT", 1
        elif actual > line:
            res, sign = "MISS", -1
        else:
            res, sign = "VOID", 0

    void_reason_out = "" if res != "VOID" else "PUSH"

    return {
        "prop_type_norm": prop_norm,
        "stat_prefix": prefix,
        "allowed_dir": allowed_dir,
        "bet_dir": forced_dir,
        "platform_ok": platform_ok,
        "dir_ok": dir_ok,
        "actual": float(actual),
        "result": res,
        "void_reason": void_reason_out,
        "result_sign": sign,
    }


def grade_df(df: pd.DataFrame, date_str: str, actuals_by_id: Dict[int, Dict[str, float]], name_to_id: Dict[str, int]) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]

    graded_rows = []
    for _, row in df.iterrows():
        graded_rows.append(grade_row(row, actuals_by_id, name_to_id))

    g = pd.DataFrame(graded_rows).add_prefix("graded_")
    out = pd.concat([df.reset_index(drop=True), g.reset_index(drop=True)], axis=1)

    res = out["graded_result"]
    out["decided"] = np.where(res.isin(["HIT", "MISS"]), 1, 0)
    out["hit"] = np.where(res == "HIT", 1, 0)
    out["miss"] = np.where(res == "MISS", 1, 0)
    out["void"] = np.where(res == "VOID", 1, 0)
    out["graded_date"] = date_str
    return out


# -----------------------------
# Main
# -----------------------------

def resolve_date(s: Optional[str]) -> str:
    if s:
        return s
    return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default=None, help="YYYY-MM-DD (default: yesterday)")
    ap.add_argument("--ranked", default=None, help="ranked csv/xlsx to grade")
    ap.add_argument("--top", action="store_true", help="grade overs/unders files too")
    ap.add_argument("--overs", default=None, help="overs csv/xlsx (used with --top)")
    ap.add_argument("--unders", default=None, help="unders csv/xlsx (used with --top)")
    ap.add_argument("--outdir", default=".", help="output directory")
    ap.add_argument("--all", action="store_true", help="convenience flag; grades whatever inputs are provided")
    args = ap.parse_args()

    date_str = resolve_date(args.date)
    os.makedirs(args.outdir, exist_ok=True)

    print(f"📅 Grading date: {date_str}")
    actuals_by_id, name_to_id = build_player_actuals_for_date(date_str)
    if not actuals_by_id:
        print("⚠️ No game/player actuals found for this date (no games or API issue).")

    # Ranked
    if args.ranked:
        df_ranked = read_any(args.ranked)
        graded_ranked = grade_df(df_ranked, date_str, actuals_by_id, name_to_id)

        base = f"graded_ranked_{date_str}_graded"
        out_csv = os.path.join(args.outdir, base + ".csv")
        out_xlsx = os.path.join(args.outdir, base + ".xlsx")

        graded_ranked.to_csv(out_csv, index=False)
        write_xlsx_with_tabs(graded_ranked, out_xlsx)

        print("✅ Saved:", out_csv)
        print("✅ Saved:", out_xlsx)

    # Top (overs/unders)
    if args.top or args.all:
        parts = []

        if args.overs:
            df_over = read_any(args.overs)
            graded_over = grade_df(df_over, date_str, actuals_by_id, name_to_id)
            graded_over["top_bucket"] = "OVERS"
            parts.append(graded_over)

        if args.unders:
            df_under = read_any(args.unders)
            graded_under = grade_df(df_under, date_str, actuals_by_id, name_to_id)
            graded_under["top_bucket"] = "UNDERS"
            parts.append(graded_under)

        if parts:
            graded_top = pd.concat(parts, axis=0, ignore_index=True)

            base = f"graded_top_{date_str}_graded"
            out_csv = os.path.join(args.outdir, base + ".csv")
            out_xlsx = os.path.join(args.outdir, base + ".xlsx")

            graded_top.to_csv(out_csv, index=False)
            write_xlsx_with_tabs(graded_top, out_xlsx)

            print("✅ Saved:", out_csv)
            print("✅ Saved:", out_xlsx)

    if not args.ranked and not (args.top or args.all):
        print("ℹ️ Nothing to do. Provide --ranked and/or --top with --overs/--unders.")


if __name__ == "__main__":
    main()
