#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
from typing import Optional, List, Dict

import numpy as np
import pandas as pd

# Excel (installed)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# ================= CONFIG =================
W_LAST5, W_LAST10, W_SEASON = 0.50, 0.30, 0.20
W_EDGE = 1.6
W_MINUTES = 0.35
W_FORM = 0.55

MAX_EXPORT_ROWS = 1_000_000

# Prop-type weighting (stat_prefix-based)
PROP_WEIGHTS = {
    # ── Core single stats (boosted) ──
    "pts": 1.06,
    "reb": 1.06,
    "ast": 1.06,

    # ── Combo stats (discounted) ──
    "pr": 0.98,
    "pa": 0.98,
    "ra": 0.98,
    "pra": 0.96,

    # ── Shooting volume / makes ──
    "fg3a": 1.1,
    "fg3m": 1.1,
    "fga": 1.1,
    "fgm": 1.02,
    "fg2a": 1.1,
    "fg2m": 1.01,

    # ── Free throws ──
    "fta": 1.00,
    "ftm": 1.00,

    # ── Fantasy (inflated) ──
    "fantasy": 0.95,

    # ── Defensive / misc ──
    "stocks": 1.00,

    # ── Turnovers ──
    "tov": 0.95,
}


PROP_TO_PREFIX = {
    "points": "pts", "pts": "pts",
    "rebounds": "reb", "rebs": "reb",
    "assists": "ast", "asts": "ast",
    "fantasy": "fantasy",
    "fantasy score": "fantasy",
    "turnovers": "tov",
    "stocks": "stocks",
    "pts+rebs": "pr", "points+rebounds": "pr",
    "pts+asts": "pa", "points+assists": "pa",
    "rebs+asts": "ra", "rebounds+assists": "ra",
    "pts+rebs+asts": "pra",
    "points+rebounds+assists": "pra",
    "pra": "pra",
    "3pta": "fg3a", "3ptm": "fg3m",
    "fga": "fga", "fgm": "fgm",
    "fta": "fta", "ftm": "ftm",
    "2pa": "fg2a",
    "2pm": "fg2m",
    "two pointers attempted": "fg2a",
    "two pointers made": "fg2m",
# PrizePicks “Attempts/Made” labels after your normalization
    "fg attempted": "fga",
    "fg made": "fgm",

    "3pt attempted": "fg3a",
    "3pt made": "fg3m",

}

COMBO_MAP = {
    "pts": ["pts"],
    "reb": ["reb"],
    "ast": ["ast"],
    "fantasy": ["fantasy"],
    "tov": ["tov"],
    "stocks": ["stocks"],
    "pr": ["pts", "reb"],
    "pa": ["pts", "ast"],
    "ra": ["reb", "ast"],
    "pra": ["pts", "reb", "ast"],
    "fg3a": ["fg3a"],
    "fg3m": ["fg3m"],
    "fga": ["fga"],
    "fgm": ["fgm"],
    "fta": ["fta"],
    "ftm": ["ftm"],
    "fg2a": ["fg2a"],
    "fg2m": ["fg2m"],
}

TEAM_NAME_MAP = {
    "ATLANTA HAWKS": "ATL", "BOSTON CELTICS": "BOS", "BROOKLYN NETS": "BKN",
    "CHARLOTTE HORNETS": "CHA", "CHICAGO BULLS": "CHI",
    "CLEVELAND CAVALIERS": "CLE", "DALLAS MAVERICKS": "DAL",
    "DENVER NUGGETS": "DEN", "DETROIT PISTONS": "DET",
    "GOLDEN STATE WARRIORS": "GSW", "HOUSTON ROCKETS": "HOU",
    "INDIANA PACERS": "IND", "LA CLIPPERS": "LAC", "LOS ANGELES CLIPPERS": "LAC",
    "LA LAKERS": "LAL", "LOS ANGELES LAKERS": "LAL",
    "MEMPHIS GRIZZLIES": "MEM", "MIAMI HEAT": "MIA",
    "MILWAUKEE BUCKS": "MIL", "MINNESOTA TIMBERWOLVES": "MIN",
    "NEW ORLEANS PELICANS": "NOP", "NEW YORK KNICKS": "NYK",
    "OKLAHOMA CITY THUNDER": "OKC", "ORLANDO MAGIC": "ORL",
    "PHILADELPHIA 76ERS": "PHI", "PHOENIX SUNS": "PHX",
    "PORTLAND TRAIL BLAZERS": "POR", "SACRAMENTO KINGS": "SAC",
    "SAN ANTONIO SPURS": "SAS", "TORONTO RAPTORS": "TOR",
    "UTAH JAZZ": "UTA", "WASHINGTON WIZARDS": "WAS",
}

# Columns you want visible (and ONLY these visible)
VISIBLE_COLS = [
    "nba_player_id",
    "player",
    "pos",
    "team",
    "opp_team",
    "tier",
    "line",
    "prop_type",
    "pick_type",
    "bet_dir",
    "min_season_avg",
    "last5_over",
    "last5_under",
    "last5_hit_rate",
    "last5_form_z",
    "TEAM_NAME",
    "OVERALL_DEF_RANK",
    "DEF_TIER",
    "edge",
    "abs_edge",
    "confidence",
    "prop_weight",
]


# ================= HELPERS =================
def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def normalize_stat_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    rename = {}
    for c in df.columns:
        cl = c.lower()
        if cl.endswith("_season") or cl.endswith("_last10") or cl.endswith("_last5"):
            rename[c] = cl + "_avg"
    if rename:
        df = df.rename(columns=rename)
    return df


def merge_defense(df: pd.DataFrame, path: Optional[str]) -> pd.DataFrame:
    if not path:
        return df

    df = df.copy()
    d = pd.read_csv(path)

    df.columns = [c.lower().strip() for c in df.columns]
    d.columns = [c.upper().strip() for c in d.columns]

    if "opp_team" in df.columns:
        prop_key = "opp_team"
    elif "opponent" in df.columns:
        prop_key = "opponent"
    else:
        print("⚠️ Props missing opponent column – skipping defense merge")
        return df

    if "TEAM_ABBREVIATION" in d.columns:
        def_key = "TEAM_ABBREVIATION"
        d[def_key] = d[def_key].astype(str).str.upper().str.strip()
    elif "TEAM_NAME" in d.columns:
        def_key = "TEAM_NAME"
        d[def_key] = (
            d[def_key]
            .astype(str)
            .str.upper()
            .str.strip()
            .map(TEAM_NAME_MAP)
        )
    else:
        print("⚠️ Defense CSV missing TEAM_NAME / TEAM_ABBREVIATION")
        return df

    df[prop_key] = df[prop_key].astype(str).str.upper().str.strip()
    d = d.drop_duplicates(subset=[def_key])

    df["opp_norm"] = df[prop_key]
    df = df.merge(d, left_on="opp_norm", right_on=def_key, how="left")
    df.drop(columns=["opp_norm"], inplace=True, errors="ignore")

    print("✅ Defense stats merged")
    return df


def add_two_pointer_stats(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for w in ["last5", "last10", "season"]:
        if f"fga_{w}_avg" in df.columns and f"fg3a_{w}_avg" in df.columns:
            df[f"fg2a_{w}_avg"] = pd.to_numeric(df[f"fga_{w}_avg"], errors="coerce") - pd.to_numeric(df[f"fg3a_{w}_avg"], errors="coerce")
        if f"fgm_{w}_avg" in df.columns and f"fg3m_{w}_avg" in df.columns:
            df[f"fg2m_{w}_avg"] = pd.to_numeric(df[f"fgm_{w}_avg"], errors="coerce") - pd.to_numeric(df[f"fg3m_{w}_avg"], errors="coerce")

    for i in range(1, 6):
        if f"fga_g{i}" in df.columns and f"fg3a_g{i}" in df.columns:
            df[f"fg2a_g{i}"] = pd.to_numeric(df[f"fga_g{i}"], errors="coerce") - pd.to_numeric(df[f"fg3a_g{i}"], errors="coerce")
        if f"fgm_g{i}" in df.columns and f"fg3m_g{i}" in df.columns:
            df[f"fg2m_g{i}"] = pd.to_numeric(df[f"fgm_g{i}"], errors="coerce") - pd.to_numeric(df[f"fg3m_g{i}"], errors="coerce")
    return df


def norm_prop_type(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip().lower().replace("-", "").replace("score", "")
    if "pts+rebs+asts" in s or s == "pra":
        return "pts+rebs+asts"
    if "pts+rebs" in s:
        return "pts+rebs"
    if "pts+asts" in s:
        return "pts+asts"
    if "rebs+asts" in s:
        return "rebs+asts"
    if "fantasy" in s:
        return "fantasy"
    if "2pa" in s or "twopointersattempted" in s:
        return "2pa"
    if "2pm" in s or "twopointersmade" in s:
        return "2pm"
    return s


def z_from_rank(r: pd.Series) -> pd.Series:
    r = pd.to_numeric(r, errors="coerce")
    if r.notna().sum() < 2:
        return pd.Series(0, index=r.index)

    mn, mx = r.min(), r.max()
    if pd.isna(mn) or pd.isna(mx) or mx == mn:
        return pd.Series(0, index=r.index)

    x = (r - mn) / (mx - mn)
    return (x - 0.5) * 2


def minutes_projection(df: pd.DataFrame) -> pd.Series:
    cols = [c for c in df.columns if c.startswith("min_")]
    if not cols:
        return pd.Series(np.nan, index=df.index)
    return df[cols].apply(pd.to_numeric, errors="coerce").mean(axis=1)


def get_dynamic_avg(row: pd.Series, prefix: str, window: str) -> float:
    if not prefix:
        return np.nan
    parts = COMBO_MAP.get(prefix, [prefix])
    total, found = 0.0, False
    for p in parts:
        c = f"{p}_{window}_avg"
        if c in row.index:
            v = pd.to_numeric(row[c], errors="coerce")
            if not pd.isna(v):
                total += float(v)
                found = True
    return total if found else np.nan


def last5_form(row: pd.Series, prefix: str, line: float):
    if not prefix or pd.isna(line):
        return 0, 0, 0.0

    parts = COMBO_MAP.get(prefix, [prefix])
    vals = []

    for i in range(1, 6):
        total, found = 0.0, False
        for p in parts:
            c = f"{p}_g{i}"
            if c in row.index:
                v = pd.to_numeric(row[c], errors="coerce")
                if not pd.isna(v):
                    total += float(v)
                    found = True
        if found:
            vals.append(total)

    if not vals:
        return 0, 0, 0.0

    over = sum(v > line for v in vals)
    under = sum(v < line for v in vals)
    hit_rate = round(max(over, under) / len(vals), 3)
    return over, under, hit_rate


def assign_tiers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    q = df["rank_score"].quantile([0.9, 0.65, 0.3]).values

    def tier(x):
        if x >= q[0]:
            return "A"
        elif x >= q[1]:
            return "B"
        elif x >= q[2]:
            return "C"
        else:
            return "D"

    df["tier"] = df["rank_score"].apply(tier)
    return df


def _ensure_nba_player_id(df: pd.DataFrame) -> pd.DataFrame:
    """
    User wants nba_player_id visible. If input uses player_id, map it.
    """
    df = df.copy()
    if "nba_player_id" not in df.columns:
        if "player_id" in df.columns:
            df["nba_player_id"] = df["player_id"]
        elif "nba_playerid" in df.columns:
            df["nba_player_id"] = df["nba_playerid"]
    return df


def _round_float_columns(df: pd.DataFrame, decimals: int = 2) -> pd.DataFrame:
    df = df.copy()
    float_cols = df.select_dtypes(include=["float", "float64", "float32"]).columns
    if len(float_cols) > 0:
        df[float_cols] = df[float_cols].round(decimals)
    return df


def write_xlsx_visible_only(
    df: pd.DataFrame,
    out_path: str,
    visible_cols: List[str],
    sheet_name: str = "Ranked",
) -> None:
    # Case-insensitive match for visible columns, while preserving desired casing/order
    col_map = {c.lower(): c for c in df.columns}

    resolved_visible = []
    for c in visible_cols:
        key = c.lower()
        if key in col_map:
            resolved_visible.append(col_map[key])
        else:
            # allow missing columns silently (e.g., defense columns not present)
            continue

    # Put visible first; keep everything else after (hidden)
    remaining = [c for c in df.columns if c not in set(resolved_visible)]
    df_out = df[resolved_visible + remaining].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write rows
    for row in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(row)

    # Freeze header, enable filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Hide non-visible columns
    visible_set = set(resolved_visible)
    for j, col_name in enumerate(df_out.columns, start=1):
        col_letter = get_column_letter(j)
        if col_name not in visible_set:
            ws.column_dimensions[col_letter].hidden = True

    # Format numbers to 0.00 for numeric columns (Excel formatting)
    numeric_cols = set(df_out.select_dtypes(include=["number"]).columns)
    for j, col_name in enumerate(df_out.columns, start=1):
        if col_name in numeric_cols:
            col_letter = get_column_letter(j)
            for cell in ws[col_letter][1:]:  # skip header row
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = "0.00"

    # Reasonable column widths for visible columns
    for j, col_name in enumerate(df_out.columns, start=1):
        col_letter = get_column_letter(j)
        if col_name in visible_set:
            ws.column_dimensions[col_letter].width = max(12, min(28, len(str(col_name)) + 4))

    wb.save(out_path)


# ================= MAIN =================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True, help="Use .csv or .xlsx")
    ap.add_argument("--defense-csv", required=False, default=None)  # <-- keeps your CLI working
    args = ap.parse_args()

    df = pd.read_csv(args.input)

    # normalize/standardize
    df = normalize_stat_columns(df)
    df = standardize_columns(df)
    df = add_two_pointer_stats(df)
    df = merge_defense(df, args.defense_csv)

    # required columns checks (soft)
    if "prop_type" not in df.columns:
        raise SystemExit("❌ Missing required column: prop_type")
    if "line" not in df.columns:
        raise SystemExit("❌ Missing required column: line")

    # prop parsing
    df["prop_type_norm"] = df["prop_type"].astype(str).apply(norm_prop_type)
    df["stat_prefix"] = df["prop_type_norm"].map(PROP_TO_PREFIX).fillna("")

    # minutes + z
    df["minutes_proj"] = minutes_projection(df)
    df["minutes_z"] = z_from_rank(df["minutes_proj"])

    # line
    df["line"] = pd.to_numeric(df["line"], errors="coerce")

    # prop weighting
    df["prop_weight"] = df["stat_prefix"].map(PROP_WEIGHTS).fillna(1.0)

    projections, o5, u5, r5 = [], [], [], []

    for _, row in df.iterrows():
        prefix = row["stat_prefix"]
        line = row["line"]

        l5 = get_dynamic_avg(row, prefix, "last5")
        l10 = get_dynamic_avg(row, prefix, "last10")
        seas = get_dynamic_avg(row, prefix, "season")

        parts, weights = [], []
        if not pd.isna(l5):
            parts.append(l5)
            weights.append(W_LAST5)
        if not pd.isna(l10):
            parts.append(l10)
            weights.append(W_LAST10)
        if not pd.isna(seas):
            parts.append(seas)
            weights.append(W_SEASON)

        proj = (sum(p * w for p, w in zip(parts, weights)) / sum(weights)) if parts else np.nan
        projections.append(proj)

        o, u, r = last5_form(row, prefix, line)
        o5.append(o)
        u5.append(u)
        r5.append(r)

    # ✅ (Indentation fixed) projection/edge
    df["projection"] = projections
    df["edge"] = df["projection"] - df["line"]

    # ✅ ONLY CHANGE ADDED: COMBO / FANTASY EDGE PENALTY
    EDGE_PENALTY = {
        "pr": 0.92,
        "pa": 0.92,
        "ra": 0.92,
        "pra": 0.88,
        "fantasy": 0.88,
    }
    m = df["stat_prefix"].isin(EDGE_PENALTY)
    df.loc[m, "edge"] = df.loc[m, "edge"] * df.loc[m, "stat_prefix"].map(EDGE_PENALTY)

    # abs edge
    df["abs_edge"] = df["edge"].abs()

    # bet direction (AFTER penalty)
    df["bet_dir"] = np.where(df["edge"] >= 0, "OVER", "UNDER")

    df["last5_over"] = o5
    df["last5_under"] = u5
    df["last5_hit_rate"] = r5
    df["last5_form_z"] = z_from_rank(df["last5_hit_rate"])

    # pick_type rules
    if "pick_type" in df.columns:
        df["pick_type"] = df["pick_type"].astype(str).str.lower()

        # Goblins & Demons = OVER only
        mask_invalid = df["pick_type"].isin(["goblin", "demon"]) & (df["bet_dir"] == "UNDER")
        df = df[~mask_invalid].copy()

    # rank score (multiply by prop_weight)
    base_score = (
        W_EDGE * df["abs_edge"]
        + W_MINUTES * df["minutes_z"].fillna(0)
        + W_FORM * df["last5_form_z"].fillna(0)
    )
    df["rank_score"] = base_score * df["prop_weight"]

    df = df.sort_values("rank_score", ascending=False)
    df = assign_tiers(df)

    if len(df) > MAX_EXPORT_ROWS:
        df = df.head(MAX_EXPORT_ROWS)

    # Ensure nba_player_id exists
    df = _ensure_nba_player_id(df)

    # min_season_avg: if not present, fall back to minutes_proj for display
    if "min_season_avg" not in df.columns:
        if "minutes_proj" in df.columns:
            df["min_season_avg"] = df["minutes_proj"]

    # ---------------- NEW CONFIDENCE (CORRELATION-FRIENDLY) ----------------
    # Blend edge + form + minutes; readable 0–10; then apply prop_weight
    edge_z = z_from_rank(df["abs_edge"].fillna(0))
    form_z = df["last5_form_z"].fillna(0)
    mins_z = df["minutes_z"].fillna(0)

    base = (0.55 * edge_z) + (0.35 * form_z) + (0.10 * mins_z)

    df["confidence"] = (5.0 + 2.0 * base) * df["prop_weight"]
    df["confidence"] = pd.to_numeric(df["confidence"], errors="coerce").fillna(0).clip(0, 10)


    # Round floats to 2 decimals for output
    df = _round_float_columns(df, decimals=2)

    out = args.output
    if out.lower().endswith(".xlsx"):
        # If defense cols exist in lowercase, expose uppercase aliases for visibility.
        for col in ["TEAM_NAME", "OVERALL_DEF_RANK", "DEF_TIER"]:
            if col not in df.columns and col.lower() in df.columns:
                df[col] = df[col.lower()]

        write_xlsx_visible_only(df, out, VISIBLE_COLS, sheet_name="Ranked")
        print("\n✅ Saved XLSX (visible cols only; extras hidden):", out)
    else:
        df.to_csv(out, index=False)
        print("\n✅ Saved CSV:", out)


if __name__ == "__main__":
    main()
