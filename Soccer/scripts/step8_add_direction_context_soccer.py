#!/usr/bin/env python3
"""
step8_add_direction_context_soccer.py  (Soccer Pipeline)

Mirrors NBA step8_add_direction_context.py.
Reads step7_soccer_ranked.xlsx, appends direction context,
outputs full CSV + clean formatted XLSX.

Run:
  py -3.14 step8_add_direction_context_soccer.py \
    --input step7_soccer_ranked.xlsx \
    --output step8_soccer_direction.csv
"""

from __future__ import annotations

import argparse
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
HEADER_COLOR = "1C1C1C"


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name: str, data: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    tier_bg, tier_fg = TIER_COLORS.get(name, ("333333", "FFFFFF"))
    headers = list(data.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        direction = row[headers.index("Direction")] if "Direction" in headers else ""
        for ci, val in enumerate(row, 1):
            col_name    = headers[ci - 1]
            display_val = "" if pd.isna(val) else val
            cell        = ws.cell(row=ri, column=ci, value=display_val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

            if col_name == "Tier":
                cell.fill = PatternFill("solid", start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name="Arial", size=9)
            elif col_name == "Direction":
                bg = "C8F7C5" if val == "OVER" else "F7C5C5" if val == "UNDER" else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, name="Arial", size=9)
            elif col_name == "League":
                cell.fill = PatternFill("solid", start_color="EBF5FB")
            else:
                cell.fill = PatternFill("solid", start_color="F9F9F9" if ri % 2 == 0 else "FFFFFF")

    col_widths = {
        "Tier": 6, "Rank Score": 10, "Player": 20, "Pos": 6, "Pos Group": 9,
        "Team": 12, "Opp": 12, "League": 12, "Game Time": 10,
        "Prop": 18, "Pick Type": 10, "Line": 7,
        "Direction": 9, "Edge": 7, "Projection": 10,
        "Hit Rate (5g)": 12, "Hit Rate (10g)": 12, "Last 5 Avg": 10, "Season Avg": 10,
        "L5 Over": 8, "L5 Under": 8,
        "Def Rank": 9, "Def Tier": 10,
        "Min Tier": 9, "Shot Role": 10, "Usage Role": 10,
        "Void Reason": 20,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%-I:%M %p")

    keep = [
        "tier", "rank_score",
        "player", "pos", "position_group", "team", "opp_team", "league", "game_time",
        "prop_type", "pick_type", "line",
        "final_bet_direction",
        "edge", "projection",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "OVERALL_DEF_RANK", "DEF_TIER",
        "minutes_tier", "shot_role", "usage_role",
        "void_reason",
    ]
    keep  = [c for c in keep if c in df2.columns]
    clean = df2[keep].copy()

    for col in ["rank_score", "edge", "projection", "line_hit_rate_over_ou_5", "line_hit_rate_over_ou_10"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(2)
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "player": "Player", "pos": "Pos", "position_group": "Pos Group",
        "team": "Team", "opp_team": "Opp", "league": "League", "game_time": "Game Time",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "final_bet_direction": "Direction",
        "edge": "Edge", "projection": "Projection",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "line_hit_rate_over_ou_10": "Hit Rate (10g)",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier",
        "minutes_tier": "Min Tier", "shot_role": "Shot Role", "usage_role": "Usage Role",
        "void_reason": "Void Reason",
    }
    clean = clean.rename(columns=rename)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "ALL", clean)
    for tier in ["A", "B", "C", "D"]:
        subset = clean[clean["Tier"] == tier].copy()
        # Always write every Tier sheet (even if empty) so step9 never crashes
        # on a missing sheet_name
        write_sheet(wb, f"Tier {tier}", subset if len(subset) else clean.head(0))

    wb.save(xlsx_path)
    print(f"📊 Clean XLSX saved → {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="s7_soccer_ranked.xlsx")
    ap.add_argument("--sheet",  default="ALL")
    ap.add_argument("--output", default="step8_soccer_direction.csv")
    ap.add_argument("--xlsx",   default="step8_soccer_direction_clean.xlsx")
    ap.add_argument("--date",   default="", help="YYYY-MM-DD target date (default: today ET)")
    args = ap.parse_args()

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    df  = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")

    if df.empty:
        print("❌ [SlateIQ-Soccer-S8] Empty input from S7 — aborting.")
        sys.exit(1)

    # ── Date filter: keep only target date's games ───────────────────────────
    import datetime, zoneinfo
    eastern = zoneinfo.ZoneInfo("America/New_York")
    target_str = (args.date.strip()[:10] if args.date
                  else datetime.datetime.now(tz=eastern).date().isoformat())
    if "start_time" in df.columns:
        before_filter = len(df)
        def _to_et_date(val):
            try:
                dt = pd.to_datetime(val)          # handles -04:00 offset natively
                if dt.tzinfo is None:
                    dt = dt.tz_localize("UTC")
                return dt.tz_convert(eastern).date().isoformat()
            except Exception:
                return ""
        df["_et_date"] = df["start_time"].apply(_to_et_date)
        df = df[df["_et_date"] == target_str].drop(columns="_et_date")
        dropped = before_filter - len(df)
        print(f"[DateFilter] Kept {len(df)}/{before_filter} rows for {target_str} ET (dropped {dropped} rows)")
    else:
        print("[DateFilter] WARNING: no start_time column — skipping date filter")

    out = df.copy()

    if "edge" not in out.columns:
        proj = pd.to_numeric(out.get("projection", ""), errors="coerce")
        line = pd.to_numeric(out.get("line",        ""), errors="coerce")
        out["edge"] = proj - line

    edge     = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = edge.abs()

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    forced    = pick_type.isin(["Goblin", "Demon"])

    # BUG FIX: edge >= 0 is False for NaN, which wrongly assigns UNDER to rows
    # with no projection. Use explicit NaN check.
    has_edge  = edge.notna()
    model_dir = pd.Series(
        np.where(has_edge & (edge >= 0), "OVER",
                 np.where(has_edge & (edge < 0), "UNDER", "NO_EDGE")),
        index=out.index
    )
    out["model_dir"] = model_dir

    final_dir = model_dir.copy()
    reason    = pd.Series(
        np.where(has_edge & (abs_edge < 0.03), "MODEL_TIEBREAK_DIFF<0.03", ""),
        index=out.index
    )
    final_dir = final_dir.where(~forced, "OVER")
    reason    = reason.where(~forced, "FORCED_OVER_ONLY_GOB_DEM")

    out["final_bet_direction"] = final_dir
    out["final_dir_reason"]    = reason

    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"✅ Saved → {args.output}")

    if out.empty:
        print("❌ [SlateIQ-Soccer-S8] Output is empty — aborting.")
        sys.exit(1)

    print("final_bet_direction counts:")
    print(pd.Series(out["final_bet_direction"]).value_counts().to_string())
    if "tier" in out.columns:
        print("tier counts:")
        print(out["tier"].value_counts().to_string())

    # Always use explicit --xlsx path (default: step8_soccer_direction_clean.xlsx)
    xlsx_path = args.xlsx if args.xlsx else "step8_soccer_direction_clean.xlsx"
    build_clean_xlsx(out, xlsx_path)


if __name__ == "__main__":
    main()
