#!/usr/bin/env python3
"""
step9_build_tickets.py
-----------------------
Reads step8_all_direction_clean.xlsx (Tier A sheet) and builds
optimized PrizePicks tickets prioritizing 3-leg tickets for best risk/reward.

Ticket building rules:
- Only Tier A props with Hit Rate >= min_hit_rate
- One player per game (no correlated legs)
- One prop per player per ticket
- Deduplicated so same leg doesn't appear in multiple tickets of same type
- Tickets ranked by average rank score

PrizePicks payouts (approximate):
  2-leg: 3x  |  3-leg: 5x  |  4-leg: 10x  |  5-leg: 20x

Run:
  py -3.14 step9_build_tickets.py --input step8_all_direction_clean.xlsx --output best_tickets.xlsx
"""

from __future__ import annotations

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PrizePicks Power Play payouts by leg count and pick type
# Standard = full odds | Goblin/Demon = reduced odds
PRIZEPICKS_PAYOUTS = {
    'Standard': {2: 3.0,  3: 5.0,  4: 10.0, 5: 20.0},
    'Goblin':   {2: 1.25, 3: 1.7,  4: 2.5,  5: 3.5 },
    'Demon':    {2: 1.85, 3: 3.0,  4: 5.5,  5: 10.0},
    'Best Mix': {2: 3.0,  3: 5.0,  4: 10.0, 5: 20.0},  # assumes Standard rates
}

def get_payout(pick_label: str, n_legs: int) -> float:
    table = PRIZEPICKS_PAYOUTS.get(pick_label, PRIZEPICKS_PAYOUTS['Standard'])
    return table.get(n_legs, 1.0)

def stake_to_win(target: float, pick_label: str, n_legs: int) -> float:
    """How much you need to stake to win target amount."""
    payout = get_payout(pick_label, n_legs)
    return round(target / payout, 2)

COLORS = {
    'Standard': '2874A6',
    'Goblin':   '6C3483',
    'Demon':    'C0392B',
    'Best Mix': '1E8449',
}
HDR_COLOR  = '1C1C1C'
DIR_OVER   = 'C8F7C5'
DIR_UNDER  = 'F7C5C5'
TIER_COLORS = {
    'Elite':     'D5F5E3',
    'Above Avg': 'EBF5FB',
    'Avg':       'FDFEFE',
    'Weak':      'FDEDEC',
}

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _norm_pick_type(x):
    t = str(x).strip().lower()
    if 'gob' in t: return 'Goblin'
    if 'dem' in t: return 'Demon'
    return 'Standard'

def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets: int, used_legs: set) -> list:
    """
    Greedy ticket builder.
    - Picks highest ranked anchor, fills with diverse games
    - Tracks used legs globally so same leg doesn't repeat across tickets
    """
    pool = pool[pool['Hit Rate (5g)'] >= 0.0].sort_values('Rank Score', ascending=False).copy()
    tickets = []
    used_ticket_keys = set()

    for _, anchor in pool.iterrows():
        leg_key = (anchor['Player'], anchor['Prop'], anchor['Line'])
        if leg_key in used_legs:
            continue

        ticket = [anchor]
        used_games = {
            (anchor['Team'], anchor['Opp']),
            (anchor['Opp'],  anchor['Team']),
        }

        for _, row in pool.iterrows():
            if len(ticket) >= n_legs:
                break
            if row['Player'] == anchor['Player']:
                continue
            lk = (row['Player'], row['Prop'], row['Line'])
            if lk in used_legs:
                continue
            game  = (row['Team'], row['Opp'])
            game_r= (row['Opp'],  row['Team'])
            if game in used_games or game_r in used_games:
                continue
            ticket.append(row)
            used_games.update([game, game_r])

        if len(ticket) == n_legs:
            tkey = frozenset((r['Player'], r['Prop'], str(r['Line'])) for r in ticket)
            if tkey not in used_ticket_keys:
                used_ticket_keys.add(tkey)
                tickets.append(ticket)
                # Mark legs as used so they don't repeat
                for r in ticket:
                    used_legs.add((r['Player'], r['Prop'], r['Line']))

        if len(tickets) >= max_tickets:
            break

    return tickets


def write_tickets_sheet(wb, sheet_name: str, tickets: list, n_legs: int, pick_label: str):
    if not tickets:
        return

    ws = wb.create_sheet(sheet_name)
    tab_color = COLORS.get(pick_label, '1E8449')
    payout = get_payout(pick_label, n_legs)

    cols    = ['#', 'Player', 'Team', 'Opp', 'Prop', 'Pick Type', 'Line',
               'Direction', 'Edge', 'Hit Rate', 'L5 Avg', 'Season Avg',
               'L5 Over', 'L5 Under', 'Rank Score', 'Def Tier']
    col_w   = [4,   22,      6,      6,     18,     10,          7,
               10,          7,      10,       9,          11,
               8,        8,          11,       11]

    # Set column widths once
    for ci, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    current_row = 1

    for t_idx, ticket in enumerate(tickets, 1):
        avg_score = sum(float(r['Rank Score']) for r in ticket) / len(ticket)
        avg_hr    = sum(float(r['Hit Rate (5g)']) for r in ticket) / len(ticket)
        exp_val   = avg_hr ** n_legs * payout  # rough EV per $1

        # Ticket header banner
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(cols))
        stake_for_100 = stake_to_win(100, pick_label, n_legs)
        hcell = ws.cell(
            row=current_row, column=1,
            value=(f"  Ticket #{t_idx}  ·  {n_legs}-Leg {pick_label}"
                   f"  ·  Payout: {payout}x"
                   f"  ·  Stake ${stake_for_100:.0f} to win $100"
                   f"  ·  Avg Hit Rate: {avg_hr:.0%}"
                   f"  ·  Est. Win Prob: {avg_hr**n_legs:.0%}"
                   f"  ·  Avg Rank Score: {avg_score:.2f}")
        )
        hcell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        hcell.fill      = PatternFill('solid', start_color=tab_color)
        hcell.alignment = Alignment(vertical='center', horizontal='left')
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Column headers
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=9)
            c.fill      = PatternFill('solid', start_color=HDR_COLOR)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = thin_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Legs
        for leg_i, row in enumerate(ticket, 1):
            row_bg  = 'F8F9FA' if leg_i % 2 == 0 else 'FFFFFF'
            def_bg  = TIER_COLORS.get(str(row.get('Def Tier', '')), 'FFFFFF')

            l5o = row.get('L5 Over',  '')
            l5u = row.get('L5 Under', '')
            l5o = '' if pd.isna(l5o) else int(l5o)
            l5u = '' if pd.isna(l5u) else int(l5u)

            vals = [
                leg_i,
                row['Player'],
                row['Team'],
                row['Opp'],
                row['Prop'],
                row['Pick Type'],
                row['Line'],
                row['Direction'],
                round(float(row['Edge']), 1),
                row['Hit Rate (5g)'],
                round(float(row['Last 5 Avg']), 1),
                round(float(row['Season Avg']), 1),
                l5o, l5u,
                round(float(row['Rank Score']), 2),
                row['Def Tier'],
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = thin_border()

                col_name = cols[ci - 1]
                if col_name == 'Direction':
                    bg = DIR_OVER if val == 'OVER' else DIR_UNDER
                    c.fill = PatternFill('solid', start_color=bg)
                    c.font = Font(bold=True, name='Arial', size=9)
                elif col_name == 'Def Tier':
                    c.fill = PatternFill('solid', start_color=def_bg)
                else:
                    c.fill = PatternFill('solid', start_color=row_bg)

            current_row += 1

        current_row += 1  # spacer


def write_summary_sheet(wb, all_ticket_sets: list):
    """Overview sheet showing all tickets at a glance."""
    ws = wb.create_sheet('SUMMARY', 0)

    headers = ['Sheet', 'Ticket #', 'Legs', 'Pick Type', 'Payout',
               'Avg Hit Rate', 'Est Win %', 'Avg Rank Score', 'Players']
    col_w   = [20,      10,       6,      12,          9,
               13,           12,          15,              50]

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=HDR_COLOR)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    row = 2
    for sheet_name, tickets, n_legs, pick_label in all_ticket_sets:
        payout = get_payout(pick_label, n_legs)
        tab_color = COLORS.get(pick_label, '1E8449')
        for t_idx, ticket in enumerate(tickets, 1):
            avg_score = sum(float(r['Rank Score']) for r in ticket) / len(ticket)
            avg_hr    = sum(float(r['Hit Rate (5g)']) for r in ticket) / len(ticket)
            players   = ' | '.join(f"{r['Player']} {r['Direction']} {r['Prop']} {r['Line']}" for r in ticket)

            vals = [sheet_name, t_idx, n_legs, pick_label, f"{payout}x",
                    f"{avg_hr:.0%}", f"{avg_hr**n_legs:.0%}",
                    round(avg_score, 2), players]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                c.border    = thin_border()
                bg = 'F8F9FA' if row % 2 == 0 else 'FFFFFF'
                c.fill = PatternFill('solid', start_color=bg)
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input',       default='step8_all_direction_clean.xlsx')
    ap.add_argument('--sheet',       default='Tier A')
    ap.add_argument('--output',      default='best_tickets.xlsx')
    ap.add_argument('--min_hit_rate',type=float, default=0.8)
    ap.add_argument('--max_tickets', type=int,   default=10)
    ap.add_argument('--legs',        default='2,3,4',
                    help='Comma-separated leg counts to generate, e.g. 2,3,4,5')
    args = ap.parse_args()

    leg_counts = [int(x.strip()) for x in args.legs.split(',')]

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    df = pd.read_excel(args.input, sheet_name=args.sheet)
    df['Pick Type'] = df['Pick Type'].astype(str).apply(_norm_pick_type)

    # Filter by min hit rate
    df = df[pd.to_numeric(df['Hit Rate (5g)'], errors='coerce') >= args.min_hit_rate].copy()
    print(f"  Props with hit rate >= {args.min_hit_rate}: {len(df)}")

    # Deduplicate - best rank score per player+prop+line
    df = df.sort_values('Rank Score', ascending=False).drop_duplicates(subset=['Player','Prop','Line','Direction'])

    # Separate pools
    std_pool = df[df['Pick Type']=='Standard'].drop_duplicates(subset=['Player'])
    gob_pool = df[df['Pick Type']=='Goblin'].drop_duplicates(subset=['Player'])
    all_pool = df.sort_values('Rank Score', ascending=False).drop_duplicates(subset=['Player'])

    pools = [
        ('Standard', std_pool),
        ('Goblin',   gob_pool),
        ('Best Mix', all_pool),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    all_ticket_sets = []

    for pick_label, pool in pools:
        if len(pool) < 2:
            print(f"  ⚠ Skipping {pick_label} — not enough props in pool ({len(pool)})")
            continue
        # Fresh used_legs per pool type so legs can appear across Standard/Goblin/Mix
        used_legs: set = set()
        for n_legs in leg_counts:
            if len(pool) < n_legs:
                continue
            tickets = build_tickets(pool.copy(), n_legs, args.max_tickets, used_legs)
            sheet_name = f"{pick_label} {n_legs}-Leg"
            write_tickets_sheet(wb, sheet_name, tickets, n_legs, pick_label)
            all_ticket_sets.append((sheet_name, tickets, n_legs, pick_label))
            print(f"  {sheet_name}: {len(tickets)} tickets")

    write_summary_sheet(wb, all_ticket_sets)

    wb.save(args.output)
    print(f"\n✅ Saved → {args.output}")
    print(f"\nPrizePicks Power Play payout reference:")
    print(f"  {'Type':<12} {'2-leg':>6} {'3-leg':>6} {'4-leg':>6} {'5-leg':>6}  {'Stake needed to win $100 (3-leg)':>35}")
    for label, table in PRIZEPICKS_PAYOUTS.items():
        if label == 'Best Mix': continue
        stake = stake_to_win(100, label, 3)
        print(f"  {label:<12} {table.get(2,''):>6} {table.get(3,''):>6} {table.get(4,''):>6} {table.get(5,''):>6}  ${stake:>6.2f} stake to win $100 on 3-leg")

if __name__ == '__main__':
    main()
