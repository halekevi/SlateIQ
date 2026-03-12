"""
Step 7 — Rank + Tier NHL Props
Scores each prop using a composite model and assigns A/B/C/D tier.

Scoring model:
  - composite_hit_rate (primary signal, weighted by stat stability)
  - defense tier (opponent difficulty)
  - scoring tier (player quality)
  - PP tier (power play usage context)
  - sample size confidence
  - home/road adjustment

Usage:
    py step7_rank_props_nhl.py --input step6_nhl_role_context.csv \
        --output step7_nhl_ranked.xlsx
"""

import argparse
import csv
import subprocess
import sys
try:
    from tqdm import tqdm as _tqdm
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm", "--break-system-packages", "-q"])
    from tqdm import tqdm as _tqdm

# ── Head-to-Head (H2H) utility ────────────────────────────────────────────────
def _attach_h2h(df: "pd.DataFrame", cache_path: str, sport: str,
                player_col: str, opp_col: str, prop_col: str, line_col: str) -> "pd.DataFrame":
    """
    Attach H2H stats per row: how did this player perform vs this opponent
    historically in the boxscore cache?

    Adds columns:
      h2h_games      – number of H2H games found
      h2h_avg        – player's average stat value vs this opp
      h2h_over_rate  – fraction of those games where they hit OVER the current line
      h2h_last        – most recent game value vs this opp
    """
    import os, pandas as pd, numpy as np

    if not cache_path or not os.path.exists(cache_path):
        df["h2h_games"]     = 0
        df["h2h_avg"]       = np.nan
        df["h2h_over_rate"] = np.nan
        df["h2h_last"]      = np.nan
        return df

    try:
        cache = pd.read_csv(cache_path, low_memory=False)
    except Exception:
        df["h2h_games"]     = 0
        df["h2h_avg"]       = np.nan
        df["h2h_over_rate"] = np.nan
        df["h2h_last"]      = np.nan
        return df

    # Normalise cache columns: need player, opponent, stat_type, value, date
    cache.columns = [c.lower().strip() for c in cache.columns]

    # Detect player col in cache
    p_col  = next((c for c in ["player_norm","player_name","player","name"] if c in cache.columns), None)
    o_col  = next((c for c in ["opp_team","opp","opponent","opp_team_abbr"] if c in cache.columns), None)
    s_col  = next((c for c in ["stat_type","stat","prop_type","stat_norm"]   if c in cache.columns), None)
    v_col  = next((c for c in ["value","stat_value","actual","val"]          if c in cache.columns), None)
    d_col  = next((c for c in ["date","game_date","event_date"]              if c in cache.columns), None)

    if not all([p_col, o_col, v_col]):
        df["h2h_games"]     = 0
        df["h2h_avg"]       = np.nan
        df["h2h_over_rate"] = np.nan
        df["h2h_last"]      = np.nan
        return df

    cache[v_col] = pd.to_numeric(cache[v_col], errors="coerce")

    def _norm(x):
        return str(x).strip().lower() if x and str(x).strip() else ""

    # Build lookup: (player_norm, opp_norm) -> list of (value, date)
    lookup: dict = {}
    for row in cache.itertuples(index=False):
        pk = (_norm(getattr(row, p_col)), _norm(getattr(row, o_col)))
        v  = getattr(row, v_col)
        dt = getattr(row, d_col, "") if d_col else ""
        if pk not in lookup:
            lookup[pk] = []
        lookup[pk].append((v, str(dt)))

    h2h_games, h2h_avg, h2h_over, h2h_last = [], [], [], []

    for _, r in df.iterrows():
        player = _norm(r.get(player_col, ""))
        opp    = _norm(r.get(opp_col, ""))
        line   = r.get(line_col, None)
        try:
            line_f = float(line)
        except (TypeError, ValueError):
            line_f = None

        entries = lookup.get((player, opp), [])
        # sort by date desc, take up to 10
        try:
            entries_sorted = sorted(entries, key=lambda x: x[1], reverse=True)[:10]
        except Exception:
            entries_sorted = entries[:10]

        vals = [v for v, _ in entries_sorted if v is not None and not (isinstance(v, float) and v != v)]

        if not vals:
            h2h_games.append(0)
            h2h_avg.append(np.nan)
            h2h_over.append(np.nan)
            h2h_last.append(np.nan)
        else:
            avg = round(float(np.mean(vals)), 2)
            last = vals[0]
            over_rate = round(sum(1 for v in vals if line_f is not None and v > line_f) / len(vals), 3) if line_f else np.nan
            h2h_games.append(len(vals))
            h2h_avg.append(avg)
            h2h_over.append(over_rate)
            h2h_last.append(round(float(last), 2))

    df["h2h_games"]     = h2h_games
    df["h2h_avg"]       = h2h_avg
    df["h2h_over_rate"] = h2h_over
    df["h2h_last"]      = h2h_last
    return df
# ─────────────────────────────────────────────────────────────────────────────


# Stat stability weights (higher = more predictable/trackable)
STAT_STABILITY = {
    "shots_on_goal": 1.12,   # most consistent NHL prop — volume metric
    "saves":         1.10,   # consistent for starting goalies
    "assists":       1.00,
    "points":        1.00,
    "hits":          1.05,   # consistent for physical players
    "blocked_shots": 1.05,
    "goals":         0.82,   # highest variance in hockey
    "goals_allowed": 0.85,
    "fantasy_score": 0.95,
}

# Defense tier adjustments for OVER hit rate
DEF_TIER_BOOST = {
    "WEAK":    +0.04,    # easiest opp = small boost to over
    "AVERAGE": +0.00,
    "SOLID":   -0.02,
    "ELITE":   -0.05,    # hardest opp = penalize over
}

# Scoring tier adjustments
SCORING_TIER_BOOST = {
    "ELITE":    +0.05,
    "SECONDARY": +0.02,
    "DEPTH":     -0.02,
    "SHUTDOWN":  -0.04,
    "GOALIE":    +0.0,
    "UNKNOWN":   +0.0,
}

PP_TIER_BOOST = {
    "PP1_STAR":   +0.04,
    "PP_REGULAR": +0.01,
    "PP_OCC":     +0.0,
    "NO_PP":      -0.01,
    "N/A":        +0.0,
}

HOME_BOOST = 0.01
MIN_SAMPLE = 5  # Minimum games to be rankable


def install_openpyxl():
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])


def read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def safe_float(val, default=0.0) -> float:
    try:
        return float(val)
    except Exception:
        return default


def score_prop(row: dict) -> float:
    stat = row.get("stat_norm", "")
    composite = safe_float(row.get("composite_hit_rate"))
    if composite == 0.0 and row.get("composite_hit_rate", "") == "":
        # No hit rate data — return a neutral score so prop still appears with D tier
        return 0.0
    recommended_side = row.get("recommended_side", "OVER")

    # Flip composite to represent confidence in recommended side
    confidence = composite if recommended_side == "OVER" else 1 - composite

    # Stat stability weight
    stab = STAT_STABILITY.get(stat, 1.0)
    base_score = confidence * stab

    # Defense adjustment (applies to OVER on skater props; reverse for UNDER)
    def_tier = row.get("def_tier", "AVERAGE")
    def_adj = DEF_TIER_BOOST.get(def_tier, 0.0)
    if recommended_side == "UNDER":
        def_adj = -def_adj  # Under benefits from tougher defense

    # Scoring tier
    scoring_tier = row.get("scoring_tier", "DEPTH")
    score_adj = SCORING_TIER_BOOST.get(scoring_tier, 0.0)

    # PP tier
    pp_tier = row.get("pp_tier", "N/A")
    pp_adj = PP_TIER_BOOST.get(pp_tier, 0.0)

    # Home/road
    is_home = str(row.get("is_home", "0")) == "1"
    home_adj = HOME_BOOST if is_home else 0.0

    # Sample confidence (penalize small samples)
    sample = safe_float(row.get("sample_L10", 0))
    sample_conf = min(sample / 10.0, 1.0)

    total = (base_score + def_adj + score_adj + pp_adj + home_adj) * sample_conf
    return round(total, 5)


def assign_tier(score: float, sample: float) -> str:
    if sample < MIN_SAMPLE:
        return "D"
    if score >= 0.68:
        return "A"
    elif score >= 0.60:
        return "B"
    elif score >= 0.50:
        return "C"
    else:
        return "D"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="step6_nhl_role_context.csv")
    parser.add_argument("--output", default="step7_nhl_ranked.xlsx")
    parser.add_argument("--min-sample", type=int, default=MIN_SAMPLE)
    parser.add_argument("--cache", default="", help="Path to NHL boxscore cache CSV")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        install_openpyxl()
        import openpyxl

    rows = read_csv(args.input)

    scored = []
    for row in _tqdm(rows, desc="  Scoring props", unit="prop"):
        prop_score = score_prop(row)
        sample = safe_float(row.get("sample_L10", 0))
        tier = assign_tier(prop_score, sample)
        row["prop_score"] = prop_score
        row["tier"] = tier
        scored.append(row)

    scored.sort(key=lambda x: -safe_float(x.get("prop_score", 0)))

    # ── Pass Demons through; only drop neg-edge Goblins to audit sheet ──────────
    def _norm_pt(x: str) -> str:
        t = (x or "").strip().lower()
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    active  = []
    dropped = []
    for row in scored:
        pt    = _norm_pt(row.get("pick_type", ""))
        edge  = safe_float(row.get("edge", 0))
        if pt == "Goblin" and edge < 0:
            row["void_reason"] = "DROPPED_NEG_EDGE_GOBLIN"
            dropped.append(row)
        else:
            # Demons pass through for data/tracking — excluded from tickets in combined_slate
            active.append(row)

    # Add rank only on active rows
    for i, row in enumerate(active):
        row["rank"] = i + 1
    for row in dropped:
        row["rank"] = ""

    print(f"  Active: {len(active)} | Dropped (neg-edge Goblin only): {len(dropped)}")

    # ── Head-to-Head stats ────────────────────────────────────────────────────
    if args.cache:
        import pandas as _pd
        df_h2h = _pd.DataFrame(active)
        player_col = next((c for c in ["player_name","player_norm","player"] if c in df_h2h.columns), "")
        opp_col    = next((c for c in ["opp_team","opp","pp_opp_team","opp_team_abbr"] if c in df_h2h.columns), "")
        line_col   = next((c for c in ["line_score","line"] if c in df_h2h.columns), "line_score")
        prop_col   = next((c for c in ["stat_norm","prop_type"] if c in df_h2h.columns), "stat_norm")
        if player_col and opp_col:
            df_h2h = _attach_h2h(df_h2h, args.cache, "nhl", player_col, opp_col, prop_col, line_col)
            active = df_h2h.to_dict("records")
            print(f"  H2H: {sum(1 for r in active if r.get('h2h_games',0) > 0)}/{len(active)} rows matched")

    # Write XLSX with multiple tabs
    wb = openpyxl.Workbook()

    # ── All Props tab ──────────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Props"

    headers = list(active[0].keys()) if active else (list(dropped[0].keys()) if dropped else [])
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    TIER_COLORS = {"A": "C6EFCE", "B": "FFEB9C", "C": "FFCCCC", "D": "E0E0E0"}
    for row in _tqdm(active, desc="  Writing All Props sheet", unit="row"):
        ws_all.append([row.get(h, "") for h in headers])
        last_row = ws_all.max_row
        tier_color = TIER_COLORS.get(row.get("tier", "D"), "FFFFFF")
        for col in range(1, len(headers) + 1):
            ws_all.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor=tier_color)

    # ── Skaters tab ────────────────────────────────────────────────────────────
    ws_sk = wb.create_sheet("Skaters")
    skaters = [r for r in active if r.get("player_role") == "SKATER"]
    if skaters:
        sk_headers = list(skaters[0].keys())
        for col, h in enumerate(sk_headers, 1):
            cell = ws_sk.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in skaters:
            ws_sk.append([row.get(h, "") for h in sk_headers])
            last_row = ws_sk.max_row
            tier_color = TIER_COLORS.get(row.get("tier", "D"), "FFFFFF")
            for col in range(1, len(sk_headers) + 1):
                ws_sk.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor=tier_color)

    # ── Goalies tab ────────────────────────────────────────────────────────────
    ws_g = wb.create_sheet("Goalies")
    goalies = [r for r in active if r.get("player_role") == "GOALIE"]
    if goalies:
        g_headers = list(goalies[0].keys())
        for col, h in enumerate(g_headers, 1):
            cell = ws_g.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in goalies:
            ws_g.append([row.get(h, "") for h in g_headers])
            last_row = ws_g.max_row
            tier_color = TIER_COLORS.get(row.get("tier", "D"), "FFFFFF")
            for col in range(1, len(g_headers) + 1):
                ws_g.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor=tier_color)

    # ── A-Tier only ────────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("A-Tier Best")
    a_props = [r for r in active if r.get("tier") == "A"]
    if a_props:
        a_headers = list(a_props[0].keys())
        for col, h in enumerate(a_headers, 1):
            cell = ws_a.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="375623")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in a_props:
            ws_a.append([row.get(h, "") for h in a_headers])
            last_row = ws_a.max_row
            for col in range(1, len(a_headers) + 1):
                ws_a.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="C6EFCE")

    # ── DROPPED tab (neg-edge Goblin audit) ───────────────────────────────────
    ws_drop = wb.create_sheet("DROPPED")
    if dropped:
        d_headers = list(dropped[0].keys())
        for col, h in enumerate(d_headers, 1):
            cell = ws_drop.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="7B241C")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in dropped:
            ws_drop.append([row.get(h, "") for h in d_headers])
            last_row = ws_drop.max_row
            for col in range(1, len(d_headers) + 1):
                ws_drop.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FADBD8")

    # Autofit columns (approximate)
    for ws in [ws_all, ws_sk, ws_g, ws_a, ws_drop]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    wb.save(args.output)
    print(f"Saved ranked props -> {args.output}")

    # Summary
    tier_counts = {}
    for r in active:
        t = r.get("tier", "?")
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"Tier breakdown: {tier_counts}")
    print(f"Dropped (neg-edge Goblin only): {len(dropped)}")
    print(f"\nTop 10 props:")
    for r in active[:10]:
        print(f"  #{r['rank']} [{r['tier']}] {r['player_name']} {r['stat_norm']} "
              f"{r['line_score']} {r['recommended_side']} | score={r['prop_score']:.4f}")


if __name__ == "__main__":
    main()
