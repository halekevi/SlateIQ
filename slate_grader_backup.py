#!/usr/bin/env python3
"""
Full Slate Grader — grades every prop in step8 NBA or step6 CBB
against actual results, producing breakdowns matching step6_graded_props format.

Usage:
  # NBA
  py -3.14 slate_grader.py --sport NBA --slate step8_all_direction_clean.xlsx --actuals actuals_nba.csv --output nba_graded_2026-02-20.xlsx

  # CBB
  py -3.14 slate_grader.py --sport CBB --slate step6_ranked_cbb.xlsx --actuals actuals_cbb.csv --output cbb_graded_2026-02-20.xlsx

  # Template
  py -3.14 slate_grader.py --template --sport NBA
  py -3.14 slate_grader.py --template --sport CBB

actuals CSV columns: player, prop_type, actual
(just the actual stat value — grader computes HIT/MISS/PUSH)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
from datetime import datetime

# ── Colors ────────────────────────────────────────────────────────────────────
C = {
    'hit':      '27AE60', 'miss':    'E74C3C', 'push':    'F39C12',
    'void':     '95A5A6', 'hdr':     '1C1C1C', 'hdr2':    '1A5276',
    'hdr3':     '1E8449', 'hdr4':    '7D6608', 'hdr5':    '922B21',
    'alt':      'F2F3F4', 'white':   'FFFFFF',
    'tier_a':   'D5F5E3', 'tier_b':  'D6EAF8',
    'tier_c':   'FEF9E7', 'tier_d':  'FDEDEC',
}

def side(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hc(ws, r, c, v, bg=None, fc='FFFFFF', bold=True, sz=9, align='center', wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=fc, name='Arial', size=sz)
    if bg: cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = side()
    return cell

def dc(ws, r, c, v, bg=None, bold=False, sz=9, align='center', fmt=None, fc='000000'):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, name='Arial', size=sz, color=fc)
    cell.fill = PatternFill('solid', start_color=bg or C['white'])
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = side()
    if fmt: cell.number_format = fmt
    return cell

def res_bg(r):
    r = str(r).upper()
    return {'HIT': C['hit'], 'MISS': C['miss'], 'PUSH': C['push'], 'VOID': C['void']}.get(r, 'DDDDDD')

def hr_bg(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return 'DDDDDD'
    if v >= 0.65: return C['hit']
    if v >= 0.50: return C['push']
    return C['miss']

def tier_bg(t):
    return {'A': C['tier_a'], 'B': C['tier_b'], 'C': C['tier_c'], 'D': C['tier_d']}.get(str(t).upper(), C['white'])

# ── Grade a single prop ───────────────────────────────────────────────────────
def grade(row, actual):
    if pd.isna(actual):
        return 'VOID', 'NO_ACTUAL', 0
    line = float(row['line']) if 'line' in row else float(row.get('Line', 0))
    direction = str(row.get('bet_direction', row.get('Direction', 'OVER'))).upper()
    if actual == line:
        return 'VOID', 'PUSH', 0
    if direction == 'OVER':
        result = 'HIT' if actual > line else 'MISS'
    else:
        result = 'HIT' if actual < line else 'MISS'
    margin = round(actual - line if direction == 'OVER' else line - actual, 2)
    return result, None, margin

# ── Load and normalize NBA slate ──────────────────────────────────────────────
def load_nba(path):
    df = pd.read_excel(path, sheet_name='ALL')
    df = df.rename(columns={
        'Player': 'player', 'Team': 'team', 'Opp': 'opp_team',
        'Line': 'line', 'Prop': 'prop_type_norm', 'Pick Type': 'pick_type',
        'Direction': 'bet_direction', 'Tier': 'tier', 'Edge': 'edge',
        'Rank Score': 'rank_score', 'Hit Rate (5g)': 'last5_hit_rate',
        'Last 5 Avg': 'last5_avg', 'Season Avg': 'season_avg',
        'Def Tier': 'def_tier', 'Def Rank': 'def_rank',
        'L5 Over': 'last5_over', 'L5 Under': 'last5_under',
        'Projection': 'projection', 'Pos': 'pos',
        'Void Reason': 'void_reason', 'Min Tier': 'min_tier',
    })
    df['abs_edge'] = df['edge'].abs()
    df['abs_edge_bucket'] = pd.cut(df['abs_edge'],
        bins=[-np.inf, 1, 2, 3, np.inf],
        labels=['0-0.99','1-1.99','2-2.99','3+'])
    df['player_key'] = df['player'].str.lower().str.strip() + '|' + df['prop_type_norm'].str.lower().str.strip()
    return df

# ── Load and normalize CBB slate ──────────────────────────────────────────────
def load_cbb(path):
    df = pd.read_excel(path, sheet_name='ALL')
    df['player_key'] = df['player'].str.lower().str.strip() + '|' + df['prop_type'].str.lower().str.strip()
    df['prop_type_norm'] = df['prop_type']
    df['bet_direction'] = df['final_bet_direction']
    df['last5_hit_rate'] = df['line_hit_rate']
    df['last5_avg'] = df['stat_last5_avg']
    df['season_avg'] = df['stat_season_avg']
    df['opp_team'] = df['opp_team_abbr']
    df['abs_edge_bucket'] = pd.cut(df['abs_edge'],
        bins=[-np.inf, 1, 2, 3, np.inf],
        labels=['0-0.99','1-1.99','2-2.99','3+'])
    return df

# ── Apply actuals ─────────────────────────────────────────────────────────────
def apply_actuals(df, actuals_path):
    act = pd.read_csv(actuals_path)
    act['player_key'] = act['player'].str.lower().str.strip() + '|' + act['prop_type'].str.lower().str.strip()
    act_map = dict(zip(act['player_key'], act['actual']))

    results, void_reasons, margins = [], [], []
    for _, row in df.iterrows():
        key = row.get('player_key', '')
        actual = act_map.get(key, np.nan)
        if pd.notna(row.get('void_reason', np.nan)):
            results.append('VOID')
            void_reasons.append(row['void_reason'])
            margins.append(np.nan)
        else:
            r, vr, m = grade(row, actual)
            results.append(r)
            void_reasons.append(vr)
            margins.append(m)

    df['result'] = results
    df['void_reason_grade'] = void_reasons
    df['margin'] = margins
    df['actual'] = df.get('player_key', pd.Series()).map(act_map)
    df['result_sign'] = df['result'].map({'HIT': 1, 'MISS': -1, 'VOID': 0, 'PUSH': 0})
    return df

# ── Breakdown helper ──────────────────────────────────────────────────────────
def breakdown(df, group_col):
    grp = df.groupby(group_col, dropna=False)
    rows = []
    for key, sub in grp:
        total  = len(sub)
        hit    = (sub['result'] == 'HIT').sum()
        miss   = (sub['result'] == 'MISS').sum()
        void   = (sub['result'].isin(['VOID','PUSH'])).sum()
        decided= hit + miss
        hr     = hit / decided if decided > 0 else np.nan
        rows.append({group_col: key, 'total': total, 'hit': hit,
                     'miss': miss, 'void': void, 'decided': decided,
                     'hit_rate': round(hr, 4) if not np.isnan(hr) else np.nan})
    return pd.DataFrame(rows).sort_values('total', ascending=False)

# ── Write breakdown sheet ─────────────────────────────────────────────────────
def write_breakdown(wb, df_b, sheet_name, group_col, bg_hdr):
    ws = wb.create_sheet(sheet_name)
    cols = [group_col, 'total', 'hit', 'miss', 'void', 'decided', 'hit_rate']
    widths = [28, 8, 8, 8, 8, 10, 12]
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        hc(ws, 1, ci, col, bg=bg_hdr)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for ri, row in enumerate(df_b.itertuples(), 2):
        bg = C['alt'] if ri % 2 == 0 else C['white']
        hr = getattr(row, 'hit_rate', np.nan)
        hr_b = hr_bg(hr) if not (isinstance(hr, float) and np.isnan(hr)) else 'DDDDDD'
        dc(ws, ri, 1, getattr(row, group_col.replace(' ', '_'), ''), bg=bg, align='left')
        dc(ws, ri, 2, getattr(row, 'total', ''), bg=bg)
        dc(ws, ri, 3, int(getattr(row, 'hit', 0)), bg=bg)
        dc(ws, ri, 4, int(getattr(row, 'miss', 0)), bg=bg)
        dc(ws, ri, 5, int(getattr(row, 'void', 0)), bg=bg)
        dc(ws, ri, 6, int(getattr(row, 'decided', 0)), bg=bg)
        c = dc(ws, ri, 7, hr if not (isinstance(hr, float) and np.isnan(hr)) else '', bg=hr_b, bold=True)
        if not (isinstance(hr, float) and np.isnan(hr)):
            c.number_format = '0.0%'
            c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')

# ── Write Box Raw sheet ───────────────────────────────────────────────────────
def write_raw(wb, df):
    ws = wb.create_sheet('Box Raw')
    cols = ['player','team','opp_team','prop_type_norm','pick_type','line',
            'bet_direction','tier','edge','abs_edge','last5_hit_rate',
            'last5_avg','season_avg','last5_over','last5_under',
            'projection','rank_score','actual','result','margin','void_reason_grade']
    widths = [22,6,6,16,10,7,10,5,8,8,13,10,12,9,10,12,12,9,8,8,16]
    present = [c for c in cols if c in df.columns]
    for ci, col in enumerate(present, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci-1] if ci <= len(widths) else 12
        hc(ws, 1, ci, col, bg=C['hdr'])
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for ri, row in enumerate(df[present].itertuples(), 2):
        bg = C['alt'] if ri % 2 == 0 else C['white']
        res = str(getattr(row, 'result', '')).upper()
        for ci, col in enumerate(present, 1):
            val = getattr(row, col, '')
            c_bg = res_bg(res) if col == 'result' else (
                tier_bg(val) if col == 'tier' else bg)
            c = dc(ws, ri, ci, val, bg=c_bg,
                   align='left' if col == 'player' else 'center')
            if col == 'result' and res in ('HIT','MISS','PUSH','VOID'):
                c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
            if col in ('last5_hit_rate',) and val != '' and val is not None:
                try: c.number_format = '0%'
                except: pass
    ws.auto_filter.ref = f"A1:{get_column_letter(len(present))}1"

# ── Write Summary / Dashboard ─────────────────────────────────────────────────
def write_dashboard(wb, df, sport, date_str):
    ws = wb.create_sheet('Summary', 0)
    ws.column_dimensions['A'].width = 28
    for ci in range(2, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f"{sport} SLATE GRADE  |  {date_str}  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(bold=True, name='Arial', size=12, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=C['hdr'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    decided = df[df['result'].isin(['HIT','MISS'])]
    total_d = len(decided)
    hits    = (decided['result'] == 'HIT').sum()
    misses  = (decided['result'] == 'MISS').sum()
    voids   = df['result'].isin(['VOID','PUSH']).sum()
    hr      = hits / total_d if total_d else 0

    # Overall stats
    row = 2
    hc(ws, row, 1, 'OVERALL', bg=C['hdr2'])
    for ci, h in enumerate(['Total Props','Decided','Hits','Misses','Voids','Hit Rate'], 2):
        hc(ws, row, ci, h, bg=C['hdr2'])
    row += 1
    dc(ws, row, 1, 'Full Slate', bold=True, align='left')
    dc(ws, row, 2, len(df))
    dc(ws, row, 3, total_d)
    dc(ws, row, 4, int(hits))
    dc(ws, row, 5, int(misses))
    dc(ws, row, 6, int(voids))
    c = dc(ws, row, 7, hr, bg=hr_bg(hr), bold=True)
    c.number_format = '0.0%'
    c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')

    # By pick type
    row += 2
    hc(ws, row, 1, 'BY PICK TYPE', bg=C['hdr3'])
    for ci, h in enumerate(['Total','Decided','Hits','Misses','Voids','Hit Rate'], 2):
        hc(ws, row, ci, h, bg=C['hdr3'])
    row += 1
    for pt in sorted(df['pick_type'].dropna().unique()):
        sub = df[df['pick_type'] == pt]
        dec = sub[sub['result'].isin(['HIT','MISS'])]
        h   = (dec['result'] == 'HIT').sum()
        m   = (dec['result'] == 'MISS').sum()
        v   = sub['result'].isin(['VOID','PUSH']).sum()
        hr2 = h / len(dec) if len(dec) else 0
        bg  = C['alt'] if row % 2 == 0 else C['white']
        dc(ws, row, 1, pt, bold=True, align='left')
        dc(ws, row, 2, len(sub), bg=bg)
        dc(ws, row, 3, len(dec), bg=bg)
        dc(ws, row, 4, int(h),   bg=bg)
        dc(ws, row, 5, int(m),   bg=bg)
        dc(ws, row, 6, int(v),   bg=bg)
        c = dc(ws, row, 7, hr2, bg=hr_bg(hr2), bold=True)
        c.number_format = '0.0%'
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        row += 1

    # By tier
    row += 1
    hc(ws, row, 1, 'BY TIER', bg=C['hdr4'])
    for ci, h in enumerate(['Total','Decided','Hits','Misses','Voids','Hit Rate'], 2):
        hc(ws, row, ci, h, bg=C['hdr4'])
    row += 1
    for t in ['A','B','C','D']:
        sub = df[df['tier'].astype(str).str.upper() == t]
        if len(sub) == 0: continue
        dec = sub[sub['result'].isin(['HIT','MISS'])]
        h   = (dec['result'] == 'HIT').sum()
        m   = (dec['result'] == 'MISS').sum()
        v   = sub['result'].isin(['VOID','PUSH']).sum()
        hr2 = h / len(dec) if len(dec) else 0
        tb  = tier_bg(t)
        dc(ws, row, 1, f'Tier {t}', bold=True, align='left', bg=tb)
        dc(ws, row, 2, len(sub), bg=tb)
        dc(ws, row, 3, len(dec), bg=tb)
        dc(ws, row, 4, int(h),   bg=tb)
        dc(ws, row, 5, int(m),   bg=tb)
        dc(ws, row, 6, int(v),   bg=tb)
        c = dc(ws, row, 7, hr2, bg=hr_bg(hr2), bold=True)
        c.number_format = '0.0%'
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        row += 1

    # By direction
    row += 1
    hc(ws, row, 1, 'BY DIRECTION', bg=C['hdr5'])
    for ci, h in enumerate(['Total','Decided','Hits','Misses','Voids','Hit Rate'], 2):
        hc(ws, row, ci, h, bg=C['hdr5'])
    row += 1
    for direction in ['OVER','UNDER']:
        sub = df[df['bet_direction'].str.upper() == direction]
        dec = sub[sub['result'].isin(['HIT','MISS'])]
        h   = (dec['result'] == 'HIT').sum()
        m   = (dec['result'] == 'MISS').sum()
        v   = sub['result'].isin(['VOID','PUSH']).sum()
        hr2 = h / len(dec) if len(dec) else 0
        bg  = C['alt'] if row % 2 == 0 else C['white']
        dc(ws, row, 1, direction, bold=True, align='left')
        dc(ws, row, 2, len(sub), bg=bg)
        dc(ws, row, 3, len(dec), bg=bg)
        dc(ws, row, 4, int(h),   bg=bg)
        dc(ws, row, 5, int(m),   bg=bg)
        dc(ws, row, 6, int(v),   bg=bg)
        c = dc(ws, row, 7, hr2, bg=hr_bg(hr2), bold=True)
        c.number_format = '0.0%'
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        row += 1

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sport',   default='NBA', choices=['NBA','CBB'])
    ap.add_argument('--slate',   default='', help='step8_all_direction_clean.xlsx or step6_ranked_cbb.xlsx')
    ap.add_argument('--actuals', default='', help='CSV with player,prop_type,actual columns')
    ap.add_argument('--output',  default='')
    ap.add_argument('--template',action='store_true')
    ap.add_argument('--date',    default=datetime.now().strftime('%Y-%m-%d'))
    args = ap.parse_args()

    if args.template:
        pd.DataFrame(columns=['player','prop_type','actual']).to_csv(
            f'actuals_{args.sport.lower()}.csv', index=False)
        print(f'Saved actuals_{args.sport.lower()}.csv')
        print('Fill in player names (exactly as they appear in the slate), prop type, and actual stat value.')
        print('prop_type examples: Points, Rebounds, Assists, PRA, Pts+Asts, Fantasy Score, Blocked Shots')
        return

    if not args.slate:
        print('ERROR: --slate required. Use --template to get the actuals CSV format.')
        return

    if not args.output:
        args.output = f'{args.sport.lower()}_graded_{args.date}.xlsx'

    # Load slate
    print(f'Loading {args.sport} slate...')
    if args.sport == 'NBA':
        df = load_nba(args.slate)
    else:
        df = load_cbb(args.slate)

    print(f'  {len(df)} props loaded')

    # Apply actuals if provided
    if args.actuals:
        print(f'Applying actuals from {args.actuals}...')
        df = apply_actuals(df, args.actuals)
        decided = df[df['result'].isin(['HIT','MISS'])]
        hits = (decided['result'] == 'HIT').sum()
        print(f'  Graded: {len(decided)} props — {hits} HIT / {len(decided)-hits} MISS')
    else:
        df['result'] = 'PENDING'
        df['void_reason_grade'] = ''
        df['margin'] = np.nan
        df['actual'] = np.nan
        df['result_sign'] = 0
        print('  No actuals provided — outputting slate with PENDING results')

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    write_dashboard(wb, df, args.sport, args.date)
    write_raw(wb, df)

    # Breakdowns
    if 'pick_type' in df.columns:
        write_breakdown(wb, breakdown(df, 'pick_type'), 'By Pick Type', 'pick_type', C['hdr3'])
    if 'tier' in df.columns:
        write_breakdown(wb, breakdown(df, 'tier'), 'By Tier', 'tier', C['hdr4'])
    if 'prop_type_norm' in df.columns:
        write_breakdown(wb, breakdown(df, 'prop_type_norm'), 'By Prop Type', 'prop_type_norm', C['hdr2'])
    if 'bet_direction' in df.columns:
        write_breakdown(wb, breakdown(df, 'bet_direction'), 'By Direction', 'bet_direction', C['hdr5'])
    if 'abs_edge_bucket' in df.columns:
        write_breakdown(wb, breakdown(df, 'abs_edge_bucket'), 'By Edge Bucket', 'abs_edge_bucket', C['hdr'])
    if 'def_tier' in df.columns:
        write_breakdown(wb, breakdown(df, 'def_tier'), 'By Opp Def', 'def_tier', C['hdr2'])
    if 'void_reason' in df.columns:
        vr_col = 'void_reason'
        write_breakdown(wb, breakdown(df, vr_col), 'Void Reasons', vr_col, C['hdr'])

    wb.save(args.output)
    print(f'\nSaved -> {args.output}')
    print('Sheets:', wb.sheetnames)
    print()
    print('Usage:')
    print(f'  Template: py -3.14 slate_grader.py --sport {args.sport} --template')
    print(f'  Grade:    py -3.14 slate_grader.py --sport {args.sport} --slate [file] --actuals actuals_{args.sport.lower()}.csv --output {args.output}')

if __name__ == '__main__':
    main()
