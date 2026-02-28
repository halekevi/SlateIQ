"""
build_tickets_html.py
=====================
Converts combined_slate_tickets_*.xlsx into a styled tickets_latest.html
and saves it to ui_runner/docs/tickets_latest.html.

Usage:
    py -3.14 build_tickets_html.py
    py -3.14 build_tickets_html.py --date 2026-02-24
    py -3.14 build_tickets_html.py --input path\\to\\file.xlsx
"""

from __future__ import annotations

import argparse
import html as html_lib
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).resolve().parent
OUTPUTS_DIR   = SCRIPT_DIR / "outputs"
UI_DOCS_DIR   = SCRIPT_DIR / "ui_runner" / "templates"


# ── Helpers ───────────────────────────────────────────────────────────────────
def find_latest_tickets(date_str: str | None = None) -> Path:
    pattern = "combined_slate_tickets_*.xlsx"
    candidates = []
    for d in [SCRIPT_DIR] + sorted(OUTPUTS_DIR.glob("*"), reverse=True):
        if Path(d).is_dir():
            for f in Path(d).glob(pattern):
                if "TOP3" not in f.name:
                    candidates.append(f)
    if not candidates:
        raise FileNotFoundError(f"No combined_slate_tickets_*.xlsx found under {SCRIPT_DIR}")
    if date_str:
        matches = [c for c in candidates if date_str in c.name]
        if matches:
            return sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def extract_date(path: Path) -> tuple[str, str]:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    if m:
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d")
            return d.strftime("%b %d, %Y").upper(), m.group(1)
        except ValueError:
            pass
    now = datetime.now()
    return now.strftime("%b %d, %Y").upper(), now.strftime("%Y-%m-%d")


def h(v: Any) -> str:
    return html_lib.escape(str(v) if v is not None else "")


def fmt(v: Any, dec: int = 2) -> str:
    try:
        return f"{float(v):.{dec}f}"
    except (TypeError, ValueError):
        return str(v) if v is not None else "—"


def pct(v: Any) -> str:
    try:
        f = float(v)
        return f"{f*100:.0f}%" if f <= 1.0 else f"{f:.0f}%"
    except (TypeError, ValueError):
        return str(v) if v is not None else "—"


def rate_color(v: Any) -> tuple[str, str]:
    try:
        f = float(v)
        if f <= 1.0: f *= 100
    except (TypeError, ValueError):
        return "#94a3b8", "var(--muted2)"
    if f >= 80:  return "#6ee7b7", "var(--green)"
    if f >= 65:  return "#6ee7b7", "var(--green)"
    if f >= 55:  return "#fcd34d", "var(--amber)"
    if f >= 50:  return "#93c5fd", "var(--blue)"
    return "#fca5a5", "var(--red)"


def rate_bar(v: Any) -> str:
    try:
        f = float(v)
        if f <= 1.0: f *= 100
    except (TypeError, ValueError):
        return f'<span class="mono muted">{h(v)}</span>'
    tc, bc = rate_color(v)
    return (f'<div class="rbar"><div class="rbar-bg">'
            f'<div class="rbar-fill" style="width:{min(f,100):.1f}%;background:{bc}"></div>'
            f'</div><span class="rbar-num" style="color:{tc}">{f:.0f}%</span></div>')


def tier_chip(t: Any) -> str:
    s = str(t).strip().upper() if t else ""
    cls = {"A":"chip-a","B":"chip-b","C":"chip-c","D":"chip-d"}.get(s, "chip-d")
    return f'<span class="chip {cls}">T{h(s)}</span>' if s else "—"


def pick_chip(p: Any) -> str:
    s = str(p).strip().lower() if p else ""
    if "goblin" in s:  return '<span class="chip chip-goblin">&#x1F47A; Goblin</span>'
    if "demon"  in s:  return '<span class="chip chip-demon">&#x1F608; Demon</span>'
    return '<span class="chip chip-std">&#x2B50; Std</span>'


def dir_chip(d: Any) -> str:
    s = str(d).strip().upper() if d else ""
    if s == "OVER":  return '<span class="chip chip-over">&#x25B2; OVER</span>'
    if s == "UNDER": return '<span class="chip chip-under">&#x25BC; UNDER</span>'
    return f'<span class="chip chip-d">{h(d)}</span>' if d else "—"


def def_chip(d: Any) -> str:
    s = str(d).strip().title() if d else ""
    cls = {"Elite":"chip-demon","Strong":"chip-c","Average":"chip-std",
           "Weak":"chip-a","Very Weak":"chip-a"}.get(s, "chip-d")
    return f'<span class="chip {cls}">{h(s)}</span>' if s else "—"


def sport_chip(s: Any) -> str:
    v = str(s).strip().upper() if s else ""
    if "NBA" in v: return '<span class="chip chip-b">NBA</span>'
    if "CBB" in v: return '<span class="chip chip-goblin">CBB</span>'
    return f'<span class="chip chip-d">{h(v)}</span>'


# ── Sheet parsers ─────────────────────────────────────────────────────────────
def read_flat_sheet(ws) -> list[dict]:
    """Sheets like Full Slate / NBA Slate — first row is header."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [str(c).strip() if c else f"_c{i}" for i, c in enumerate(rows[0])]
    return [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]


def read_ticket_sheet(ws) -> list[dict]:
    """
    Ticket sheets pattern:
      Row N:   '  Ticket #1  · 3-Leg NBA Goblin · Power: 4.37x ...'  (title)
      Row N+1: '#', 'Player', 'Team', ...  (headers)
      Row N+2+: data rows
      Row M:   next ticket title...
    """
    rows = list(ws.iter_rows(values_only=True))
    tickets = []
    i = 0
    while i < len(rows):
        first = str(rows[i][0]).strip() if rows[i][0] else ""
        if first.startswith("Ticket"):
            title = first
            i += 1
            if i >= len(rows): break
            headers = [str(c).strip() if c else f"_c{j}" for j, c in enumerate(rows[i])]
            i += 1
            legs = []
            while i < len(rows):
                dr = rows[i]
                fd = str(dr[0]).strip() if dr[0] else ""
                if fd.startswith("Ticket"): break
                if any(v is not None for v in dr):
                    legs.append(dict(zip(headers, dr)))
                i += 1
            tickets.append({"title": title, "legs": legs})
        else:
            i += 1
    return tickets


def parse_ticket_title(title: str) -> dict:
    info = {"num":"","desc":"","power":"","flex":"","avg_hit":"","est_prob":"","avg_score":""}
    m = re.search(r"Ticket\s*#?(\d+)", title)
    if m: info["num"] = m.group(1)
    m = re.search(r"·\s*([^·]+Leg[^·]*)", title)
    if m: info["desc"] = m.group(1).strip()
    m = re.search(r"Power:\s*([\d.]+x)", title)
    if m: info["power"] = m.group(1)
    m = re.search(r"Flex:\s*([\d.]+x)", title)
    if m: info["flex"] = m.group(1)
    m = re.search(r"Avg Hit Rate:\s*([\d.]+%)", title)
    if m: info["avg_hit"] = m.group(1)
    m = re.search(r"Est Win Prob:\s*([\d.]+%)", title)
    if m: info["est_prob"] = m.group(1)
    m = re.search(r"Avg Rank Score:\s*([\d.]+)", title)
    if m: info["avg_score"] = m.group(1)
    return info


# ── Slate table ───────────────────────────────────────────────────────────────
def build_slate_table(rows: list[dict], limit: int = 500) -> str:
    if not rows:
        return '<div class="alert alert-amber"><div class="alert-title">No data.</div></div>'
    rows = rows[:limit]
    body = ""
    for r in rows:
        body += f"""<tr>
          <td>{sport_chip(r.get('Sport'))}</td>
          <td>{tier_chip(r.get('Tier'))}</td>
          <td class="mono right">{fmt(r.get('Rank Score'))}</td>
          <td><strong>{h(r.get('Player',''))}</strong><div class="sub">{h(r.get('Team',''))} vs {h(r.get('Opp',''))}</div></td>
          <td class="mono">{h(r.get('Prop',''))}</td>
          <td>{pick_chip(r.get('Pick Type'))}</td>
          <td class="mono right">{fmt(r.get('Line'),1)}</td>
          <td>{dir_chip(r.get('Dir'))}</td>
          <td class="mono right pos">{fmt(r.get('Edge'),2)}</td>
          <td>{rate_bar(r.get('Hit Rate'))}</td>
          <td class="mono right">{fmt(r.get('L5 Avg'),1)}</td>
          <td>{def_chip(r.get('Def Tier'))}</td>
          <td class="mono muted small">{h(r.get('Game Time',''))}</td>
        </tr>"""
    return f"""<div class="table-wrap scrollx">
  <table>
    <thead><tr>
      <th></th><th>TIER</th><th class="right">SCORE</th><th>PLAYER</th>
      <th>PROP</th><th>TYPE</th><th class="right">LINE</th><th>DIR</th>
      <th class="right">EDGE</th><th>HIT RATE</th><th class="right">L5</th>
      <th>DEF</th><th>TIME</th>
    </tr></thead>
    <tbody>{body}</tbody>
  </table>
</div>"""


# ── Ticket card ───────────────────────────────────────────────────────────────
def build_ticket_card(ticket: dict) -> str:
    info = parse_ticket_title(ticket["title"])
    pills = ""
    if info["power"]:    pills += f'<span class="pill pill-green">&#x26A1; {h(info["power"])} Power</span>'
    if info["flex"]:     pills += f'<span class="pill pill-blue">&#x1F500; {h(info["flex"])} Flex</span>'
    if info["avg_hit"]:  pills += f'<span class="pill pill-amber">&#x1F3AF; {h(info["avg_hit"])} Hit</span>'
    if info["est_prob"]: pills += f'<span class="pill pill-purple">&#x1F4CA; {h(info["est_prob"])} Win</span>'

    legs_html = ""
    for leg in ticket["legs"]:
        legs_html += f"""<tr>
          <td><strong>{h(leg.get('Player',''))}</strong><div class="sub">{h(leg.get('Team',''))} vs {h(leg.get('Opp',''))}</div></td>
          <td class="mono">{h(leg.get('Prop',''))}</td>
          <td>{pick_chip(leg.get('Pick Type'))}</td>
          <td class="mono right">{fmt(leg.get('Line'),1)}</td>
          <td>{dir_chip(leg.get('Dir'))}</td>
          <td class="mono right pos">{fmt(leg.get('Edge'),2)}</td>
          <td>{rate_bar(leg.get('Hit Rate'))}</td>
          <td class="mono right muted">{fmt(leg.get('L5 Avg'),1)}</td>
          <td class="mono right muted">{fmt(leg.get('Rank Score'),2)}</td>
        </tr>"""

    return f"""<div class="ticket-card">
  <div class="ticket-header">
    <div class="ticket-num">#{h(info['num'])}</div>
    <div class="ticket-desc">{h(info['desc'])}</div>
    <div class="ticket-pills">{pills}</div>
  </div>
  <div class="table-wrap">
    <table>
      <thead><tr>
        <th>PLAYER</th><th>PROP</th><th>TYPE</th><th class="right">LINE</th>
        <th>DIR</th><th class="right">EDGE</th><th>HIT RATE</th>
        <th class="right">L5</th><th class="right">SCORE</th>
      </tr></thead>
      <tbody>{legs_html}</tbody>
    </table>
  </div>
</div>"""


def build_ticket_group(tickets: list[dict], limit: int = 60) -> str:
    if not tickets:
        return '<div class="muted" style="padding:20px;font-family:\'DM Mono\',monospace;font-size:12px">No tickets in this group.</div>'
    cards = "".join(build_ticket_card(t) for t in tickets[:limit])
    note  = (f'<div class="muted small" style="padding:8px;font-family:\'DM Mono\',monospace">'
             f'Showing {min(len(tickets),limit)} of {len(tickets)} tickets</div>') if len(tickets) > limit else ""
    return cards + note


# ── Tab section builder ───────────────────────────────────────────────────────
def build_tab_section(wb, sheet_names: list[str], id_prefix: str) -> tuple[str, str]:
    """Group sheets by pick type, sub-tab by leg count."""
    # group by pick type: Goblin, Standard, Demon, Mix
    type_map: dict[str, dict[str, list]] = {}
    for name in sheet_names:
        if name not in wb.sheetnames: continue
        # extract pick type from sheet name
        # handles: "NBA Goblin 3-Leg", "MIX Standard 3-Leg", "COMBO Mix 3-Leg"
        m_leg  = re.search(r"(\d+)-Leg", name)
        m_type = re.search(r"\b(Goblin|Standard|Demon|Mix)\b", name, re.I)
        if not m_leg: continue
        leg  = m_leg.group(1)
        ptype = m_type.group(1).title() if m_type else "Other"
        tickets = read_ticket_sheet(wb[name])
        if not tickets: continue
        type_map.setdefault(ptype, {}).setdefault(leg, [])
        type_map[ptype][leg].extend(tickets)

    if not type_map:
        return "", ""

    type_order = ["Goblin", "Standard", "Demon", "Mix", "Other"]
    btns_html   = ""
    panels_html = ""
    first_type  = True

    for ptype in type_order:
        if ptype not in type_map: continue
        leg_data = type_map[ptype]
        type_id  = f"{id_prefix}-{ptype.lower()}"

        # Inner leg stabs
        inner_btns   = ""
        inner_panels = ""
        first_leg    = True
        for leg in sorted(leg_data.keys(), key=lambda x: int(x) if x.isdigit() else 99):
            lid     = f"{type_id}-{leg}"
            tickets = leg_data[leg]
            active  = "active" if first_leg else ""
            inner_btns   += f'<button class="stab {active}" onclick="switchStab(event,\'{lid}\')">{leg}-Leg <span class="count-badge">{len(tickets)}</span></button>'
            inner_panels += f'<div id="{lid}" class="stab-panel {active}">{build_ticket_group(tickets)}</div>'
            first_leg = False

        total = sum(len(v) for v in leg_data.values())
        type_label = {"Goblin":"👺 Goblin","Standard":"⭐ Standard","Demon":"😈 Demon","Mix":"🔀 Mix"}.get(ptype, ptype)
        active_type = "active" if first_type else ""

        btns_html   += f'<button class="top-tab {active_type}" onclick="switchTop(event,\'{type_id}\')">{type_label} <span class="count-badge" style="margin-left:4px">{total}</span></button>'
        panels_html += f"""<div id="{type_id}" class="top-panel {active_type}">
  <div class="stab-bar" style="margin-bottom:16px">{inner_btns}</div>
  {inner_panels}
</div>"""
        first_type = False

    return btns_html, panels_html


# ── KPI cards ─────────────────────────────────────────────────────────────────
def build_kpi(rows: list[dict], total_tickets: int, display_date: str) -> str:
    total_props = len(rows)
    nba  = sum(1 for r in rows if str(r.get("Sport","")).upper()=="NBA")
    cbb  = sum(1 for r in rows if str(r.get("Sport","")).upper()=="CBB")
    hrs  = []
    for r in rows:
        try:
            f = float(r.get("Hit Rate", 0))
            hrs.append(f if f > 1 else f*100)
        except (TypeError, ValueError):
            pass
    avg_hr = sum(hrs)/len(hrs) if hrs else 0
    tc, _  = rate_color(avg_hr/100)
    # parse date for display
    m = re.search(r"(\w+)\s+(\d+),\s+(\d+)", display_date)
    mon = m.group(1)[:3] if m else display_date[:6]
    day = m.group(2) if m else ""

    return f"""<div class="stat-grid stat-grid-4">
  <div class="stat-card green">
    <div class="stat-label">TOTAL PROPS</div>
    <div class="stat-val" style="color:#10b981">{total_props:,}</div>
    <div class="stat-sub">{nba:,} NBA &nbsp;·&nbsp; {cbb:,} CBB</div>
  </div>
  <div class="stat-card blue">
    <div class="stat-label">TOTAL TICKETS</div>
    <div class="stat-val" style="color:#60a5fa">{total_tickets:,}</div>
    <div class="stat-sub">All types &amp; leg counts</div>
  </div>
  <div class="stat-card amber">
    <div class="stat-label">AVG HIT RATE</div>
    <div class="stat-val" style="color:{tc}">{avg_hr:.1f}%</div>
    <div class="stat-sub">Full slate average</div>
  </div>
  <div class="stat-card purple">
    <div class="stat-label">SLATE DATE</div>
    <div class="stat-val" style="color:#c4b5fd;font-size:28px">{mon} {day}</div>
    <div class="stat-sub">2026</div>
  </div>
</div>"""


# ── CSS ───────────────────────────────────────────────────────────────────────
CSS = """
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&display=swap');
:root{
  --bg:#05050f;--bg2:#0d0d1f;--bg3:#111128;--border:#1e1e3a;--bd2:#2a2a4a;
  --text:#e8e8f0;--muted:#999;--muted2:#666;
  --accent:#c8ff00;--cyan:#00e5ff;
  --green:#39ff6e;--amber:#f0a500;--red:#ff4d4d;--purple:#a78bfa;--blue:#00e5ff;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Share Tech Mono',monospace;background:var(--bg);color:var(--text);min-height:100vh;padding-bottom:80px;overflow-x:hidden;}
body::before{content:'';position:fixed;inset:0;background-image:linear-gradient(rgba(200,255,0,.03) 1px,transparent 1px),linear-gradient(90deg,rgba(200,255,0,.03) 1px,transparent 1px);background-size:40px 40px;animation:gridScroll 20s linear infinite;pointer-events:none;z-index:0;}
@keyframes gridScroll{from{background-position:0 0;}to{background-position:0 40px;}}
body::after{content:'';position:fixed;inset:0;background:repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(0,0,0,.18) 2px,rgba(0,0,0,.18) 4px);pointer-events:none;z-index:0;}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:var(--bg2)}::-webkit-scrollbar-thumb{background:var(--bd2);border-radius:4px}
/* brain logo */
.brain-wrap{position:relative;width:46px;height:46px;flex-shrink:0;}
.brain-slate{position:absolute;inset:0;border-radius:7px;background:linear-gradient(145deg,#12122a 0%,#080818 100%);border:1px solid #252545;animation:slateBreak 3.5s ease-in-out infinite;}
.brain-slate::before{content:'';position:absolute;inset:0;background:linear-gradient(to bottom right,transparent 47%,#c8ff0044 49%,transparent 51%),linear-gradient(to bottom left,transparent 44%,#c8ff0022 46%,transparent 48%);border-radius:7px;animation:crackGlow 3.5s ease-in-out infinite;}
@keyframes slateBreak{0%,100%{transform:scale(1);box-shadow:0 0 0px #c8ff0000;}48%{transform:scale(1.06) rotate(-0.5deg);box-shadow:0 0 24px #c8ff0055;}50%{transform:scale(1.10) rotate(0.5deg);box-shadow:0 0 40px #c8ff0088;}52%{transform:scale(1.06);box-shadow:0 0 24px #c8ff0055;}}
@keyframes crackGlow{0%,100%{opacity:0.2;}50%{opacity:1;}}
.brain-svg-el{position:absolute;inset:3px;animation:brainBT 3.5s ease-in-out infinite;transform-origin:center bottom;}
@keyframes brainBT{0%,100%{transform:scale(1) translateY(0px);filter:drop-shadow(0 0 5px #c8ff0099) drop-shadow(0 0 2px #00e5ff66);}48%{transform:scale(1.07) translateY(-1px);filter:drop-shadow(0 0 12px #c8ff00cc) drop-shadow(0 0 8px #00e5ffaa);}50%{transform:scale(1.18) translateY(-3px);filter:drop-shadow(0 0 20px #c8ff00ff) drop-shadow(0 0 14px #00e5ffcc);}52%{transform:scale(1.07) translateY(-1px);}}
.bpr{position:absolute;border-radius:9px;border:1.5px solid #c8ff00;opacity:0;animation:bRing 3.5s ease-out infinite;inset:-3px;}
.bpr:nth-child(2){border-color:#00e5ff;animation-delay:.15s;}.bpr:nth-child(3){border-color:#c8ff0088;animation-delay:.3s;}.bpr:nth-child(4){border-color:#00e5ff66;animation-delay:.45s;}
@keyframes bRing{0%,48%{transform:scale(1);opacity:0;}50%{opacity:.9;}85%{transform:scale(2.4);opacity:0;}100%{transform:scale(2.4);opacity:0;}}
.bsp{position:absolute;border-radius:50%;opacity:0;animation:bSpark 3.5s ease-out infinite;}
.bsp.lg{width:4px;height:4px;background:#c8ff00;box-shadow:0 0 6px #c8ff00;}.bsp.md{width:3px;height:3px;background:#00e5ff;box-shadow:0 0 5px #00e5ff;}.bsp.sm{width:2px;height:2px;background:#c8ff00cc;}.bsp.cy{width:2px;height:2px;background:#00e5ffcc;}
.bsp:nth-child(5){top:10%;left:5%;--tx:-16px;--ty:-14px;animation-delay:.50s;}.bsp:nth-child(6){top:5%;left:40%;--tx:2px;--ty:-20px;animation-delay:.52s;}.bsp:nth-child(7){top:8%;left:75%;--tx:14px;--ty:-16px;animation-delay:.54s;}.bsp:nth-child(8){top:30%;left:96%;--tx:20px;--ty:-8px;animation-delay:.51s;}.bsp:nth-child(9){top:55%;left:96%;--tx:18px;--ty:8px;animation-delay:.53s;}.bsp:nth-child(10){top:80%;left:85%;--tx:12px;--ty:14px;animation-delay:.55s;}.bsp:nth-child(11){top:92%;left:55%;--tx:4px;--ty:20px;animation-delay:.50s;}.bsp:nth-child(12){top:90%;left:25%;--tx:-10px;--ty:18px;animation-delay:.52s;}.bsp:nth-child(13){top:70%;left:2%;--tx:-18px;--ty:10px;animation-delay:.54s;}.bsp:nth-child(14){top:45%;left:0%;--tx:-20px;--ty:0px;animation-delay:.51s;}.bsp:nth-child(15){top:20%;left:2%;--tx:-16px;--ty:-10px;animation-delay:.56s;}.bsp:nth-child(16){top:15%;left:60%;--tx:8px;--ty:-18px;animation-delay:.53s;}.bsp:nth-child(17){top:18%;left:20%;--tx:-12px;--ty:-16px;animation-delay:.65s;}.bsp:nth-child(18){top:12%;left:55%;--tx:6px;--ty:-18px;animation-delay:.67s;}.bsp:nth-child(19){top:25%;left:88%;--tx:16px;--ty:-12px;animation-delay:.66s;}.bsp:nth-child(20){top:60%;left:93%;--tx:16px;--ty:10px;animation-delay:.68s;}
@keyframes bSpark{0%,47%{opacity:0;transform:translate(0,0) scale(0);}50%{opacity:1;transform:translate(0,0) scale(1);}75%{opacity:.5;}95%{opacity:0;transform:translate(var(--tx),var(--ty)) scale(.2);}100%{opacity:0;}}
/* nav */
.snav{position:sticky;top:0;z-index:200;background:rgba(2,4,8,.92);backdrop-filter:blur(32px) saturate(1.8);border-bottom:1px solid var(--border);padding:0 40px;display:flex;align-items:center;height:72px;gap:0;}
.nav-accent{position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent 0%,var(--lime,#c6ff00) 15%,var(--cyan) 40%,var(--purple) 65%,var(--lime,#c6ff00) 85%,transparent 100%);opacity:.7;animation:accentShift 6s ease-in-out infinite alternate;}
@keyframes accentShift{from{opacity:.5;filter:hue-rotate(0deg);}to{opacity:.9;filter:hue-rotate(20deg);}}
.snav-brand{display:flex;align-items:center;gap:14px;text-decoration:none;margin-right:48px;flex-shrink:0;}
.snav-name{font-family:'Bebas Neue',sans-serif;font-size:22px;letter-spacing:4px;color:var(--text);}
.snav-name span{color:var(--accent);}
.snav-links{display:flex;align-items:stretch;list-style:none;flex:1;gap:4px;}
.snav-links li a{display:flex;align-items:center;gap:10px;padding:0 20px;height:72px;font-family:'Share Tech Mono',monospace;font-size:12px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#7a9ac8;text-decoration:none;border-bottom:2px solid transparent;transition:color .2s,border-color .2s,background .2s;position:relative;}
.snav-links li a:hover{color:var(--text);background:rgba(255,255,255,.02);}
.snav-links li a.active{color:var(--accent);border-bottom-color:var(--accent);}
.snav-links li a.active::after{content:'';position:absolute;bottom:-1px;left:50%;transform:translateX(-50%);width:6px;height:6px;border-radius:50%;background:var(--accent);box-shadow:0 0 12px var(--accent),0 0 24px rgba(200,255,0,.5);}
.ni{width:28px;height:28px;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0;transition:transform .2s;}
.snav-links li a:hover .ni{transform:scale(1.1) translateY(-1px);}
.ni-ctrl{background:rgba(198,255,0,.1);border:1px solid rgba(198,255,0,.2);}
.ni-tick{background:rgba(0,212,255,.1);border:1px solid rgba(0,212,255,.2);}
.ni-slate{background:rgba(198,255,0,.1);border:1px solid rgba(198,255,0,.2);}
.ni-pay{background:rgba(0,255,170,.1);border:1px solid rgba(0,255,170,.2);}
.ni-grade{background:rgba(168,85,247,.1);border:1px solid rgba(168,85,247,.2);}
.snav-right{display:flex;align-items:center;gap:16px;margin-left:auto;}
.live-pill{font-family:'Share Tech Mono',monospace;font-size:10px;letter-spacing:2px;padding:6px 14px;border-radius:100px;border:1px solid rgba(0,255,170,.25);color:var(--green);background:rgba(0,255,170,.04);display:flex;align-items:center;gap:8px;}
.live-dot{width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 8px var(--green);animation:blink 2s ease-in-out infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}
/* header */
header{position:relative;z-index:1;background:rgba(5,5,15,.92);backdrop-filter:blur(12px);border-bottom:1px solid var(--border);padding:18px 32px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;position:sticky;top:62px;z-index:90;}
.logo{display:flex;align-items:center;gap:14px;}
.logo-title{font-family:'Bebas Neue',sans-serif;font-size:28px;letter-spacing:.1em;color:var(--accent);}
.logo-sub{font-size:10px;color:var(--muted);letter-spacing:2.5px;margin-top:2px;}
.date-badge{font-size:11px;color:var(--muted);background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:6px 14px;letter-spacing:1px;}
/* layout */
.main{position:relative;z-index:1;max-width:1300px;margin:0 auto;padding:28px 20px;}
.sport-header{display:flex;align-items:center;gap:14px;margin-bottom:20px;}
.sport-label{font-family:'Bebas Neue',sans-serif;font-size:30px;letter-spacing:.08em;line-height:1;}
.sport-header-line{flex:1;height:1px;background:linear-gradient(90deg,var(--bd2),transparent);}
.section-label{font-size:10px;color:var(--muted);letter-spacing:3px;display:flex;align-items:center;gap:10px;margin-bottom:14px;}
.section-label::after{content:'';flex:1;height:1px;background:var(--border);}
/* stat cards */
.stat-grid{display:grid;gap:14px;margin-bottom:28px;}
.stat-grid-4{grid-template-columns:repeat(4,1fr);}
.stat-card{background:var(--bg2);border:1px solid var(--border);border-radius:14px;padding:16px 18px;position:relative;overflow:hidden;transition:border-color .2s;}
.stat-card:hover{border-color:var(--bd2);}
.stat-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;}
.stat-card.green::before{background:linear-gradient(90deg,var(--green),transparent);}
.stat-card.blue::before{background:linear-gradient(90deg,var(--cyan),transparent);}
.stat-card.amber::before{background:linear-gradient(90deg,var(--amber),transparent);}
.stat-card.purple::before{background:linear-gradient(90deg,var(--purple),transparent);}
.stat-label{font-size:9px;color:var(--muted);letter-spacing:2.5px;margin-bottom:8px;}
.stat-val{font-family:'Bebas Neue',sans-serif;font-size:36px;letter-spacing:1px;line-height:1;}
.stat-sub{font-size:12px;color:var(--muted2);margin-top:5px;}
/* tabs */
.top-tabs{display:flex;gap:0;border-bottom:1px solid var(--border);margin-bottom:24px;flex-wrap:wrap;}
.top-tab{font-size:11px;letter-spacing:1px;padding:10px 18px;cursor:pointer;color:var(--muted);border:none;background:none;border-bottom:2px solid transparent;transition:all .15s;display:flex;align-items:center;gap:6px;font-family:'Share Tech Mono',monospace;}
.top-tab:hover{color:var(--text);}.top-tab.active{color:var(--accent);border-bottom-color:var(--accent);}
.top-panel{display:none;}.top-panel.active{display:block;}
.stab-bar{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:16px;align-items:center;}
.stab{font-size:10px;letter-spacing:1px;padding:5px 14px;cursor:pointer;color:var(--muted2);border:1px solid var(--border);background:var(--bg2);border-radius:20px;transition:all .15s;display:flex;align-items:center;gap:5px;font-family:'Share Tech Mono',monospace;}
.stab:hover{color:var(--text);border-color:var(--bd2);}.stab.active{color:var(--accent);border-color:var(--accent);background:rgba(200,255,0,.06);}
.stab-panel{display:none;}.stab-panel.active{display:block;}
.count-badge{background:var(--bg3);border-radius:10px;padding:1px 6px;font-size:9px;color:var(--muted2);}
/* ticket cards */
.ticket-card{background:var(--bg2);border:1px solid var(--border);border-radius:14px;margin-bottom:12px;overflow:hidden;transition:transform .2s,box-shadow .2s;}
.ticket-card:hover{transform:translateY(-2px);box-shadow:0 8px 32px rgba(200,255,0,.08);border-color:var(--bd2);}
.ticket-header{display:flex;align-items:center;gap:12px;padding:11px 16px;background:var(--bg3);border-bottom:1px solid var(--border);border-left:4px solid var(--accent);flex-wrap:wrap;}
.ticket-num{font-family:'Bebas Neue',sans-serif;font-size:20px;color:var(--accent);min-width:32px;}
.ticket-desc{font-size:11px;letter-spacing:1px;color:var(--text);flex:1;}
.ticket-pills{display:flex;gap:5px;flex-wrap:wrap;}
.pill{font-size:10px;padding:2px 9px;border-radius:20px;letter-spacing:.5px;}
.pill-green{background:rgba(57,255,110,.12);color:var(--green);border:1px solid rgba(57,255,110,.25);}
.pill-blue{background:rgba(0,229,255,.12);color:var(--cyan);border:1px solid rgba(0,229,255,.25);}
.pill-amber{background:rgba(240,165,0,.12);color:var(--amber);border:1px solid rgba(240,165,0,.25);}
.pill-purple{background:rgba(167,139,250,.12);color:var(--purple);border:1px solid rgba(167,139,250,.25);}
/* tables */
.table-wrap{background:var(--bg2);border:1px solid var(--border);border-radius:14px;overflow:hidden;margin-bottom:16px;}
.ticket-card .table-wrap{border:none;border-radius:0;margin-bottom:0;}
.scrollx{overflow-x:auto;}
table{width:100%;border-collapse:collapse;font-size:13px;}
th{font-size:10px;letter-spacing:.08em;color:var(--accent);padding:9px 12px;text-align:left;background:rgba(200,255,0,.04);border-bottom:1px solid var(--border);white-space:nowrap;font-family:'Bebas Neue',sans-serif;}
th.right{text-align:right;}
td{padding:9px 12px;border-bottom:1px solid rgba(255,255,255,.04);vertical-align:middle;}
tr:last-child td{border-bottom:none;}
tr:hover td{background:rgba(200,255,0,.02);}
td.right{text-align:right;}td.mono{font-family:'Share Tech Mono',monospace;font-size:12px;}
td.muted{color:var(--muted2);}td.small{font-size:11px;}
/* rate bar */
.rbar{display:flex;align-items:center;gap:8px;min-width:90px;}
.rbar-bg{flex:1;height:5px;background:rgba(255,255,255,.06);border-radius:3px;overflow:hidden;}
.rbar-fill{height:100%;border-radius:3px;}
.rbar-num{font-size:11px;width:36px;text-align:right;flex-shrink:0;}
/* chips */
.chip{display:inline-block;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:700;letter-spacing:.5px;white-space:nowrap;font-family:'Share Tech Mono',monospace;}
.chip-a{background:rgba(57,255,110,.12);color:var(--green);border:1px solid rgba(57,255,110,.25);}
.chip-b{background:rgba(0,229,255,.12);color:var(--cyan);border:1px solid rgba(0,229,255,.25);}
.chip-c{background:rgba(240,165,0,.12);color:var(--amber);border:1px solid rgba(240,165,0,.25);}
.chip-d{background:rgba(153,153,153,.1);color:#aaa;border:1px solid rgba(153,153,153,.2);}
.chip-goblin{background:rgba(167,139,250,.12);color:var(--purple);border:1px solid rgba(167,139,250,.25);}
.chip-demon{background:rgba(255,77,77,.12);color:#ff8080;border:1px solid rgba(255,77,77,.25);}
.chip-std{background:rgba(0,229,255,.12);color:var(--cyan);border:1px solid rgba(0,229,255,.25);}
.chip-over{background:rgba(57,255,110,.12);color:var(--green);border:1px solid rgba(57,255,110,.25);}
.chip-under{background:rgba(240,165,0,.12);color:var(--amber);border:1px solid rgba(240,165,0,.25);}
.pos{color:var(--green);font-weight:700;}.neg{color:var(--red);font-weight:700;}.muted{color:var(--muted2);}
.sub{font-size:11px;color:var(--muted2);margin-top:2px;}
.alert{border-radius:12px;padding:14px 18px;margin-bottom:20px;border:1px solid;font-size:13px;line-height:1.6;}
.alert-amber{background:rgba(240,165,0,.06);border-color:rgba(240,165,0,.25);}
.footer{font-size:10px;color:var(--muted);text-align:center;margin-top:40px;letter-spacing:1.5px;}
@media(max-width:768px){.stat-grid-4{grid-template-columns:repeat(2,1fr)}.snav-name{display:none}}
"""


# ── Full HTML ─────────────────────────────────────────────────────────────────
def build_html(xlsx_path: Path) -> str:
    print(f"  Loading: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    display_date, iso_date = extract_date(xlsx_path)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Main slate rows
    full_slate = read_flat_sheet(wb["Full Slate"]) if "Full Slate" in wb.sheetnames else []
    nba_slate  = read_flat_sheet(wb["NBA Slate"])  if "NBA Slate"  in wb.sheetnames else []
    cbb_slate  = read_flat_sheet(wb["CBB Slate"])  if "CBB Slate"  in wb.sheetnames else []
    main_slate = full_slate or nba_slate
    print(f"  Slate rows: {len(main_slate)}")

    # Ticket sheets
    nba_sheets   = [s for s in wb.sheetnames if re.match(r"NBA (Goblin|Standard|Demon|Mix)", s)]
    cbb_sheets   = [s for s in wb.sheetnames if re.match(r"CBB (Goblin|Standard|Demon|Mix)", s)]
    # COMBO Mix sheets + MIX Standard/MIX Goblin cross-sport sheets
    combo_sheets = [s for s in wb.sheetnames if re.match(r"COMBO ", s)]
    mix_sheets   = [s for s in wb.sheetnames if re.match(r"MIX ", s)]

    # Count all tickets for KPI
    total_tickets = 0
    for sheets in [nba_sheets, cbb_sheets, combo_sheets, mix_sheets]:
        for sname in sheets:
            if sname in wb.sheetnames:
                total_tickets += len(read_ticket_sheet(wb[sname]))
    print(f"  Total tickets: {total_tickets}")

    kpi = build_kpi(main_slate, total_tickets, display_date)

    # Slate tables
    t_all = build_slate_table(main_slate)
    t_nba = build_slate_table([r for r in main_slate if str(r.get("Sport","")).upper()=="NBA"])
    t_cbb = build_slate_table(cbb_slate or [r for r in main_slate if str(r.get("Sport","")).upper()=="CBB"])

    # Ticket section builder
    def ticket_section(label: str, color: str, sheets: list[str], prefix: str) -> str:
        btns, panels = build_tab_section(wb, sheets, prefix)
        if not btns: return ""
        return f"""<div id="tab-{prefix}" class="top-panel">
  <div class="sport-header" style="margin-top:4px">
    <div class="sport-label" style="color:{color}">{label}</div>
    <div class="sport-header-line"></div>
  </div>
  <div class="top-tabs" style="margin-bottom:20px">{btns}</div>
  {panels}
</div>"""

    nba_section   = ticket_section("NBA TICKETS",         "var(--accent)", nba_sheets,   "nba")
    cbb_section   = ticket_section("CBB TICKETS",         "var(--cyan)",   cbb_sheets,   "cbb")
    combo_section = ticket_section("COMBO MIX TICKETS",  "var(--green)",  combo_sheets, "combo")
    mix_section   = ticket_section("CROSS-SPORT TICKETS", "var(--amber)", mix_sheets,  "mix")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Tickets — {display_date} — NBA Pipelines</title>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&display=swap" rel="stylesheet"/>
<style>{CSS}</style>
</head>
<body>

<nav class="snav">
  <div class="nav-accent"></div>
  <a class="snav-brand" href="/">
    <div class="brain-wrap">
      <div class="brain-slate"></div>
      <div class="bpr"></div><div class="bpr"></div><div class="bpr"></div><div class="bpr"></div>
      <div class="bsp lg"></div><div class="bsp md"></div><div class="bsp sm"></div><div class="bsp lg"></div>
      <div class="bsp cy"></div><div class="bsp md"></div><div class="bsp sm"></div><div class="bsp lg"></div>
      <div class="bsp cy"></div><div class="bsp md"></div><div class="bsp lg"></div><div class="bsp sm"></div>
      <svg class="brain-svg-el" viewBox="0 0 50 50" fill="none" xmlns="http://www.w3.org/2000/svg">
        <defs>
          <linearGradient id="lgL2" x1="6" y1="6" x2="25" y2="44" gradientUnits="userSpaceOnUse"><stop offset="0%" stop-color="#c8ff00" stop-opacity="0.35"/><stop offset="100%" stop-color="#c8ff00" stop-opacity="0.06"/></linearGradient>
          <linearGradient id="lgR2" x1="44" y1="6" x2="25" y2="44" gradientUnits="userSpaceOnUse"><stop offset="0%" stop-color="#00e5ff" stop-opacity="0.35"/><stop offset="100%" stop-color="#00e5ff" stop-opacity="0.06"/></linearGradient>
          <filter id="ng2"><feGaussianBlur stdDeviation="0.8" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>
        </defs>
        <path d="M25 7 C22 7 18 8 15 10 C12 12 10 15 9 18 C8 21 8.5 24 9 26 C7.5 27.5 7 30 7.5 32.5 C8 35 10 37.5 13 39 C15 40 17 40 19 39.5 C20.5 39 22 38 23 37 L23 9 C23.5 8 24 7.5 25 7Z" fill="url(#lgL2)" stroke="#c8ff00" stroke-width="0.9"/>
        <path d="M9 19 C10.5 18 12 19 13.5 18" stroke="#c8ff00" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M8.5 23 C10 22 12 23 13.5 22" stroke="#c8ff00" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M8 27 C9.5 26 11.5 27 13 26.5" stroke="#c8ff00" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M8.5 31 C10 30.5 12 31 13.5 30.5" stroke="#c8ff00" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M25 7 C28 7 32 8 35 10 C38 12 40 15 41 18 C42 21 41.5 24 41 26 C42.5 27.5 43 30 42.5 32.5 C42 35 40 37.5 37 39 C35 40 33 40 31 39.5 C29.5 39 28 38 27 37 L27 9 C26.5 8 26 7.5 25 7Z" fill="url(#lgR2)" stroke="#00e5ff" stroke-width="0.9"/>
        <path d="M41 19 C39.5 18 38 19 36.5 18" stroke="#00e5ff" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M41.5 23 C40 22 38 23 36.5 22" stroke="#00e5ff" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M42 27 C40.5 26 38.5 27 37 26.5" stroke="#00e5ff" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M41.5 31 C40 30.5 38 31 36.5 30.5" stroke="#00e5ff" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.7"/>
        <line x1="25" y1="8" x2="25" y2="38" stroke="#ffffff22" stroke-width="0.6" stroke-dasharray="2.5,2"/>
        <circle cx="13" cy="16" r="1.4" fill="#c8ff00" filter="url(#ng2)"><animate attributeName="opacity" values="1;0.15;1" dur="1.7s" repeatCount="indefinite"/></circle>
        <circle cx="11" cy="22" r="1.2" fill="#c8ff00" filter="url(#ng2)"><animate attributeName="opacity" values="0.8;0.1;0.8" dur="2.2s" repeatCount="indefinite" begin="0.3s"/></circle>
        <circle cx="12" cy="28.5" r="1.3" fill="#c8ff00" filter="url(#ng2)"><animate attributeName="opacity" values="0.9;0.2;0.9" dur="1.9s" repeatCount="indefinite" begin="0.6s"/></circle>
        <circle cx="37" cy="16" r="1.4" fill="#00e5ff" filter="url(#ng2)"><animate attributeName="opacity" values="1;0.15;1" dur="2.0s" repeatCount="indefinite" begin="0.2s"/></circle>
        <circle cx="39" cy="22" r="1.2" fill="#00e5ff" filter="url(#ng2)"><animate attributeName="opacity" values="0.8;0.1;0.8" dur="1.8s" repeatCount="indefinite" begin="0.5s"/></circle>
        <circle cx="38" cy="28.5" r="1.3" fill="#00e5ff" filter="url(#ng2)"><animate attributeName="opacity" values="0.9;0.2;0.9" dur="2.3s" repeatCount="indefinite" begin="0.8s"/></circle>
        <line x1="13" y1="16" x2="37" y2="16" stroke="#c8ff0030" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.9;0.2" dur="1.7s" repeatCount="indefinite"/></line>
        <line x1="11" y1="22" x2="39" y2="22" stroke="#00e5ff30" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.9;0.2" dur="2.2s" repeatCount="indefinite" begin="0.4s"/></line>
        <line x1="12" y1="28.5" x2="38" y2="28.5" stroke="#c8ff0030" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.8;0.2" dur="1.9s" repeatCount="indefinite" begin="0.7s"/></line>
      </svg>
    </div>
    <span class="snav-name">Slate<span>IQ</span></span>
  </a>
  <ul class="snav-links">
    <li><a href="/"><span class="ni ni-ctrl">&#x26A1;</span> Control</a></li>
    <li><a href="/tickets" class="active"><span class="ni ni-tick">&#x1F39F;</span> Tickets</a></li>
    <li><a href="/slate"><span class="ni ni-slate">&#x1F4CB;</span> Slate</a></li>
    <li><a href="/payout"><span class="ni ni-pay">&#x1F4B0;</span> Payout</a></li>
    <li><a href="/grades"><span class="ni ni-grade">&#x270F;</span> Grades</a></li>
  </ul>
  <div class="snav-right">
    <div class="live-pill"><div class="live-dot"></div>SLATEIQ &nbsp;·&nbsp; LIVE</div>
  </div>
</nav>

<header>
  <div class="logo">
    <div>
      <div class="logo-title">SLATE VIEWER</div>
      <div class="logo-sub">COMBINED NBA + CBB SLATE</div>
    </div>
  </div>
  <div class="date-badge">&#x1F4C5; {display_date}</div>
</header>

<div class="main">

  {kpi}

  <div class="top-tabs" id="main-tabs">
    <button class="top-tab active" onclick="switchTop(event,'tab-slate')">&#x1F4CB; Full Slate</button>
    <button class="top-tab" onclick="switchTop(event,'tab-nba')">&#x1F3C0; NBA Tickets</button>
    <button class="top-tab" onclick="switchTop(event,'tab-cbb')">&#x1F3EB; CBB Tickets</button>
    <button class="top-tab" onclick="switchTop(event,'tab-combo')">&#x1F500; Combo Mix</button>
    <button class="top-tab" onclick="switchTop(event,'tab-mix')">&#x1F91D; Cross-Sport</button>
  </div>

  <div id="tab-slate" class="top-panel active">
    <div class="top-tabs">
      <button class="top-tab active" onclick="switchTop(event,'sl-all')">All ({len(main_slate):,})</button>
      <button class="top-tab" onclick="switchTop(event,'sl-nba')">&#x1F3C0; NBA</button>
      <button class="top-tab" onclick="switchTop(event,'sl-cbb')">&#x1F3EB; CBB</button>
    </div>
    <div id="sl-all" class="top-panel active">{t_all}</div>
    <div id="sl-nba" class="top-panel">{t_nba}</div>
    <div id="sl-cbb" class="top-panel">{t_cbb}</div>
  </div>

  {nba_section}
  {cbb_section}
  {combo_section}
  {mix_section}

  <div class="footer">GENERATED {generated} &nbsp;·&nbsp; {h(xlsx_path.name)}</div>
</div>

<script>
function switchTop(e, id) {{
  const bar = e.target.closest('.top-tabs');
  bar.querySelectorAll('.top-tab').forEach(t => t.classList.remove('active'));
  e.target.classList.add('active');
  const container = bar.parentElement;
  container.querySelectorAll(':scope > .top-panel').forEach(p => p.classList.remove('active'));
  const el = document.getElementById(id);
  if (el) el.classList.add('active');
}}
function switchStab(e, id) {{
  const bar = e.target.closest('.stab-bar');
  bar.querySelectorAll('.stab').forEach(t => t.classList.remove('active'));
  e.target.classList.add('active');
  const section = bar.parentElement;
  section.querySelectorAll(':scope > .stab-panel').forEach(p => p.classList.remove('active'));
  const el = document.getElementById(id);
  if (el) el.classList.add('active');
}}
</script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=str)
    parser.add_argument("--date",  type=str)
    parser.add_argument("--out",   type=str)
    args = parser.parse_args()

    if args.input:
        xlsx_path = Path(args.input).resolve()
        if not xlsx_path.exists():
            print(f"ERROR: Not found: {xlsx_path}"); sys.exit(1)
    else:
        xlsx_path = find_latest_tickets(args.date)
        print(f"  Auto-detected: {xlsx_path}")

    html = build_html(xlsx_path)
    out  = Path(args.out).resolve() if args.out else UI_DOCS_DIR / "tickets_latest.html"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    print(f"  Saved  -> {out}  ({len(html):,} bytes)")
    print("  Done.")

if __name__ == "__main__":
    main()
