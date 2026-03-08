#!/usr/bin/env python3
"""
combined_slate_tickets.py

Combined NBA + CBB + NHL + Soccer Slate & Ticket Generator
Merges NBA (step8_all_direction_clean.xlsx) and CBB (step6_ranked_cbb.xlsx ELIGIBLE)
Outputs:
  - combined_slate_tickets_YYYY-MM-DD.xlsx
  - tickets_latest.json / tickets_latest.html (web-friendly, static)
  - docs/tickets_latest.json / docs/tickets_latest.html (for GitHub Pages /docs)

Sheets: SUMMARY, Full Slate, NBA Slate, CBB Slate,
        NBA 3/4/5/6-Leg tickets (Goblin/Standard/Demon/Mix),
        CBB 3/4/5/6-Leg tickets, Combined 3/4/5/6-Leg tickets,
        Cross-sport Standard Mix, Cross-sport Goblin Mix

NEW (Web):
- Adds player headshot thumbnails when an ID is available:
    NBA: uses nba_player_id (if present) -> cdn.nba.com headshot
    CBB: uses espn_player_id (if present) -> espncdn headshot
  If no ID exists, it falls back to a simple initials avatar.
- JSON includes image_url per leg.
- More helpful file-path resolution (tries script dir + recursive search if file not found)

HOTFIX:
- Fixes crash when CBB "direction" becomes a DataFrame due to duplicate columns.
  We de-duplicate columns BEFORE touching df["direction"].str.upper().
"""

from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime, timezone
from typing import Optional, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Color palette ─────────────────────────────────────────────────────────────
C = {
    "hdr": "1C1C1C",
    "hdr_nba": "1A5276",
    "hdr_cbb": "1E8449",
    "hdr_mix": "6C3483",
    "hdr_sum": "117A65",
    "hit": "27AE60",
    "miss": "E74C3C",
    "push": "F39C12",
    "tier_a": "D5F5E3",
    "tier_b": "D6EAF8",
    "tier_c": "FEF9E7",
    "tier_d": "FDEDEC",
    "goblin": "E8D5F5",
    "demon": "FDEDEC",
    "standard": "F2F3F4",
    "over": "D6EAF8",
    "under": "FDEBD0",
    "alt": "F2F3F4",
    "white": "FFFFFF",
    "nba": "EBF5FB",
    "cbb": "EAFAF1",
    "nhl": "EBF4FD",
    "hdr_nhl": "1A3A5C",
    "hdr_soccer": "1A5C2E",
    "soccer": "EAFBF1",
    "mix": "F5EEF8",
    "gold": "F9E79F",
}

PAYOUT = {
    2: {"power": 3.0,  "flex": 3.0},
    3: {"power": 6.0,  "flex": 3.0},   # Updated: power=6x, flex=3x
    4: {"power": 10.0, "flex": 6.0},
    5: {"power": 20.0, "flex": 10.0},
    6: {"power": 37.5, "flex": 25.0},  # Updated: power=37.5x, flex=25x
}

# ── Per-leg count quality thresholds (used by smart ticket builder) ───────────
# Min hit rate required per leg depending on ticket length
# Longer tickets need higher floor because win prob = product of all hit rates
LEG_MIN_HIT_RATE = {
    3: 0.58,   # 3-leg: 0.58^3 = 19.5% win prob floor
    4: 0.62,   # 4-leg: 0.62^4 = 14.8% win prob floor
    5: 0.65,   # 5-leg: 0.65^5 = 11.6% win prob floor
    6: 0.68,   # 6-leg: 0.68^6 = 9.8% win prob floor
}

# Min tier per leg count for Power mode tickets
POWER_MIN_TIER = {
    3: ["A", "B", "C"],   # 3-leg power: Tier A/B/C ok
    4: ["A", "B", "C"],   # 4-leg power: Tier A/B/C ok
    5: ["A", "B"],         # 5-leg power: Tier A/B only
    6: ["A", "B"],         # 6-leg power: Tier A/B only
}

# Demon legs are only allowed in Flex-mode analysis (too low hit rate for Power)
# This is enforced in build_tickets_smart() below


# ── Excel style helpers ───────────────────────────────────────────────────────
def side(color: str = "CCCCCC") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def hc(ws, r, c, v, bg=None, fc="FFFFFF", bold=True, sz=9, align="center"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=fc, name="Arial", size=sz)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = side()
    return cell


def dc(ws, r, c, v, bg=None, bold=False, sz=9, align="center", fc="000000", fmt=None):
    if v is pd.NA or (isinstance(v, float) and np.isnan(v)) or v is None:
        v = ""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, name="Arial", size=sz, color=fc)
    cell.fill = PatternFill("solid", start_color=bg or C["white"])
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = side()
    if fmt:
        cell.number_format = fmt
    return cell


def sw(ws, widths: List[int]):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def tier_bg(t) -> str:
    return {"A": C["tier_a"], "B": C["tier_b"], "C": C["tier_c"], "D": C["tier_d"]}.get(
        str(t).upper(), C["white"]
    )


def pt_bg(pt) -> str:
    return {"Goblin": C["goblin"], "Demon": C["demon"], "Standard": C["standard"]}.get(pt, C["white"])


def hr_bg(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "DDDDDD"
    if v >= 0.65:
        return C["hit"]
    if v >= 0.50:
        return C["push"]
    return C["miss"]


def pct_cell(ws, r, c, val):
    nan = val is None or (isinstance(val, float) and np.isnan(val))
    bg = hr_bg(val) if not nan else "DDDDDD"
    cell = dc(ws, r, c, val if not nan else "", bg=bg, bold=True)
    if not nan:
        cell.number_format = "0%"
        cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
    return cell


def win_prob(hit_rates, _n_legs: int) -> float:
    vals = []
    for h in hit_rates:
        try:
            if h is None:
                continue
            if isinstance(h, float) and np.isnan(h):
                continue
            vals.append(float(h))
        except Exception:
            continue
    if not vals:
        return 0.0
    return float(np.prod(vals))


# ──────────────────────────────────────────────────────────────────────────────
# Path resolution helpers (fixes the “file not found” headaches)
# ──────────────────────────────────────────────────────────────────────────────
def _norm_path(p: str) -> str:
    p = (p or "").strip().strip('"').strip("'")
    p = os.path.expanduser(p)
    return os.path.abspath(p)


def _find_first_by_filename(root_dir: str, filename: str) -> Optional[str]:
    try:
        for base, _dirs, files in os.walk(root_dir):
            for f in files:
                if f.lower() == filename.lower():
                    return os.path.join(base, f)
    except Exception:
        return None
    return None


def resolve_input_path(path: str, fallback_filename: Optional[str] = None) -> str:
    """
    Tries:
    1) exact path as provided
    2) relative to script directory
    3) recursive search from script directory by filename (case-insensitive)
    """
    if not path:
        raise FileNotFoundError("Empty input path.")

    raw = path.strip().strip('"').strip("'")
    p = _norm_path(raw)
    if os.path.exists(p):
        return p

    script_dir = os.path.dirname(os.path.abspath(__file__))
    p2 = os.path.abspath(os.path.join(script_dir, raw))
    if os.path.exists(p2):
        return p2

    filename = fallback_filename or os.path.basename(raw)
    found = _find_first_by_filename(script_dir, filename)
    if found and os.path.exists(found):
        return os.path.abspath(found)

    raise FileNotFoundError(
        f"Could not find file: {path}\nTried:\n- {p}\n- {p2}\n- recursive search for: {filename}"
    )


# ──────────────────────────────────────────────────────────────────────────────
# Web outputs (static HTML + JSON) + player images
# ──────────────────────────────────────────────────────────────────────────────
def _safe_float(x):
    try:
        if x is None:
            return None
        if isinstance(x, float) and np.isnan(x):
            return None
        return float(x)
    except Exception:
        return None


def _clean_id(x) -> str:
    """Return a clean integer-like string for IDs, or ''."""
    if x is None:
        return ""
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return ""
    # handle 1628368.0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    if re.fullmatch(r"\d+", s):
        return s
    return ""

def attach_standard_refs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds Standard sibling references to every row (Standard/Goblin/Demon):
      - standard_line
      - standard_edge
      - standard_projection
      - line_discount_vs_standard (direction-aware)

    Matching key uses: sport, player, team, opp, prop_type, game_time
    Bulletproof: supports 'Projection' vs 'projection' and missing cols.
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    # --- unify projection column name (Projection -> projection) ---
    if "projection" not in out.columns and "Projection" in out.columns:
        out["projection"] = out["Projection"]

    # Ensure required columns exist
    for c in [
        "sport", "player", "team", "opp", "prop_type", "pick_type",
        "direction", "line", "edge", "projection", "game_time"
    ]:
        if c not in out.columns:
            out[c] = pd.NA

    key_cols = ["sport", "player", "team", "opp", "prop_type", "game_time"]

    # Build Standard reference table
    std = out[out["pick_type"].astype(str).str.lower() == "standard"].copy()
    if std.empty:
        out["standard_line"] = pd.NA
        out["standard_edge"] = pd.NA
        out["standard_projection"] = pd.NA
        out["line_discount_vs_standard"] = pd.NA
        return out

    std_ref = (
        std[key_cols + ["line", "edge", "projection"]]
        .rename(columns={
            "line": "standard_line",
            "edge": "standard_edge",
            "projection": "standard_projection",
        })
        .drop_duplicates(subset=key_cols, keep="first")
    )

    out = out.merge(std_ref, on=key_cols, how="left")

    # Direction-aware "discount vs standard"
    def _discount(row):
        try:
            s = row.get("standard_line", pd.NA)
            l = row.get("line", pd.NA)
            if pd.isna(s) or pd.isna(l):
                return pd.NA
            d = str(row.get("direction", "")).upper().strip()
            s = float(s)
            l = float(l)
            if d == "OVER":
                return s - l
            if d == "UNDER":
                return l - s
            return pd.NA
        except Exception:
            return pd.NA

    out["line_discount_vs_standard"] = out.apply(_discount, axis=1)
    return out

def player_initials(name: str) -> str:
    s = (name or "").strip()
    if not s:
        return "?"
    parts = [p for p in re.split(r"\s+", s) if p]
    if len(parts) == 1:
        return parts[0][:1].upper()
    return (parts[0][:1] + parts[-1][:1]).upper()


def compute_image_url(leg: dict) -> Optional[str]:
    """
    NBA:
      needs nba_player_id -> https://cdn.nba.com/headshots/nba/latest/1040x760/<id>.png
    CBB:
      needs espn_player_id -> https://a.espncdn.com/i/headshots/mens-college-basketball/players/full/<id>.png
    """
    sport = (leg.get("sport") or "").upper()
    if sport == "NBA":
        pid = _clean_id(leg.get("nba_player_id"))
        if pid:
            return f"https://cdn.nba.com/headshots/nba/latest/1040x760/{pid}.png"
        return None
    if sport == "CBB":
        eid = _clean_id(leg.get("espn_player_id"))
        if eid:
            return f"https://a.espncdn.com/i/headshots/mens-college-basketball/players/full/{eid}.png"
        return None
    return None


def ticket_groups_to_payload(all_ticket_groups, date_str, thresholds):
    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "date": date_str,
        "filters": thresholds,
        "groups": [],
    }

    for group_name, tickets, _bg in all_ticket_groups:
        if not tickets:
            continue

        group = {
            "group_name": str(group_name),
            "n_legs": int(tickets[0].get("n_legs", 0) or 0),
            "power_payout": _safe_float(tickets[0].get("power_payout")),
            "flex_payout": _safe_float(tickets[0].get("flex_payout")),
            "tickets": [],
        }

        for ti, t in enumerate(tickets, start=1):
            rows = t.get("rows", [])
            slip = {
                "ticket_no": ti,
                "avg_hit_rate": _safe_float(t.get("avg_hit_rate")),
                "avg_rank_score": _safe_float(t.get("avg_rank_score")),
                "est_win_prob": _safe_float(t.get("est_win_prob")),
                "legs": [],
            }

            for row in rows:

                def gv(field):
                    return row.get(field, "") if isinstance(row, dict) else getattr(row, field, "")

                leg = {
                    "sport": str(gv("sport") or ""),
                    "player": str(gv("player") or ""),
                    "team": str(gv("team") or ""),
                    "opp": str(gv("opp") or ""),
                    "prop_type": str(gv("prop_type") or ""),
                    "pick_type": str(gv("pick_type") or ""),
                    "direction": str(gv("direction") or ""),
                    "line": _safe_float(gv("line")),
                    "edge": _safe_float(gv("edge")),
                    "standard_line": _safe_float(gv("standard_line")),
                    "standard_edge": _safe_float(gv("standard_edge")),
                    "standard_projection": _safe_float(gv("standard_projection")),
                    "line_discount_vs_standard": _safe_float(gv("line_discount_vs_standard")),
                    "hit_rate": _safe_float(gv("hit_rate")),
                    "rank_score": _safe_float(gv("rank_score")),
                    "game_time": str(gv("game_time") or ""),
                    "nba_player_id": gv("nba_player_id"),
                    "espn_player_id": gv("espn_player_id"),
                    "min_tier": str(gv("min_tier") or gv("minutes_tier") or gv("Min Tier") or ""),
                    "shot_role": str(gv("shot_role") or gv("Shot Role") or ""),
                    "usage_role": str(gv("usage_role") or gv("Usage Role") or ""),
                }
                leg["image_url"] = compute_image_url(leg)
                leg["initials"] = player_initials(leg.get("player", ""))

                slip["legs"].append(leg)

            group["tickets"].append(slip)

        payload["groups"].append(group)

    return payload


def write_web_outputs(payload, outdir: str):
    os.makedirs(outdir, exist_ok=True)
    json_path = os.path.join(outdir, "tickets_latest.json")
    html_path = os.path.join(outdir, "tickets_latest.html")

    def fmt_pct(x) -> str:
        try:
            if x is None:
                return ""
            return f"{float(x) * 100:.2f}%"
        except Exception:
            return ""

    def fmt_2(x) -> str:
        try:
            if x is None:
                return ""
            return f"{float(x):.2f}"
        except Exception:
            return ""

    def fmt_line(x) -> str:
        # keep lines readable (avoid 5.5000000003)
        try:
            if x is None:
                return ""
            xf = float(x)
            if abs(xf - round(xf)) < 1e-9:
                return str(int(round(xf)))
            return f"{xf:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(x) if x is not None else ""

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    # ── helpers ────────────────────────────────────────────────────────────────
    def hit_color(x) -> str:
        try:
            v = float(x)
            if v >= 0.65:
                return "#39ff6e"
            if v >= 0.50:
                return "#f0a500"
            return "#aaa"
        except Exception:
            return "#aaa"

    def sport_badge(sport: str) -> str:
        s = (sport or "").upper()
        if "NBA" in s:
            return "<span style='background:#c8ff00;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>NBA</span>"
        if "CBB" in s or "NCAA" in s:
            return "<span style='background:#00e5ff;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>CBB</span>"
        if "NHL" in s:
            return "<span style='background:#5bc4f5;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>NHL</span>"
        if "SOCCER" in s:
            return "<span style='background:#57e87d;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>SOC</span>"
        return f"<span style='background:#333;color:#ccc;font-size:11px;padding:2px 7px;border-radius:4px;'>{sport or ''}</span>"

    def badge(val, color="#39ff6e") -> str:
        if not val:
            return "<span style='color:#555;font-size:12px;'>—</span>"
        return f"<span style='background:rgba(0,0,0,.35);color:{color};font-size:12px;padding:2px 8px;border-radius:4px;border:1px solid {color}33;'>{val}</span>"

    def wp_bar(wp) -> str:
        try:
            pct = float(wp) * 100
            w = max(2, min(100, pct))
            col = "#39ff6e" if pct >= 50 else "#f0a500"
            return (
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<div style='flex:1;height:6px;background:#1a1a2e;border-radius:3px;overflow:hidden;'>"
                f"<div style='width:{w:.1f}%;height:100%;background:{col};border-radius:3px;'></div></div>"
                f"<span style='color:{col};font-family:\"Bebas Neue\",sans-serif;font-size:15px;letter-spacing:.05em;'>{pct:.1f}%</span>"
                f"</div>"
            )
        except Exception:
            return ""

    # ── HTML ───────────────────────────────────────────────────────────────────
    filters = payload.get("filters", {})
    gen_at  = payload.get("generated_at", "")
    date    = payload.get("date", "")

    CSS = """
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&display=swap');
*{box-sizing:border-box;margin:0;padding:0;}
:root{
  --bg:#05050f;--surface:#0d0d1f;--card:#111128;--border:#1e1e3a;
  --accent:#c8ff00;--cyan:#00e5ff;--muted:#999;--text:#e8e8f0;
}
body{background:var(--bg);color:var(--text);font-family:'Share Tech Mono',monospace;min-height:100vh;overflow-x:hidden;}

/* scrolling grid */
body::before{
  content:'';position:fixed;inset:0;
  background-image:linear-gradient(rgba(200,255,0,.03) 1px,transparent 1px),
                   linear-gradient(90deg,rgba(200,255,0,.03) 1px,transparent 1px);
  background-size:40px 40px;
  animation:gridScroll 20s linear infinite;pointer-events:none;z-index:0;
}
@keyframes gridScroll{from{background-position:0 0;}to{background-position:0 40px;}}

/* scanlines */
body::after{
  content:'';position:fixed;inset:0;
  background:repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(0,0,0,.18) 2px,rgba(0,0,0,.18) 4px);
  pointer-events:none;z-index:0;
}

#app{position:relative;z-index:1;max-width:1400px;margin:0 auto;padding:24px 20px;}

/* nav */
nav{display:flex;align-items:center;gap:16px;padding:12px 0 24px;border-bottom:1px solid var(--border);flex-wrap:wrap;}
.nav-logo{display:flex;align-items:center;gap:12px;text-decoration:none;}
.brain-wrap{position:relative;width:52px;height:52px;}
.brain-slate{position:absolute;inset:0;border-radius:7px;background:linear-gradient(145deg,#12122a 0%,#080818 100%);border:1px solid #252545;animation:slateBreak 3.5s ease-in-out infinite;}
.brain-slate::before{content:'';position:absolute;inset:0;background:linear-gradient(to bottom right,transparent 47%,#c8ff0044 49%,transparent 51%),linear-gradient(to bottom left,transparent 44%,#c8ff0022 46%,transparent 48%),linear-gradient(to right,transparent 30%,#00e5ff22 31%,transparent 33%);border-radius:7px;animation:crackGlow 3.5s ease-in-out infinite;}
@keyframes slateBreak{0%,100%{transform:scale(1);box-shadow:0 0 0px #c8ff0000;}48%{transform:scale(1.06) rotate(-0.5deg);box-shadow:0 0 24px #c8ff0055;}50%{transform:scale(1.10) rotate(0.5deg);box-shadow:0 0 40px #c8ff0088;}52%{transform:scale(1.06) rotate(-0.3deg);box-shadow:0 0 24px #c8ff0055;}}
@keyframes crackGlow{0%,100%{opacity:0.2;}50%{opacity:1;}}
.brain-svg{position:absolute;inset:3px;animation:brainBreakthrough 3.5s ease-in-out infinite;transform-origin:center bottom;}
@keyframes brainBreakthrough{0%,100%{transform:scale(1) translateY(0px);filter:drop-shadow(0 0 5px #c8ff0099) drop-shadow(0 0 2px #00e5ff66);}48%{transform:scale(1.07) translateY(-1px);filter:drop-shadow(0 0 12px #c8ff00cc) drop-shadow(0 0 8px #00e5ffaa);}50%{transform:scale(1.18) translateY(-3px);filter:drop-shadow(0 0 20px #c8ff00ff) drop-shadow(0 0 14px #00e5ffcc) drop-shadow(0 0 40px #c8ff0044);}52%{transform:scale(1.07) translateY(-1px);filter:drop-shadow(0 0 12px #c8ff00cc) drop-shadow(0 0 8px #00e5ffaa);}}
.brain-pulse-ring{position:absolute;border-radius:9px;border:1.5px solid #c8ff00;opacity:0;animation:brainRingExpand 3.5s ease-out infinite;inset:-3px;}
.brain-pulse-ring:nth-child(2){border-color:#00e5ff;animation-delay:0.15s;}
.brain-pulse-ring:nth-child(3){border-color:#c8ff0088;animation-delay:0.3s;}
.brain-pulse-ring:nth-child(4){border-color:#00e5ff66;animation-delay:0.45s;}
@keyframes brainRingExpand{0%,48%{transform:scale(1);opacity:0;}50%{transform:scale(1);opacity:0.9;}85%{transform:scale(2.4);opacity:0;}100%{transform:scale(2.4);opacity:0;}}
.bspark{position:absolute;border-radius:50%;opacity:0;animation:bsparkFly 3.5s ease-out infinite;}
.bspark.lg{width:4px;height:4px;background:#c8ff00;box-shadow:0 0 6px #c8ff00;}
.bspark.md{width:3px;height:3px;background:#00e5ff;box-shadow:0 0 5px #00e5ff;}
.bspark.sm{width:2px;height:2px;background:#c8ff00cc;}
.bspark.cy{width:2px;height:2px;background:#00e5ffcc;}
.bspark.wh{width:2px;height:2px;background:#ffffffaa;}
.bspark:nth-child(5) {top:10%;left:5%; --tx:-18px;--ty:-16px;animation-delay:0.50s;}
.bspark:nth-child(6) {top:5%; left:40%;--tx:2px;  --ty:-22px;animation-delay:0.52s;}
.bspark:nth-child(7) {top:8%; left:75%;--tx:16px; --ty:-18px;animation-delay:0.54s;}
.bspark:nth-child(8) {top:30%;left:96%;--tx:22px; --ty:-8px; animation-delay:0.51s;}
.bspark:nth-child(9) {top:55%;left:96%;--tx:20px; --ty:8px;  animation-delay:0.53s;}
.bspark:nth-child(10){top:80%;left:86%;--tx:14px; --ty:16px; animation-delay:0.55s;}
.bspark:nth-child(11){top:92%;left:55%;--tx:4px;  --ty:22px; animation-delay:0.50s;}
.bspark:nth-child(12){top:90%;left:25%;--tx:-10px;--ty:20px; animation-delay:0.52s;}
.bspark:nth-child(13){top:72%;left:2%; --tx:-20px;--ty:12px; animation-delay:0.54s;}
.bspark:nth-child(14){top:45%;left:0%; --tx:-22px;--ty:0px;  animation-delay:0.51s;}
.bspark:nth-child(15){top:20%;left:2%; --tx:-18px;--ty:-12px;animation-delay:0.56s;}
.bspark:nth-child(16){top:15%;left:60%;--tx:10px; --ty:-20px;animation-delay:0.53s;}
.bspark:nth-child(17){top:18%;left:20%;--tx:-14px;--ty:-18px;animation-delay:0.65s;}
.bspark:nth-child(18){top:12%;left:55%;--tx:6px;  --ty:-20px;animation-delay:0.67s;}
.bspark:nth-child(19){top:25%;left:88%;--tx:18px; --ty:-14px;animation-delay:0.66s;}
.bspark:nth-child(20){top:60%;left:93%;--tx:18px; --ty:10px; animation-delay:0.68s;}
.bspark:nth-child(21){top:82%;left:70%;--tx:10px; --ty:18px; animation-delay:0.65s;}
.bspark:nth-child(22){top:80%;left:10%;--tx:-16px;--ty:14px; animation-delay:0.67s;}
.bspark:nth-child(23){top:40%;left:2%; --tx:-20px;--ty:4px;  animation-delay:0.69s;}
.bspark:nth-child(24){top:35%;left:93%;--tx:20px; --ty:-4px; animation-delay:0.66s;}
.bspark:nth-child(25){top:3%; left:30%;--tx:-6px; --ty:-24px;animation-delay:0.72s;}
.bspark:nth-child(26){top:3%; left:65%;--tx:8px;  --ty:-24px;animation-delay:0.70s;}
.bspark:nth-child(27){top:50%;left:98%;--tx:24px; --ty:2px;  animation-delay:0.73s;}
.bspark:nth-child(28){top:50%;left:0%; --tx:-24px;--ty:2px;  animation-delay:0.71s;}
@keyframes bsparkFly{0%,47%{opacity:0;transform:translate(0,0) scale(0);}50%{opacity:1;transform:translate(0,0) scale(1);}75%{opacity:0.5;}95%{opacity:0;transform:translate(var(--tx),var(--ty)) scale(0.2);}100%{opacity:0;transform:translate(var(--tx),var(--ty)) scale(0);}}
.brand{font-family:'Bebas Neue',sans-serif;font-size:24px;letter-spacing:.12em;color:var(--accent);line-height:1;}
.brand span{color:var(--cyan);}
.nav-links{display:flex;gap:8px;margin-left:auto;flex-wrap:wrap;}
.nav-links a{color:#aaa;text-decoration:none;font-size:13px;padding:6px 14px;border-radius:6px;border:1px solid transparent;transition:all .2s;}
.nav-links a:hover{color:var(--text);border-color:var(--border);}
.nav-links a.active{color:var(--accent);border-color:var(--accent);background:rgba(200,255,0,.06);}

/* hero */
.hero{margin:28px 0 20px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end;}
.hero h1{font-family:'Bebas Neue',sans-serif;font-size:clamp(32px,5vw,52px);letter-spacing:.08em;line-height:1;color:var(--accent);}
.hero h1 span{color:var(--cyan);}
.meta{color:var(--muted);font-size:12px;margin-top:4px;}

/* filter pill */
.filter-pill{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:10px 16px;font-size:12px;color:#888;margin-bottom:24px;}
.filter-pill strong{color:var(--cyan);}

/* group */
.group{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:20px;margin-bottom:24px;}
.group-hdr{display:flex;align-items:center;gap:12px;margin-bottom:16px;}
.group-title{font-family:'Bebas Neue',sans-serif;font-size:22px;letter-spacing:.08em;color:var(--accent);}
.group-meta{color:var(--muted);font-size:12px;}

/* ticket card */
.ticket{background:var(--card);border:1px solid var(--border);border-radius:10px;margin-bottom:16px;overflow:hidden;transition:transform .2s,box-shadow .2s;}
.ticket:hover{transform:translateY(-2px);box-shadow:0 8px 32px rgba(200,255,0,.08);}
.ticket-accent{width:5px;background:linear-gradient(180deg,var(--accent),var(--cyan));flex-shrink:0;}
.ticket-inner{display:flex;}
.ticket-body{flex:1;padding:14px 16px;}
.ticket-hdr{display:flex;align-items:center;gap:10px;margin-bottom:10px;flex-wrap:wrap;}
.ticket-no{font-family:'Bebas Neue',sans-serif;font-size:18px;letter-spacing:.08em;color:var(--text);}
.kpi-row{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:12px;}
.kpi{display:flex;flex-direction:column;gap:2px;}
.kpi-label{font-size:10px;color:var(--muted);letter-spacing:.08em;text-transform:uppercase;}
.kpi-val{font-family:'Bebas Neue',sans-serif;font-size:20px;letter-spacing:.05em;}

/* table */
table{width:100%;border-collapse:collapse;}
th{background:rgba(200,255,0,.06);color:var(--accent);font-family:'Bebas Neue',sans-serif;font-size:13px;letter-spacing:.08em;padding:8px 10px;text-align:left;border-bottom:1px solid var(--border);}
td{padding:8px 10px;border-bottom:1px solid rgba(255,255,255,.04);font-size:12px;vertical-align:middle;}
tr:last-child td{border-bottom:none;}
tr:hover td{background:rgba(200,255,0,.03);}

/* player cell */
.pwrap{display:flex;gap:8px;align-items:center;}
.avatar{width:30px;height:30px;border-radius:50%;overflow:hidden;border:1px solid var(--border);flex-shrink:0;background:#1a1a2e;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:var(--accent);}
.avatar img{width:100%;height:100%;object-fit:cover;}

/* dir badges */
.dir-over{background:rgba(57,255,110,.15);color:#39ff6e;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;}
.dir-under{background:rgba(240,165,0,.15);color:#f0a500;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;}

/* responsive */
@media(max-width:640px){
  .kpi-row{gap:10px;}
  th,td{padding:6px 6px;font-size:11px;}
}
"""

    html_parts = []
    html_parts.append(f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>SlateIQ — Tickets</title>
<style>{CSS}</style>
</head>
<body>
<div id="app">

<nav>
  <a class="nav-logo" href="index.html">
    <div class="brain-wrap">
      <div class="brain-slate"></div>
      <div class="brain-pulse-ring"></div>
      <div class="brain-pulse-ring"></div>
      <div class="brain-pulse-ring"></div>
      <div class="brain-pulse-ring"></div>
      <div class="bspark lg"></div><div class="bspark md"></div>
      <div class="bspark sm"></div><div class="bspark lg"></div>
      <div class="bspark cy"></div><div class="bspark md"></div>
      <div class="bspark sm"></div><div class="bspark lg"></div>
      <div class="bspark cy"></div><div class="bspark md"></div>
      <div class="bspark lg"></div><div class="bspark sm"></div>
      <div class="bspark cy"></div><div class="bspark sm"></div>
      <div class="bspark md"></div><div class="bspark cy"></div>
      <div class="bspark sm"></div><div class="bspark md"></div>
      <div class="bspark lg"></div><div class="bspark wh"></div>
      <svg class="brain-svg" viewBox="0 0 50 50" fill="none" xmlns="http://www.w3.org/2000/svg">
        <defs>
          <linearGradient id="lgL" x1="6" y1="6" x2="25" y2="44" gradientUnits="userSpaceOnUse">
            <stop offset="0%" stop-color="#c8ff00" stop-opacity="0.35"/>
            <stop offset="60%" stop-color="#c8ff00" stop-opacity="0.12"/>
            <stop offset="100%" stop-color="#c8ff00" stop-opacity="0.06"/>
          </linearGradient>
          <linearGradient id="lgR" x1="44" y1="6" x2="25" y2="44" gradientUnits="userSpaceOnUse">
            <stop offset="0%" stop-color="#00e5ff" stop-opacity="0.35"/>
            <stop offset="60%" stop-color="#00e5ff" stop-opacity="0.12"/>
            <stop offset="100%" stop-color="#00e5ff" stop-opacity="0.06"/>
          </linearGradient>
          <filter id="nglow"><feGaussianBlur stdDeviation="0.8" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>
        </defs>
        <path d="M25 7 C22 7 18 8 15 10 C12 12 10 15 9 18 C8 21 8.5 24 9 26 C7.5 27.5 7 30 7.5 32.5 C8 35 10 37.5 13 39 C15 40 17 40 19 39.5 C20.5 39 22 38 23 37 L23 9 C23.5 8 24 7.5 25 7Z" fill="url(#lgL)" stroke="#c8ff00" stroke-width="0.9"/>
        <path d="M15 10 C13 11 11 13 11 15 C11 17 12.5 18.5 14 18" stroke="#c8ff00" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.6"/>
        <path d="M9 19 C10.5 18 12 19 13.5 18"    stroke="#c8ff00" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M8.5 23 C10 22 12 23 13.5 22"    stroke="#c8ff00" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M8 27 C9.5 26 11.5 27 13 26.5"   stroke="#c8ff00" stroke-width="0.7"  stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M8.5 31 C10 30.5 12 31 13.5 30.5" stroke="#c8ff00" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M10 35 C11.5 34.5 13.5 35 15 34.5" stroke="#c8ff00" stroke-width="0.65" stroke-linecap="round" fill="none" opacity="0.55"/>
        <path d="M16 14 C17 13.5 18.5 14 19.5 13.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M15.5 20 C17 19.5 18.5 20 20 19.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M15 27 C16.5 26.5 18 27 19.5 26.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M15 33 C16.5 32.5 18.5 33 20 32.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M25 7 C28 7 32 8 35 10 C38 12 40 15 41 18 C42 21 41.5 24 41 26 C42.5 27.5 43 30 42.5 32.5 C42 35 40 37.5 37 39 C35 40 33 40 31 39.5 C29.5 39 28 38 27 37 L27 9 C26.5 8 26 7.5 25 7Z" fill="url(#lgR)" stroke="#00e5ff" stroke-width="0.9"/>
        <path d="M35 10 C37 11 39 13 39 15 C39 17 37.5 18.5 36 18" stroke="#00e5ff" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.6"/>
        <path d="M41 19 C39.5 18 38 19 36.5 18"      stroke="#00e5ff" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M41.5 23 C40 22 38 23 36.5 22"      stroke="#00e5ff" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M42 27 C40.5 26 38.5 27 37 26.5"    stroke="#00e5ff" stroke-width="0.7"  stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M41.5 31 C40 30.5 38 31 36.5 30.5"  stroke="#00e5ff" stroke-width="0.7"  stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M40 35 C38.5 34.5 36.5 35 35 34.5"  stroke="#00e5ff" stroke-width="0.65" stroke-linecap="round" fill="none" opacity="0.55"/>
        <path d="M34 14 C33 13.5 31.5 14 30.5 13.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M34.5 20 C33 19.5 31.5 20 30 19.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M35 27 C33.5 26.5 32 27 30.5 26.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M35 33 C33.5 32.5 31.5 33 30 32.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <line x1="25" y1="8" x2="25" y2="38" stroke="#ffffff22" stroke-width="0.6" stroke-dasharray="2.5,2"/>
        <circle cx="13" cy="16" r="1.4" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="1;0.15;1" dur="1.7s" repeatCount="indefinite"/><animate attributeName="r" values="1.4;0.9;1.4" dur="1.7s" repeatCount="indefinite"/></circle>
        <circle cx="11" cy="22" r="1.2" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.8;0.1;0.8" dur="2.2s" repeatCount="indefinite" begin="0.3s"/></circle>
        <circle cx="12" cy="28.5" r="1.3" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.9;0.2;0.9" dur="1.9s" repeatCount="indefinite" begin="0.6s"/></circle>
        <circle cx="15" cy="34.5" r="1.1" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.1;0.7" dur="2.4s" repeatCount="indefinite" begin="0.9s"/></circle>
        <circle cx="19" cy="18" r="1.0" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.6;0.1;0.6" dur="2.0s" repeatCount="indefinite" begin="1.1s"/></circle>
        <circle cx="18" cy="30" r="1.0" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.15;0.7" dur="1.6s" repeatCount="indefinite" begin="0.5s"/></circle>
        <circle cx="37" cy="16" r="1.4" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="1;0.15;1" dur="2.0s" repeatCount="indefinite" begin="0.2s"/><animate attributeName="r" values="1.4;0.9;1.4" dur="2.0s" repeatCount="indefinite" begin="0.2s"/></circle>
        <circle cx="39" cy="22" r="1.2" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.8;0.1;0.8" dur="1.8s" repeatCount="indefinite" begin="0.5s"/></circle>
        <circle cx="38" cy="28.5" r="1.3" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.9;0.2;0.9" dur="2.3s" repeatCount="indefinite" begin="0.8s"/></circle>
        <circle cx="35" cy="34.5" r="1.1" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.1;0.7" dur="1.7s" repeatCount="indefinite" begin="1.0s"/></circle>
        <circle cx="31" cy="18" r="1.0" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.6;0.1;0.6" dur="2.1s" repeatCount="indefinite" begin="1.2s"/></circle>
        <circle cx="32" cy="30" r="1.0" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.15;0.7" dur="1.5s" repeatCount="indefinite" begin="0.4s"/></circle>
        <line x1="13" y1="16" x2="37" y2="16" stroke="#c8ff0030" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.9;0.2" dur="1.7s" repeatCount="indefinite"/></line>
        <line x1="11" y1="22" x2="39" y2="22" stroke="#00e5ff30" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.9;0.2" dur="2.2s" repeatCount="indefinite" begin="0.4s"/></line>
        <line x1="12" y1="28.5" x2="38" y2="28.5" stroke="#c8ff0030" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.8;0.2" dur="1.9s" repeatCount="indefinite" begin="0.7s"/></line>
        <line x1="15" y1="34.5" x2="35" y2="34.5" stroke="#00e5ff30" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.8;0.2" dur="2.4s" repeatCount="indefinite" begin="1.0s"/></line>
        <line x1="13" y1="16" x2="39" y2="22" stroke="#c8ff0018" stroke-width="0.5"><animate attributeName="opacity" values="0;0.6;0" dur="2.5s" repeatCount="indefinite" begin="0.3s"/></line>
        <line x1="11" y1="22" x2="38" y2="28.5" stroke="#00e5ff18" stroke-width="0.5"><animate attributeName="opacity" values="0;0.6;0" dur="2.1s" repeatCount="indefinite" begin="0.8s"/></line>
        <line x1="19" y1="18" x2="31" y2="18" stroke="#ffffff18" stroke-width="0.5"><animate attributeName="opacity" values="0;0.7;0" dur="1.6s" repeatCount="indefinite" begin="1.1s"/></line>
        <line x1="18" y1="30" x2="32" y2="30" stroke="#ffffff18" stroke-width="0.5"><animate attributeName="opacity" values="0;0.7;0" dur="2.0s" repeatCount="indefinite" begin="0.5s"/></line>
        <path d="M22 38 C22 40.5 23 43 25 43 C27 43 28 40.5 28 38" stroke="#c8ff0066" stroke-width="0.9" fill="none" stroke-linecap="round"/>
        <line x1="25" y1="38" x2="25" y2="43" stroke="#00e5ff55" stroke-width="0.7" stroke-dasharray="1.5,1.5"/>
      </svg>
    </div>
    <div><div class="brand">Slate<span>IQ</span></div></div>
  </a>
  <div class="nav-links">
    <a href="index.html">Home</a>
    <a href="indexGrades.html">Grades</a>
    <a href="tickets_latest.html" class="active">Tickets</a>
    <a href="payout_calculator.html">Payouts</a>
  </div>
</nav>

<div class="hero">
  <div>
    <h1>🎟 Latest <span>Tickets</span></h1>
    <div class="meta">Generated: {gen_at} &nbsp;|&nbsp; Date: {date}</div>
  </div>
</div>

<div class="filter-pill">
  Filters &rarr;
  <strong>tiers:</strong> {filters.get('tiers','ALL')} &nbsp;
  <strong>min_hit_rate:</strong> {filters.get('min_hit_rate',0)} &nbsp;
  <strong>min_edge:</strong> {filters.get('min_edge',0)} &nbsp;
  <strong>min_rank:</strong> {filters.get('min_rank','None')} &nbsp;
  <strong>pick_types:</strong> {filters.get('pick_types','ALL')}
  &nbsp;&nbsp;<a href="tickets_latest.json" style="color:var(--cyan);">⬇ JSON</a>
</div>
""")

    for g in payload.get("groups", []):
        html_parts.append(f"""
<div class="group">
  <div class="group-hdr">
    <div class="group-title">{g.get('group_name','Group')}</div>
    <div class="group-meta">Legs: {g.get('n_legs','')} &nbsp;|&nbsp; Power: {g.get('power_payout','')}x &nbsp;|&nbsp; Flex: {g.get('flex_payout','')}x</div>
  </div>
""")
        for t in g.get("tickets", []):
            avg_hr = t.get("avg_hit_rate")
            avg_rs = t.get("avg_rank_score")
            wp     = t.get("est_win_prob")

            try:
                hr_disp = f"{float(avg_hr)*100:.1f}%"
                hr_col  = hit_color(avg_hr)
            except Exception:
                hr_disp, hr_col = "—", "#aaa"

            try:
                rs_disp = f"{float(avg_rs):.2f}"
            except Exception:
                rs_disp = "—"

            html_parts.append(f"""
  <div class="ticket">
    <div class="ticket-inner">
      <div class="ticket-accent"></div>
      <div class="ticket-body">
        <div class="ticket-hdr">
          <div class="ticket-no">Ticket #{t.get('ticket_no','')}</div>
        </div>
        <div class="kpi-row">
          <div class="kpi">
            <div class="kpi-label">Hit Rate</div>
            <div class="kpi-val" style="color:{hr_col};">{hr_disp}</div>
          </div>
          <div class="kpi">
            <div class="kpi-label">Avg Rank</div>
            <div class="kpi-val" style="color:var(--cyan);">{rs_disp}</div>
          </div>
          <div class="kpi" style="flex:1;min-width:140px;">
            <div class="kpi-label">Win Prob</div>
            {wp_bar(wp)}
          </div>
        </div>
        <table>
          <thead><tr>
            <th>#</th><th>Sport</th><th>Player</th><th>Prop</th><th>Line</th>
            <th>Pick</th><th>Min</th><th>Shot</th><th>Usage</th>
            <th>Dir</th><th>Hit%</th><th>Edge</th><th>Rank</th>
          </tr></thead>
          <tbody>
""")
            for i, leg in enumerate(t.get("legs", []), start=1):
                dirv = (leg.get("direction") or "").upper()
                dir_span = (
                    "<span class='dir-over'>OVER</span>"
                    if dirv == "OVER"
                    else f"<span class='dir-under'>{dirv or '—'}</span>"
                )
                img      = leg.get("image_url")
                initials = leg.get("initials") or "?"
                if img:
                    avatar = f"<div class='avatar'><img src='{img}' alt='{initials}' onerror=\"this.style.display='none'\"></div>"
                else:
                    avatar = f"<div class='avatar'>{initials}</div>"

                player_cell = f"<div class='pwrap'>{avatar}<div>{leg.get('player','')}</div></div>"

                hr_val = leg.get("hit_rate")
                hr_fmt = fmt_pct(hr_val) if hr_val is not None else "—"
                hr_c   = hit_color(hr_val) if hr_val is not None else "#aaa"

                min_tier  = badge(leg.get("min_tier") or leg.get("minutes_tier"), "#39ff6e")
                shot_role = badge(leg.get("shot_role"), "#00e5ff")
                usg_role  = badge(leg.get("usage_role"), "#888")

                html_parts.append(
                    f"<tr>"
                    f"<td>{i}</td>"
                    f"<td>{sport_badge(leg.get('sport',''))}</td>"
                    f"<td>{player_cell}</td>"
                    f"<td>{leg.get('prop_type','')}</td>"
                    f"<td style='color:var(--text);'>{fmt_line(leg.get('line'))}</td>"
                    f"<td>{leg.get('pick_type','')}</td>"
                    f"<td>{min_tier}</td>"
                    f"<td>{shot_role}</td>"
                    f"<td>{usg_role}</td>"
                    f"<td>{dir_span}</td>"
                    f"<td style='color:{hr_c};font-weight:600;'>{hr_fmt}</td>"
                    f"<td>{fmt_2(leg.get('edge')) if leg.get('edge') is not None else '—'}</td>"
                    f"<td>{fmt_2(leg.get('rank_score')) if leg.get('rank_score') is not None else '—'}</td>"
                    f"</tr>"
                )

            html_parts.append("""
          </tbody>
        </table>
      </div>
    </div>
  </div>
""")
        html_parts.append("</div>")  # group

    html_parts.append("""
</div><!-- #app -->
</body>
</html>""")

    html_str = "\n".join(html_parts)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_str)

    print(f"✅ Web JSON  -> {json_path}")
    print(f"✅ Web HTML  -> {html_path}")


# ── Load & normalize NBA ───────────────────────────────────────────────────────
def load_nba(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_all_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(
        columns={
            "Tier": "tier",
            "Rank Score": "rank_score",
            "Player": "player",
            "Pos": "pos",
            "Team": "team",
            "Opp": "opp",
            "Game Time": "game_time",
            "Prop": "prop_type",
            "Pick Type": "pick_type",
            "Line": "line",
            "Direction": "direction",
            "Edge": "edge",
            "Projection": "projection",
            "Hit Rate (5g)": "hit_rate",
            "Last 5 Avg": "l5_avg",
            "Season Avg": "season_avg",
            "L5 Over": "l5_over",
            "L5 Under": "l5_under",
            "Def Rank": "def_rank",
            "Def Tier": "def_tier",
            "Min Tier": "min_tier",
            "Shot Role": "shot_role",
            "Usage Role": "usage_role",
            "Void Reason": "void_reason",
            # OPTIONAL if your NBA file has it:
            "nba_player_id": "nba_player_id",
            "NBA Player ID": "nba_player_id",
            "player_id": "nba_player_id",
            "Player ID": "nba_player_id",
        }
    )

    # ✅ IMPORTANT: de-dupe before using any column as Series
    df = df.loc[:, ~df.columns.duplicated()].copy()

    df["sport"] = "NBA"

    if "direction" in df.columns:
        if isinstance(df["direction"], pd.DataFrame):
            df["direction"] = df["direction"].iloc[:, 0]
        df["direction"] = df["direction"].astype(str).str.upper()

    if "tier" in df.columns:
        if isinstance(df["tier"], pd.DataFrame):
            df["tier"] = df["tier"].iloc[:, 0]
        df["tier"] = df["tier"].astype(str).str.upper()

    # Drop voids if present — BUT keep NO_PROJECTION_OR_LINE rows so that
    # shooting-split props (3-PT Made, Two Pointers Made, FT Made/Att, etc.)
    # appear in slate sheets for historical hit-rate tracking.
    # filter_eligible will still exclude them from tickets via tier/hit_rate filters.
    if "void_reason" in df.columns:
        if isinstance(df["void_reason"], pd.DataFrame):
            df["void_reason"] = df["void_reason"].iloc[:, 0]
        void_str = df["void_reason"].astype(str).str.strip()
        keep_mask = (
            df["void_reason"].isna()
            | (void_str == "")
            | (void_str == "NO_PROJECTION_OR_LINE")
        )
        df = df[keep_mask]

    # Clean ID if present
    if "nba_player_id" in df.columns:
        df["nba_player_id"] = df["nba_player_id"].apply(_clean_id)

    return df


# ── Load & normalize CBB ───────────────────────────────────────────────────────
def load_cbb(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step6_ranked_cbb.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = (
        "ELIGIBLE"
        if "ELIGIBLE" in xl.sheet_names
        else ("ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    )
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(
        columns={
            "final_bet_direction": "direction",
            "bet_direction": "direction",
            "opp_team_abbr": "opp",
            "start_time": "game_time",
            "line_hit_rate": "hit_rate",
            "stat_last5_avg": "l5_avg",
            "stat_season_avg": "season_avg",
            "line_hits_over_5": "l5_over",
            "line_hits_under_5": "l5_under",
            "Def Tier": "def_tier",
            "DEF_TIER": "def_tier",
            "Defense Tier": "def_tier",
            "minutes_tier": "min_tier",
            "Min Tier": "min_tier",
            "shot_role": "shot_role",
            "Shot Role": "shot_role",
            "usage_role": "usage_role",
            "Usage Role": "usage_role",
            # OPTIONAL IDs
            "espn_player_id": "espn_player_id",
            "ESPN Player ID": "espn_player_id",
            "player_id": "espn_player_id",
        }
    )

    # ✅ CRITICAL HOTFIX: de-duplicate columns BEFORE df["direction"].str.upper()
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # ✅ If direction is still a DataFrame for any reason, take the first column.
    if "direction" in df.columns and isinstance(df["direction"], pd.DataFrame):
        df["direction"] = df["direction"].iloc[:, 0]

    df["sport"] = "CBB"

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()

    if "tier" in df.columns:
        if isinstance(df["tier"], pd.DataFrame):
            df["tier"] = df["tier"].iloc[:, 0]
        df["tier"] = df["tier"].astype(str).str.upper()

    if "void_reason" in df.columns:
        if isinstance(df["void_reason"], pd.DataFrame):
            df["void_reason"] = df["void_reason"].iloc[:, 0]
        df = df[df["void_reason"].isna() | (df["void_reason"].astype(str).str.strip() == "")]

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    return df


# ── Load & normalize NHL ──────────────────────────────────────────────────────
def load_nhl(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_nhl_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "NHL" if "NHL" in xl.sheet_names else ("ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        "player_name":        "player",
        "position":           "pos",
        "stat_type":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
    })

    # opponent is stored in 'description' column
    if "opp" not in df.columns:
        if "description" in df.columns:
            df["opp"] = df["description"]
        else:
            df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "NHL"

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)
    forced = df["pick_type"].isin(["Goblin", "Demon"])

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
        df.loc[forced, "direction"] = "OVER"
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    for col in ["rank_score", "hit_rate", "line"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "edge" not in df.columns:
        df["edge"] = 0.0

    df = df[df["line"].notna() & (df["line"] > 0)]
    # Convert all pandas NA/NaT to None so openpyxl can handle them
    df = df.astype(object).where(df.notna(), other=None)
    return df



# ── Load & normalize Soccer ───────────────────────────────────────────────────
def load_soccer(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_soccer_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "Soccer" if "Soccer" in xl.sheet_names else (
        "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        # title-case (from step8 clean xlsx)
        "Player":           "player",
        "Tier":             "tier",
        "Rank Score":       "rank_score",
        "Pos":              "pos",
        "Team":             "team",
        "Opp":              "opp",
        "Game Time":        "game_time",
        "Prop":             "prop_type",
        "Pick Type":        "pick_type",
        "Line":             "line",
        "Direction":        "direction",
        "Edge":             "edge",
        "Projection":       "projection",
        "Hit Rate (5g)":    "hit_rate",
        "Last 5 Avg":       "l5_avg",
        "Season Avg":       "season_avg",
        "L5 Over":          "l5_over",
        "L5 Under":         "l5_under",
        "Def Rank":         "def_rank",
        "Def Tier":         "def_tier",
        "Min Tier":         "min_tier",
        "Void Reason":      "void_reason",
        # snake_case fallbacks
        "player_name":        "player",
        "stat_type":          "prop_type",
        "stat_norm":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
        "opponent":           "opp",
    })

    if "opp" not in df.columns:
        df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "Soccer"

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    for col in ["rank_score", "hit_rate", "line"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "edge" not in df.columns:
        df["edge"] = 0.0

    df = df[df["line"].notna() & (df["line"] >= 0)]
    df = df.astype(object).where(df.notna(), other=None)
    return df


# ── Merge to full slate ────────────────────────────────────────────────────────
def build_combined_slate(nba: pd.DataFrame, cbb: pd.DataFrame, nhl: pd.DataFrame = None, soccer: pd.DataFrame = None) -> pd.DataFrame:
    keep = [
        "sport",
        "tier",
        "rank_score",
        "player",
        "team",
        "opp",
        "game_time",
        "prop_type",
        "pick_type",
        "line",
        "direction",
        "edge",
        "projection",
        "hit_rate",
        "l5_avg",
        "season_avg",
        "l5_over",
        "l5_under",
        "def_tier",
        "min_tier",
        "shot_role",
        "usage_role",
        "nba_player_id",
        "espn_player_id",
    ]

    def safe_keep(df, cols):
        df = df.loc[:, ~df.columns.duplicated()].copy()
        return df[[c for c in cols if c in df.columns]].copy()

    frames = [safe_keep(nba, keep), safe_keep(cbb, keep)]
    if nhl is not None and len(nhl) > 0:
        frames.append(safe_keep(nhl, keep))
    if soccer is not None and len(soccer) > 0:
        frames.append(safe_keep(soccer, keep))
    combined = pd.concat(frames, ignore_index=True)

    if "rank_score" in combined.columns:
        combined["rank_score"] = pd.to_numeric(combined["rank_score"], errors="coerce")
    if "hit_rate" in combined.columns:
        combined["hit_rate"] = pd.to_numeric(combined["hit_rate"], errors="coerce")
    if "edge" in combined.columns:
        combined["edge"] = pd.to_numeric(combined["edge"], errors="coerce")

    combined = combined.sort_values("rank_score", ascending=False, na_position="last").reset_index(drop=True)
    return combined


# ── Filter eligible props for tickets ─────────────────────────────────────────
def filter_eligible(df: pd.DataFrame, min_hit_rate=0.55, min_edge=0.0, min_rank=None, tiers=None, pick_types=None):
    mask = pd.Series([True] * len(df), index=df.index)
    # Always exclude NO_PROJECTION_OR_LINE rows from tickets (no line = can't bet)
    if "void_reason" in df.columns:
        void_str = df["void_reason"].astype(str).str.strip()
        mask &= ~(void_str == "NO_PROJECTION_OR_LINE")
    if min_hit_rate > 0 and "hit_rate" in df.columns:
        mask &= df["hit_rate"].fillna(0) >= min_hit_rate
    if min_edge > 0 and "edge" in df.columns:
        mask &= df["edge"].fillna(0) >= min_edge
    if min_rank is not None and "rank_score" in df.columns:
        mask &= df["rank_score"].fillna(-99) >= min_rank
    if tiers and "tier" in df.columns:
        mask &= df["tier"].isin([t.upper() for t in tiers])
    if pick_types and "pick_type" in df.columns:
        mask &= df["pick_type"].isin(pick_types)
    return df[mask].copy()


# ── Build tickets ──────────────────────────────────────────────────────────────
def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets=20, require_mix=False) -> list:
    """
    Smart ticket builder with quality filters per leg count.

    Key improvements vs original:
    - Per-leg min hit rate floor (longer tickets require higher floor)
    - Tier floor per leg count for longer tickets (5/6-leg = Tier A/B only)
    - Demon legs soft-filtered: excluded from 5/6-leg tickets, capped at 1 in 3/4-leg
    - Tickets sorted by est_win_prob DESC then avg_rank_score (optimises for actual wins)
    - require_mix still enforced for cross-sport sheets
    """
    pool = pool.copy().reset_index(drop=True)
    tickets = []

    # ── Per-leg-count quality filters ─────────────────────────────────────────
    min_hr   = LEG_MIN_HIT_RATE.get(n_legs, 0.55)
    ok_tiers = POWER_MIN_TIER.get(n_legs, ["A", "B", "C", "D"])

    # Apply hit rate floor to this pool
    if "hit_rate" in pool.columns:
        pool = pool[pool["hit_rate"].fillna(0) >= min_hr].copy()

    # Apply tier floor for 5/6-leg tickets
    if n_legs >= 5 and "tier" in pool.columns:
        pool = pool[pool["tier"].isin(ok_tiers)].copy()

    # For 5/6-leg tickets: remove Demon legs entirely (38% hit rate kills these)
    if n_legs >= 5 and "pick_type" in pool.columns:
        pool = pool[pool["pick_type"] != "Demon"].copy()

    pool = pool.reset_index(drop=True)

    has_sport_col = "sport" in pool.columns
    sports_available = pool["sport"].dropna().unique().tolist() if has_sport_col else []
    can_mix = require_mix and has_sport_col and len(sports_available) >= 2

    eligible = pool.sort_values("rank_score", ascending=False, na_position="last").reset_index(drop=True)

    for _ in range(max_tickets * 5):
        if len(tickets) >= max_tickets:
            break

        ticket_rows = []
        ticket_players = set()
        sports_in_ticket = set()

        if can_mix:
            for sport in sports_available:
                sport_pool = eligible[eligible["sport"] == sport]
                for _, row in sport_pool.iterrows():
                    player = str(row.get("player", "")).strip().lower()
                    if player and player not in ticket_players:
                        ticket_rows.append(row)
                        ticket_players.add(player)
                        sports_in_ticket.add(sport)
                        break

            for _, row in eligible.iterrows():
                if len(ticket_rows) == n_legs:
                    break
                player = str(row.get("player", "")).strip().lower()
                if player and player not in ticket_players:
                    ticket_rows.append(row)
                    ticket_players.add(player)
                    sports_in_ticket.add(row.get("sport", ""))
        else:
            for _, row in eligible.iterrows():
                if len(ticket_rows) == n_legs:
                    break
                player = str(row.get("player", "")).strip().lower()
                if player and player not in ticket_players:

                    # Cap Demon legs at 1 per 3/4-leg ticket
                    if n_legs <= 4 and "pick_type" in row.index:
                        demon_count = sum(1 for r in ticket_rows
                                          if str(r.get("pick_type", "")) == "Demon")
                        if str(row.get("pick_type", "")) == "Demon" and demon_count >= 1:
                            continue

                    ticket_rows.append(row)
                    ticket_players.add(player)

        if len(ticket_rows) == n_legs:
            if can_mix and len(sports_in_ticket) < 2:
                if len(eligible) > 1:
                    eligible = eligible.iloc[1:].reset_index(drop=True)
                continue

            if can_mix:
                ticket_rows = sorted(
                    ticket_rows,
                    key=lambda r: (str(r.get("sport", "")), -float(r.get("rank_score", 0) or 0)),
                )

            key = frozenset(
                (str(r.get("player", "")) + "|" + str(r.get("prop_type", ""))).strip() for r in ticket_rows
            )

            if key not in [t["key"] for t in tickets]:
                hrs = []
                rss = []
                for r in ticket_rows:
                    hrs.append(float(r.get("hit_rate", 0.5) or 0.5))
                    rss.append(float(r.get("rank_score", 0) or 0))
                avg_hr = float(np.mean(hrs)) if hrs else 0.0
                avg_rs = float(np.mean(rss)) if rss else 0.0
                ep = win_prob(hrs, n_legs)
                pout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})
                tickets.append(
                    {
                        "key": key,
                        "rows": ticket_rows,
                        "avg_hit_rate": avg_hr,
                        "avg_rank_score": avg_rs,
                        "est_win_prob": ep,
                        "power_payout": pout["power"],
                        "flex_payout": pout["flex"],
                        "n_legs": n_legs,
                    }
                )

        if len(eligible) > n_legs:
            eligible = eligible.iloc[1:].reset_index(drop=True)
        else:
            break

    # Sort by win probability first, then rank score — optimises for actual wins
    tickets.sort(key=lambda x: (-x["est_win_prob"], -x["avg_rank_score"]))
    return tickets[:max_tickets]


# ──────────────────────────────────────────────────────────────────────────────
# FINAL web groups (ONLY the ticket sets you want) + ENFORCED Std/Gob mix
# ──────────────────────────────────────────────────────────────────────────────
def build_mixed_picktype_tickets(pool_df: pd.DataFrame, n_legs: int, max_tickets: int, min_standard: int) -> list:
    """
    Deterministic ticket builder that enforces a minimum number of Standard legs,
    while allowing remaining legs from Standard+Goblin pool.

    - Avoids duplicate players
    - Uses rank_score descending
    - Generates variety by sliding a start offset window
    """
    pool_df = pool_df.copy()
    if "rank_score" not in pool_df.columns or "pick_type" not in pool_df.columns:
        return []

    std = pool_df[pool_df["pick_type"] == "Standard"].sort_values("rank_score", ascending=False, na_position="last")
    gob = pool_df[pool_df["pick_type"] == "Goblin"].sort_values("rank_score", ascending=False, na_position="last")

    if len(std) < min_standard:
        return []

    tickets = []
    std_start = 0
    gob_start = 0
    attempts = 0
    max_attempts = max_tickets * 50

    while len(tickets) < max_tickets and attempts < max_attempts:
        attempts += 1
        legs = []
        used_players = set()

        # 1) Required Standards first
        for _, r in std.iloc[std_start:].iterrows():
            if sum(1 for x in legs if str(x.get("pick_type", "")) == "Standard") >= min_standard:
                break
            p = str(r.get("player", "")).strip().lower()
            if p and p not in used_players:
                legs.append(r)
                used_players.add(p)

        # 2) Fill remaining legs by best rank_score from (gob slice + std slice)
        combined_ranked = pd.concat([gob.iloc[gob_start:], std.iloc[std_start:]], ignore_index=True)
        combined_ranked = combined_ranked.sort_values("rank_score", ascending=False, na_position="last")

        for _, r in combined_ranked.iterrows():
            if len(legs) >= n_legs:
                break
            p = str(r.get("player", "")).strip().lower()
            if p and p not in used_players:
                legs.append(r)
                used_players.add(p)

        if len(legs) == n_legs:
            std_count = sum(1 for x in legs if str(x.get("pick_type", "")) == "Standard")
            if std_count >= min_standard:
                hrs = [float(x.get("hit_rate", 0.5) or 0.5) for x in legs]
                rss = [float(x.get("rank_score", 0) or 0) for x in legs]
                avg_hr = float(np.mean(hrs)) if hrs else 0.0
                avg_rs = float(np.mean(rss)) if rss else 0.0
                ep = win_prob(hrs, n_legs)
                pout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})

                key = frozenset((str(x.get("player", "")) + "|" + str(x.get("prop_type", ""))).strip() for x in legs)
                if key not in [t["key"] for t in tickets]:
                    tickets.append(
                        {
                            "key": key,
                            "rows": legs,
                            "avg_hit_rate": avg_hr,
                            "avg_rank_score": avg_rs,
                            "est_win_prob": ep,
                            "power_payout": pout["power"],
                            "flex_payout": pout["flex"],
                            "n_legs": n_legs,
                        }
                    )

        # Slide window to create different combos
        if len(std) > 0:
            std_start = min(std_start + 1, max(len(std) - 1, 0))
        if len(gob) > 0:
            gob_start = min(gob_start + 1, max(len(gob) - 1, 0))

    tickets.sort(key=lambda x: (-x["avg_rank_score"], -x["avg_hit_rate"]))
    return tickets[:max_tickets]


def build_final_web_ticket_groups(nba_pool: pd.DataFrame, cbb_pool: pd.DataFrame,
                                   min_hit_rate=0.70, min_edge=2.0, min_rank=5.0):
    def apply_filters(df):
        mask = pd.Series(True, index=df.index)
        if min_hit_rate > 0 and "hit_rate" in df.columns:
            mask &= df["hit_rate"].fillna(0) >= min_hit_rate
        if min_edge > 0 and "edge" in df.columns:
            mask &= df["edge"].fillna(0) >= min_edge
        if min_rank is not None and "rank_score" in df.columns:
            mask &= df["rank_score"].fillna(-99) >= min_rank
        return df[mask].copy()

    # ── NBA groups ─────────────────────────────────────────────────────────────
    nba_filtered = apply_filters(nba_pool)
    nba_mix = nba_filtered[nba_filtered["pick_type"].isin(["Standard", "Goblin"])].copy()
    nba_std = nba_filtered[nba_filtered["pick_type"].isin(["Standard"])].copy()

    groups = []

    if len(nba_mix) >= 6:
        t6 = build_mixed_picktype_tickets(nba_mix, 6, max_tickets=1, min_standard=2)
        if t6:
            groups.append(("FINAL 6-Leg (NBA Std+Gob)", t6, None))

    if len(nba_mix) >= 5:
        t5 = build_mixed_picktype_tickets(nba_mix, 5, max_tickets=1, min_standard=2)
        if t5:
            groups.append(("FINAL 5-Leg (NBA Std+Gob)", t5, None))

    if len(nba_mix) >= 4:
        t4 = build_mixed_picktype_tickets(nba_mix, 4, max_tickets=1, min_standard=2)
        if t4:
            groups.append(("FINAL 4-Leg (NBA Std+Gob)", t4, None))

    if len(nba_mix) >= 3:
        t3 = build_mixed_picktype_tickets(nba_mix, 3, max_tickets=2, min_standard=1)
        if t3:
            groups.append(("FINAL 3-Leg MIX (NBA Std+Gob)", t3, None))

    if len(nba_std) >= 3:
        groups.append(("FINAL 3-Leg STANDARD ONLY (NBA)", build_tickets(nba_std, 3, max_tickets=1), None))

    # ── CBB groups ─────────────────────────────────────────────────────────────
    if cbb_pool is not None and len(cbb_pool):
        cbb_filtered = apply_filters(cbb_pool)
        cbb_mix = cbb_filtered[cbb_filtered["pick_type"].isin(["Standard", "Goblin"])].copy() \
            if "pick_type" in cbb_filtered.columns else cbb_filtered.copy()
        cbb_std = cbb_filtered[cbb_filtered["pick_type"].isin(["Standard"])].copy() \
            if "pick_type" in cbb_filtered.columns else cbb_filtered.copy()

        if len(cbb_mix) >= 6:
            t6 = build_mixed_picktype_tickets(cbb_mix, 6, max_tickets=1, min_standard=2)
            if t6:
                groups.append(("FINAL 6-Leg (CBB Std+Gob)", t6, None))

        if len(cbb_mix) >= 5:
            t5 = build_mixed_picktype_tickets(cbb_mix, 5, max_tickets=1, min_standard=2)
            if t5:
                groups.append(("FINAL 5-Leg (CBB Std+Gob)", t5, None))

        if len(cbb_mix) >= 4:
            t4 = build_mixed_picktype_tickets(cbb_mix, 4, max_tickets=1, min_standard=2)
            if t4:
                groups.append(("FINAL 4-Leg (CBB Std+Gob)", t4, None))

        if len(cbb_mix) >= 3:
            t3 = build_mixed_picktype_tickets(cbb_mix, 3, max_tickets=2, min_standard=1)
            if t3:
                groups.append(("FINAL 3-Leg MIX (CBB Std+Gob)", t3, None))

        if len(cbb_std) >= 3:
            groups.append(("FINAL 3-Leg STANDARD ONLY (CBB)", build_tickets(cbb_std, 3, max_tickets=1), None))

    # ── NBA + CBB SPORT MIX groups ─────────────────────────────────────────────
    if cbb_pool is not None and len(cbb_pool):
        cbb_filtered = apply_filters(cbb_pool)
        cbb_mix_combo = cbb_filtered[cbb_filtered["pick_type"].isin(["Standard", "Goblin"])].copy() \
            if "pick_type" in cbb_filtered.columns else cbb_filtered.copy()
        combo = pd.concat([nba_mix, cbb_mix_combo], ignore_index=True)

        if len(combo) >= 6:
            t6 = build_mixed_picktype_tickets(combo, 6, max_tickets=1, min_standard=2)
            if t6:
                groups.append(("FINAL 6-Leg SPORT MIX (NBA+CBB)", t6, None))

        if len(combo) >= 5:
            t5 = build_mixed_picktype_tickets(combo, 5, max_tickets=1, min_standard=2)
            if t5:
                groups.append(("FINAL 5-Leg SPORT MIX (NBA+CBB)", t5, None))

        if len(combo) >= 4:
            t4 = build_mixed_picktype_tickets(combo, 4, max_tickets=1, min_standard=2)
            if t4:
                groups.append(("FINAL 4-Leg SPORT MIX (NBA+CBB)", t4, None))

        if len(combo) >= 3:
            t3 = build_mixed_picktype_tickets(combo, 3, max_tickets=2, min_standard=1)
            if t3:
                groups.append(("FINAL 3-Leg SPORT MIX (NBA+CBB)", t3, None))

    return groups


# ── Write slate sheet ──────────────────────────────────────────────────────────
SLATE_COLS = [
    "sport",
    "tier",
    "rank_score",
    "player",
    "team",
    "opp",
    "prop_type",
    "pick_type",
    "line",
    "direction",
    "edge",
    "projection",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "def_tier",
    "min_tier",
    "shot_role",
    "usage_role",
    "game_time",
]
SLATE_WIDTHS = [6, 5, 10, 20, 6, 6, 18, 10, 6, 8, 7, 10, 10, 8, 10, 7, 7, 10, 9, 10, 10, 16]
SLATE_HDRS = [
    "Sport",
    "Tier",
    "Rank Score",
    "Player",
    "Team",
    "Opp",
    "Prop",
    "Pick Type",
    "Line",
    "Dir",
    "Edge",
    "Proj",
    "Hit Rate",
    "L5 Avg",
    "Szn Avg",
    "L5 Over",
    "L5 Under",
    "Def Tier",
    "Min Tier",
    "Shot Role",
    "Usage Role",
    "Game Time",
]


def write_slate_sheet(wb, df, sheet_name, bg_hdr, sport_label=""):
    ws = wb.create_sheet(sheet_name)
    cols = [c for c in SLATE_COLS if c in df.columns]
    hdrs = [SLATE_HDRS[SLATE_COLS.index(c)] for c in cols]
    widths = [SLATE_WIDTHS[SLATE_COLS.index(c)] for c in cols]
    sw(ws, widths)
    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(hdrs, 1):
        hc(ws, 1, ci, h, bg=bg_hdr)
    ws.freeze_panes = "A2"

    for ri, row in enumerate(df[cols].itertuples(index=False), 2):
        bg = C["alt"] if ri % 2 == 0 else C["white"]
        sp = getattr(row, "sport", "")
        if sp == "NBA":
            bg_row = C["nba"] if ri % 2 == 0 else C["white"]
        elif sp == "CBB":
            bg_row = C["cbb"] if ri % 2 == 0 else C["white"]
        else:
            bg_row = bg

        for ci, col in enumerate(cols, 1):
            val = getattr(row, col, "")
            if val is None or (isinstance(val, float) and np.isnan(val)):
                val = ""
            if col == "tier":
                dc(ws, ri, ci, val, bg=tier_bg(val), bold=True, align="center")
            elif col == "pick_type":
                dc(ws, ri, ci, val, bg=pt_bg(val), align="center")
            elif col == "hit_rate":
                pct_cell(ws, ri, ci, val if val != "" else np.nan)
                continue
            elif col == "rank_score":
                dc(ws, ri, ci, round(val, 2) if val != "" else "", bg=bg_row, bold=True, fmt="0.00")
            elif col == "direction":
                dbg = C["over"] if str(val).upper() == "OVER" else C["under"]
                dc(ws, ri, ci, val, bg=dbg, bold=True)
            elif col == "sport":
                sbg = C["hdr_nba"] if val == "NBA" else C["hdr_cbb"]
                dc(ws, ri, ci, val, bg=sbg, bold=True, fc="FFFFFF")
            elif col == "player":
                dc(ws, ri, ci, val, bg=bg_row, align="left", bold=True)
            elif col == "game_time":
                try:
                    if val and val != "":
                        dt = pd.to_datetime(val)
                        dc(ws, ri, ci, dt.strftime("%m/%d %I:%M%p"), bg=bg_row, align="center")
                    else:
                        dc(ws, ri, ci, "", bg=bg_row)
                except Exception:
                    dc(ws, ri, ci, str(val)[:16], bg=bg_row)
                continue
            else:
                dc(ws, ri, ci, val, bg=bg_row, align="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── Write ticket sheet ─────────────────────────────────────────────────────────
TICKET_COLS = [
    "#",
    "player",
    "team",
    "opp",
    "prop_type",
    "pick_type",
    "line",
    "direction",
    "edge",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "rank_score",
    "def_tier",
    "sport",
]
TICKET_HDRS = [
    "#",
    "Player",
    "Team",
    "Opp",
    "Prop",
    "Pick Type",
    "Line",
    "Dir",
    "Edge",
    "Hit Rate",
    "L5 Avg",
    "Szn Avg",
    "L5 Over",
    "L5 Under",
    "Rank Score",
    "Def Tier",
    "Sport",
]
TICKET_W = [4, 20, 6, 6, 18, 10, 6, 6, 7, 9, 8, 9, 7, 8, 11, 10, 6]


def write_ticket_sheet(wb, tickets, sheet_name, bg_hdr, label=""):
    if not tickets:
        return
    ws = wb.create_sheet(sheet_name)
    sw(ws, TICKET_W)
    ws.freeze_panes = "A2"

    ri = 1
    for ti, ticket in enumerate(tickets, 1):
        n = ticket["n_legs"]
        pout = ticket["power_payout"]
        fout = ticket["flex_payout"]
        cost = round(100 / pout, 0) if pout else 0
        avg_hr = ticket["avg_hit_rate"]
        ep = ticket["est_win_prob"]
        avg_rs = ticket["avg_rank_score"]

        banner = (
            f"  Ticket #{ti}  ·  {n}-Leg {label}  ·  "
            f"Power: {pout}x (${cost:.0f} to win $100)  ·  Flex: {fout}x  ·  "
            f"Avg Hit Rate: {avg_hr:.0%}  ·  Est Win Prob: {ep:.0%}  ·  "
            f"Avg Rank Score: {avg_rs:.2f}"
        )
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(TICKET_COLS))
        hc(ws, ri, 1, banner, bg=bg_hdr, sz=9, align="left")
        ws.row_dimensions[ri].height = 16
        ri += 1

        for ci, h in enumerate(TICKET_HDRS, 1):
            hc(ws, ri, ci, h, bg=C["hdr"], sz=8)
        ws.row_dimensions[ri].height = 14
        ri += 1

        for leg_i, row in enumerate(ticket["rows"], 1):
            bg = C["alt"] if leg_i % 2 == 0 else C["white"]
            sp = row.get("sport", "")
            if sp == "NBA":
                bg = C["nba"]
            elif sp == "CBB":
                bg = C["cbb"]

            def gv(field):
                return row.get(field, "")

            dc(ws, ri, 1, leg_i, bg=bg, bold=True, align="center")
            dc(ws, ri, 2, gv("player"), bg=bg, align="left", bold=True)
            dc(ws, ri, 3, gv("team"), bg=bg)
            dc(ws, ri, 4, gv("opp"), bg=bg)
            dc(ws, ri, 5, gv("prop_type"), bg=bg, align="left")
            ptv = gv("pick_type")
            dc(ws, ri, 6, ptv, bg=pt_bg(str(ptv)), align="center")
            dc(ws, ri, 7, gv("line"), bg=bg)
            dirv = str(gv("direction")).upper()
            dc(ws, ri, 8, dirv, bg=C["over"] if dirv == "OVER" else C["under"], bold=True)
            dc(ws, ri, 9, gv("edge"), bg=bg)
            pct_cell(ws, ri, 10, gv("hit_rate") if gv("hit_rate") != "" else np.nan)
            dc(ws, ri, 11, gv("l5_avg"), bg=bg)
            dc(ws, ri, 12, gv("season_avg"), bg=bg)
            dc(ws, ri, 13, gv("l5_over"), bg=bg)
            dc(ws, ri, 14, gv("l5_under"), bg=bg)
            rs = gv("rank_score")
            try:
                rs_out = round(float(rs), 2) if rs != "" and rs is not None else ""
            except Exception:
                rs_out = ""
            dc(ws, ri, 15, rs_out, bg=bg, bold=True)
            dc(ws, ri, 16, gv("def_tier"), bg=bg)
            sv = gv("sport")
            sbg = C["hdr_nba"] if sv == "NBA" else (C["hdr_cbb"] if sv == "CBB" else C["hdr"])
            dc(ws, ri, 17, sv, bg=sbg, bold=True, fc="FFFFFF")
            ws.row_dimensions[ri].height = 14
            ri += 1

        ws.row_dimensions[ri].height = 6
        ri += 1


# ── Write SUMMARY sheet ───────────────────────────────────────────────────────
def write_summary(wb, nba, cbb, combined, all_ticket_groups, date_str, thresholds, nhl=None, soccer=None):
    ws = wb.create_sheet("SUMMARY", 0)
    sw(ws, [28, 14, 10, 10, 10, 10, 10, 12, 18])

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"COMBINED NBA + CBB SLATE  |  {date_str}  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c2 = ws["A2"]
    c2.value = (
        f"Filters: Tier {thresholds.get('tiers','ALL')} | "
        f"Min Hit Rate: {thresholds.get('min_hit_rate',0):.0%} | "
        f"Min Edge: {thresholds.get('min_edge',0)} | "
        f"Min Rank Score: {thresholds.get('min_rank','None')} | "
        f"Pick Types: {thresholds.get('pick_types','ALL')}"
    )
    c2.font = Font(bold=False, name="Arial", size=9, color="000000")
    c2.fill = PatternFill("solid", start_color=C["gold"])
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4

    def sec(r, label, bg):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        hc(ws, r, 1, label, bg=bg, sz=10, align="left")
        ws.row_dimensions[r].height = 20
        return r + 1

    def stat_row(r, label, total, elig, bg=None):
        bg = bg or (C["alt"] if r % 2 == 0 else C["white"])
        dc(ws, r, 1, label, bg=bg, align="left", bold=True)
        dc(ws, r, 2, total, bg=bg)
        dc(ws, r, 3, elig, bg=bg)
        for ci in range(4, 10):
            dc(ws, r, ci, "", bg=bg)
        return r + 1

    row = sec(row, "📊 SLATE OVERVIEW", C["hdr_sum"])
    for ci, h in enumerate(["Category", "Total Props", "Eligible", "", "", "", "", "", ""], 1):
        hc(ws, row, ci, h, bg=C["hdr"], sz=8)
    ws.row_dimensions[row].height = 14
    row += 1

    elig_nba = len(nba[nba.get("tier", "").isin(["A", "B"])]) if "tier" in nba.columns else 0
    elig_cbb = len(cbb[cbb.get("tier", "").isin(["A", "B"])]) if "tier" in cbb.columns else 0
    elig_all = len(combined[combined.get("tier", "").isin(["A", "B"])]) if "tier" in combined.columns else 0
    row = stat_row(row, "NBA Props", len(nba), elig_nba, C["nba"])
    row = stat_row(row, "CBB Props", len(cbb), elig_cbb, C["cbb"])
    if nhl is not None and len(nhl) > 0:
        elig_nhl = len(nhl[nhl.get("tier", "").isin(["A", "B"])]) if "tier" in nhl.columns else 0
        row = stat_row(row, "NHL Props", len(nhl), elig_nhl, C["nhl"])
    if soccer is not None and len(soccer) > 0:
        elig_soc = len(soccer[soccer.get("tier", "").isin(["A", "B"])]) if "tier" in soccer.columns else 0
        row = stat_row(row, "Soccer Props", len(soccer), elig_soc, C["soccer"])
    row = stat_row(row, "Combined Slate", len(combined), elig_all)
    row += 1

    row = sec(row, "🎟️ TICKET SUMMARY", C["hdr_mix"])
    for ci, h in enumerate(
        ["Sheet", "Legs", "Type", "# Tickets", "Avg Hit Rate", "Avg Win Prob", "Avg Rank Score", "Power Payout", "Players"],
        1,
    ):
        hc(ws, row, ci, h, bg=C["hdr"], sz=8)
    ws.row_dimensions[row].height = 14
    row += 1

    for group_name, tickets, bg_row in all_ticket_groups:
        if not tickets:
            continue
        avg_hr = np.mean([t["avg_hit_rate"] for t in tickets])
        avg_wp = np.mean([t["est_win_prob"] for t in tickets])
        avg_rs = np.mean([t["avg_rank_score"] for t in tickets])
        n = tickets[0]["n_legs"]
        pout = tickets[0]["power_payout"]
        bg = bg_row if bg_row else (C["alt"] if row % 2 == 0 else C["white"])
        dc(ws, row, 1, group_name, bg=bg, align="left", bold=True)
        dc(ws, row, 2, n, bg=bg)
        lbl = group_name.split(" ")[0] if group_name else ""
        dc(ws, row, 3, lbl, bg=bg)
        dc(ws, row, 4, len(tickets), bg=bg)
        pct_cell(ws, row, 5, avg_hr)
        pct_cell(ws, row, 6, avg_wp)
        dc(ws, row, 7, round(avg_rs, 2), bg=bg)
        dc(ws, row, 8, f"{pout}x", bg=bg)
        sample = " | ".join(f"{r.get('player','')}" for r in tickets[0]["rows"][:3]) + ("..." if n > 3 else "")
        dc(ws, row, 9, sample, bg=bg, align="left", sz=8)
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--nba", required=True, help="NBA step8_all_direction_clean.xlsx")
    ap.add_argument("--cbb", required=True, help="CBB step6_ranked_cbb.xlsx")
    ap.add_argument("--nhl", default="", help="NHL step8_nhl_direction_clean.xlsx (optional)")
    ap.add_argument("--soccer", default="", help="Soccer step8_soccer_direction_clean.xlsx (optional)")
    ap.add_argument("--output", default="")
    ap.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    ap.add_argument("--tiers", default="A,B,C", help="Comma-separated tiers e.g. A,B")
    ap.add_argument("--min-hit-rate", type=float, default=0.55, dest="min_hit_rate")
    ap.add_argument("--min-edge", type=float, default=0.0, dest="min_edge")
    ap.add_argument("--min-rank", type=float, default=None, dest="min_rank")
    ap.add_argument("--pick-types", default="Goblin,Standard,Demon", dest="pick_types")  # Demon kept for Flex sheets; filtered out of 5/6-leg Power by build_tickets
    ap.add_argument("--max-tickets", type=int, default=20, dest="max_tickets")

    # Web outputs
    ap.add_argument("--write-web", action="store_true", help="Write tickets_latest.html/json for GitHub Pages")
    ap.add_argument("--web-outdir", default=r"..\ui_runner\templates",help="Folder to write tickets_latest.html/json")
    ap.add_argument("--also-root", action="store_true", help="Also write tickets_latest.* in repo root")

    args = ap.parse_args()

    if not args.output:
        args.output = f"combined_slate_tickets_{args.date}.xlsx"

    tiers = [t.strip() for t in args.tiers.split(",") if t.strip()]
    pick_types = [p.strip() for p in args.pick_types.split(",") if p.strip()]
    thresholds = {
        "tiers": args.tiers,
        "min_hit_rate": args.min_hit_rate,
        "min_edge": args.min_edge,
        "min_rank": args.min_rank,
        "pick_types": args.pick_types,
    }

    print(f"Loading NBA slate from {args.nba}...")
    nba = load_nba(args.nba)
    print(f"  {len(nba)} NBA props loaded")

    cbb = load_cbb(args.cbb)
    print(f"  {len(cbb)} CBB props loaded")

    nhl = None
    if args.nhl:
        try:
            nhl = load_nhl(args.nhl)
            nhl = attach_standard_refs(nhl)
            print(f"  {len(nhl)} NHL props loaded")
        except Exception as e:
            print(f"  WARNING: Could not load NHL file: {e}")
            nhl = None

    soccer = None
    if args.soccer:
        try:
            soccer = load_soccer(args.soccer)
            soccer = attach_standard_refs(soccer)
            print(f"  {len(soccer)} Soccer props loaded")
        except Exception as e:
            print(f"  WARNING: Could not load Soccer file: {e}")
            soccer = None

    # ✅ Attach Standard sibling refs AFTER normalized columns exist
    nba = attach_standard_refs(nba)
    cbb = attach_standard_refs(cbb)

    print("Building combined slate...")
    combined = build_combined_slate(nba, cbb, nhl, soccer)

    # ✅ Attach Standard refs for combined too
    combined = attach_standard_refs(combined)

    print(f"  {len(combined)} total props")

    def pool(df, pt=None):
        return filter_eligible(
            df,
            args.min_hit_rate,
            args.min_edge,
            args.min_rank,
            tiers if tiers else None,
            pt if pt is not None else pick_types,
        )

    nba_pool = pool(nba)
    cbb_pool = pool(cbb)
    combo_pool = pool(combined)
    print(f"  NBA eligible: {len(nba_pool)} | CBB eligible: {len(cbb_pool)} | Combined: {len(combo_pool)}")

    print("Generating tickets + workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    all_ticket_groups = []
    leg_sizes = [3, 4, 5, 6]

    def gen_tickets(pool_df, sport_label, bg_hdr, sport_prefix, pick_type_filter=None):
        rows_out = []
        for n in leg_sizes:
            sub_pool = pool_df if pick_type_filter is None else pool_df[pool_df["pick_type"].isin([pick_type_filter])]
            tickets = build_tickets(sub_pool, n, args.max_tickets)
            if tickets:
                pt_label = pick_type_filter or "Mix"
                sheet_name = f"{sport_prefix} {pt_label} {n}-Leg"[:31] if pick_type_filter else f"{sport_prefix} Mix {n}-Leg"[:31]
                write_ticket_sheet(wb, tickets, sheet_name, bg_hdr, label=f"{sport_label} {pt_label}")
                rows_out.append((sheet_name, tickets, None))
                print(f"  {sheet_name}: {len(tickets)} tickets")
        return rows_out

    # NBA tickets by pick type
    for pt in ["Goblin", "Standard", "Demon"]:
        pt_pool = pool(nba, [pt])
        if len(pt_pool) >= 3:
            all_ticket_groups += gen_tickets(pt_pool, "NBA", C["hdr_nba"], "NBA", pt)

    # NBA Mix
    if len(nba_pool) >= 3:
        all_ticket_groups += gen_tickets(nba_pool, "NBA", C["hdr_nba"], "NBA Mix")

    # CBB tickets by pick type
    for pt in ["Goblin", "Standard", "Demon"]:
        pt_pool = pool(cbb, [pt])
        if len(pt_pool) >= 3:
            all_ticket_groups += gen_tickets(pt_pool, "CBB", C["hdr_cbb"], "CBB", pt)

    # CBB Mix
    if len(cbb_pool) >= 3:
        all_ticket_groups += gen_tickets(cbb_pool, "CBB", C["hdr_cbb"], "CBB Mix")


    # NHL tickets
    if nhl is not None and len(nhl) > 0:
        nhl_pool = pool(nhl)
        if len(nhl_pool) >= 3:
            for pt in ["Goblin", "Standard", "Demon"]:
                pt_pool = pool(nhl, [pt])
                if len(pt_pool) >= 3:
                    all_ticket_groups += gen_tickets(pt_pool, "NHL", C["hdr_nhl"], "NHL", pt)
            all_ticket_groups += gen_tickets(nhl_pool, "NHL", C["hdr_nhl"], "NHL Mix")

    # Soccer tickets
    if soccer is not None and len(soccer) > 0:
        soccer_pool = pool(soccer)
        if len(soccer_pool) >= 3:
            for pt in ["Goblin", "Standard", "Demon"]:
                pt_pool = pool(soccer, [pt])
                if len(pt_pool) >= 3:
                    all_ticket_groups += gen_tickets(pt_pool, "Soccer", C["hdr_soccer"], "Soccer", pt)
            all_ticket_groups += gen_tickets(soccer_pool, "Soccer", C["hdr_soccer"], "Soccer Mix")

    # Combined NBA+CBB tickets (all pick types mixed)
    if len(combo_pool) >= 3:
        all_ticket_groups += gen_tickets(combo_pool, "COMBO", C["hdr_mix"], "COMBO")

    # Cross-sport Standard Mix (enforce mix)
    nba_std = pool(nba, ["Standard"])
    cbb_std = pool(cbb, ["Standard"])
    std_mix_pool = pd.concat([nba_std, cbb_std], ignore_index=True).sort_values("rank_score", ascending=False)
    if len(std_mix_pool) >= 3:
        print("Generating cross-sport Standard Mix tickets...")
        for n in leg_sizes:
            tickets = build_tickets(std_mix_pool, n, args.max_tickets, require_mix=True)
            if tickets:
                sheet_name = f"MIX Standard {n}-Leg"[:31]
                write_ticket_sheet(wb, tickets, sheet_name, C["hdr_mix"], label="NBA+CBB Standard")
                all_ticket_groups.append((sheet_name, tickets, C["mix"]))
                print(f"  {sheet_name}: {len(tickets)} tickets")

    # Cross-sport Goblin Mix (enforce mix)
    nba_gob = pool(nba, ["Goblin"])
    cbb_gob = pool(cbb, ["Goblin"])
    gob_mix_pool = pd.concat([nba_gob, cbb_gob], ignore_index=True).sort_values("rank_score", ascending=False)
    if len(gob_mix_pool) >= 3:
        print("Generating cross-sport Goblin Mix tickets...")
        for n in leg_sizes:
            tickets = build_tickets(gob_mix_pool, n, args.max_tickets, require_mix=True)
            if tickets:
                sheet_name = f"MIX Goblin {n}-Leg"[:31]
                write_ticket_sheet(wb, tickets, sheet_name, C["goblin"], label="NBA+CBB Goblin")
                all_ticket_groups.append((sheet_name, tickets, C["goblin"]))
                print(f"  {sheet_name}: {len(tickets)} tickets")

    print("Writing slate sheets...")
    write_slate_sheet(wb, combined, "Full Slate", C["hdr"], "ALL")
    write_slate_sheet(wb, nba, "NBA Slate", C["hdr_nba"], "NBA")
    write_slate_sheet(wb, cbb, "CBB Slate", C["hdr_cbb"], "CBB")
    if nhl is not None and len(nhl) > 0:
        write_slate_sheet(wb, nhl, "NHL Slate", C["hdr_nhl"], "NHL")
    if soccer is not None and len(soccer) > 0:
        write_slate_sheet(wb, soccer, "Soccer Slate", C["hdr_soccer"], "Soccer")

    write_summary(wb, nba, cbb, combined, all_ticket_groups, args.date, thresholds, nhl=nhl, soccer=soccer)

    # Reorder: put SUMMARY + slate sheets at the front
    desired_first = ["SUMMARY", "Full Slate", "NBA Slate", "CBB Slate", "NHL Slate", "Soccer Slate"]
    for sname in reversed(desired_first):
        if sname in wb.sheetnames:
            wb.move_sheet(wb[sname], offset=-(len(wb.sheetnames) - 1))

    wb.save(args.output)
    print(f"\n✅ Saved -> {args.output}")
    print(f"   Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Web output (FINAL only)
    if args.write_web:
        print("\nWriting GitHub Pages web outputs (FINAL tickets only)...")
        final_groups = build_final_web_ticket_groups(
            nba_pool, cbb_pool,
            min_hit_rate=thresholds.get("min_hit_rate", 0.70),
            min_edge=thresholds.get("min_edge", 2.0),
            min_rank=thresholds.get("min_rank", 5.0),
        )
        payload = ticket_groups_to_payload(final_groups, args.date, thresholds)
        write_web_outputs(payload, args.web_outdir)
        if args.also_root:
            write_web_outputs(payload, outdir=".")
        print("✅ Web outputs complete (FINAL only).")


if __name__ == "__main__":
    main()