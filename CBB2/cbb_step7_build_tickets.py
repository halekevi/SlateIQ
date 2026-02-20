#!/usr/bin/env python3
"""
cbb_step7_build_tickets.py
---------------------------
Reads step6_ranked_cbb.xlsx (Tier A sheet) and builds optimized
PrizePicks tickets for CBB.

Same logic as NBA step9:
- 3-leg tickets for best risk/reward (5x payout Standard)
- One player per game (game diversity)
- Hit rate >= min_hit_rate required
- Legs deduplicated across tickets
- Correct payout rates per pick type

PrizePicks payouts:
  Standard: 2-leg 3x | 3-leg 5x | 4-leg 10x
  Goblin:   2-leg 1.25x | 3-leg 1.7x | 4-leg 2.5x
  Demon:    2-leg 1.85x | 3-leg 3x | 4-leg 5.5x

Run:
  py -3.14 cbb_step7_build_tickets.py --input step6_ranked_cbb.xlsx --output cbb_tickets.xlsx
"""

from __future__ import annotations

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Payout Tables ─────────────────────────────────────────────────────────────
# Base Power Play payouts (all correct) — Standard lines only
POWER_PLAY_BASE = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 37.5}

# Base Flex Play payouts — Standard lines only
# Format: {n_legs: {n_correct: multiplier}}
FLEX_PLAY_BASE = {
    2: {2: 3.0},
    3: {3: 3.0,  2: 1.0},
    4: {4: 6.0,  3: 1.5},
    5: {5: 10.0, 4: 2.0,  3: 0.4},
    6: {6: 25.0, 5: 2.0,  4: 0.4},
}

# Goblin deviation multipliers on Power Play top payout (6-leg observed, extrapolated for others)
# Key = deviation level (1=closest to standard, 3=furthest)
# Each Goblin in ticket compounds: multiply modifiers together
GOBLIN_POWER_MOD = {1: 0.840, 2: 0.747, 3: 0.707}   # 31.5/37.5, 28/37.5, 26.5/37.5
GOBLIN_FLEX_MOD  = {1: 0.800, 2: 0.720, 3: 0.600}   # 20/25, 18/25, 15/25

# Demon deviation multipliers on Power Play top payout (6-leg observed)
DEMON_POWER_MOD  = {1: 1.627, 2: 2.400, 3: 2.720}   # 61/37.5, 90/37.5, 102/37.5
DEMON_FLEX_MOD   = {1: 1.600, 2: 1.520, 3: 1.560}   # 40/25, 38/25, 39/25

COLORS = {
    'Standard': '2874A6',
    'Goblin':   '6C3483',
    'Demon':    'C0392B',
    'Best Mix': '1E8449',
}
HDR_COLOR = '1C1C1C'
DIR_OVER  = 'C8F7C5'
DIR_UNDER = 'F7C5C5'


def _deviation_level(pick_type: str, line: float, standard_line: float = None) -> int:
    """Estimate deviation level (1-3) for Goblin/Demon lines.
    Without a known standard line we default to level 1 (conservative).
    If standard_line is provided, compute deviation buckets.
    """
    if standard_line is None or standard_line == 0:
        return 1
    diff = abs(line - standard_line)
    pct  = diff / abs(standard_line) if standard_line != 0 else 0
    if pct < 0.10:   return 1
    elif pct < 0.20: return 2
    else:            return 3


def calc_ticket_payout(ticket: list, n_legs: int, play_type: str = 'power') -> dict:
    """
    Calculate realistic payout for a mixed ticket.
    ticket: list of row dicts with 'pick_type' and optionally 'line'
    play_type: 'power' or 'flex'
    Returns dict with top_payout, partial payouts for flex, stake_to_win_100
    """
    base_top = POWER_PLAY_BASE.get(n_legs, 37.5) if play_type == 'power' else FLEX_PLAY_BASE.get(n_legs, {}).get(n_legs, 25.0)

    power_mod = 1.0
    flex_mod  = 1.0

    for row in ticket:
        pt = str(row.get('pick_type', 'Standard')).strip().lower()
        dev = 1  # default — we don't have standard_line reference in most cases
        if 'gob' in pt:
            power_mod *= GOBLIN_POWER_MOD.get(dev, 0.840)
            flex_mod  *= GOBLIN_FLEX_MOD.get(dev, 0.800)
        elif 'dem' in pt:
            power_mod *= DEMON_POWER_MOD.get(dev, 1.627)
            flex_mod  *= DEMON_FLEX_MOD.get(dev, 1.600)

    if play_type == 'power':
        top = round(base_top * power_mod, 2)
        stake = round(100 / top, 2) if top > 0 else 0
        return {'top_payout': top, 'stake_to_win_100': stake, 'play_type': 'Power Play'}
    else:
        flex_base = FLEX_PLAY_BASE.get(n_legs, {})
        partials = {}
        for n_correct, mult in flex_base.items():
            if n_correct == n_legs:
                partials[n_correct] = round(mult * flex_mod, 2)
            else:
                partials[n_correct] = round(mult, 2)  # partial payouts unchanged by deviation
        top = partials.get(n_legs, 25.0)
        stake = round(100 / top, 2) if top > 0 else 0
        return {'top_payout': top, 'stake_to_win_100': stake, 'play_type': 'Flex Play', 'partials': partials}


def get_payout(pick_label: str, n_legs: int) -> float:
    """Legacy function — returns base Power Play payout for pure pick type pools."""
    return POWER_PLAY_BASE.get(n_legs, 37.5)


def stake_to_win(target: float, pick_label: str, n_legs: int) -> float:
    base = POWER_PLAY_BASE.get(n_legs, 37.5)
    return round(target / base, 2)


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def norm_pick_type(x: str) -> str:
    t = str(x or '').strip().lower()
    if 'gob' in t: return 'Goblin'
    if 'dem' in t: return 'Demon'
    return 'Standard'


def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets: int, used_legs: set) -> list:
    pool = pool.sort_values('rank_score', ascending=False).copy()
    tickets = []
    used_keys = set()

    # Determine game key columns
    team_col = 'team_abbr' if 'team_abbr' in pool.columns else 'team'
    opp_col  = 'opp_team_abbr' if 'opp_team_abbr' in pool.columns else 'opp_team'

    # Allow more players per game as leg count increases
    if n_legs <= 3:
        max_per_game = 1
    elif n_legs <= 5:
        max_per_game = 2
    else:
        max_per_game = 3

    for _, anchor in pool.iterrows():
        lk = (anchor['player'], anchor['prop_type'], float(anchor['line']))
        if lk in used_legs:
            continue

        ticket = [anchor]
        t = str(anchor.get(team_col, '')).strip()
        o = str(anchor.get(opp_col,  '')).strip()
        # Track how many players per game key
        game_counts = {}
        game_key = tuple(sorted([t, o]))
        game_counts[game_key] = 1

        for _, row in pool.iterrows():
            if len(ticket) >= n_legs: break
            if row['player'] == anchor['player']: continue
            lk2 = (row['player'], row['prop_type'], float(row['line']))
            if lk2 in used_legs: continue
            rt = str(row.get(team_col, '')).strip()
            ro = str(row.get(opp_col,  '')).strip()
            gk = tuple(sorted([rt, ro]))
            if game_counts.get(gk, 0) >= max_per_game:
                continue
            ticket.append(row)
            game_counts[gk] = game_counts.get(gk, 0) + 1

        if len(ticket) == n_legs:
            tkey = frozenset((r['player'], r['prop_type'], str(r['line'])) for r in ticket)
            if tkey not in used_keys:
                used_keys.add(tkey)
                tickets.append(ticket)
                for r in ticket:
                    used_legs.add((r['player'], r['prop_type'], float(r['line'])))

        if len(tickets) >= max_tickets:
            break

    return tickets


def write_tickets_sheet(wb, sheet_name, tickets, n_legs, pick_label):
    if not tickets:
        return
    ws = wb.create_sheet(sheet_name)
    tab_color = COLORS.get(pick_label, '1E8449')
    payout    = get_payout(pick_label, n_legs)

    cols  = ['#', 'Player', 'Team', 'Opp', 'Prop', 'Type', 'Line',
             'Direction', 'Edge', 'Hit Rate', 'L5 Avg', 'Season Avg', 'Rank Score']
    widths= [4,   24,       6,      6,      16,     10,     7,
             10,          7,      10,       9,           11,          11]

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    cur = 1
    for t_idx, ticket in enumerate(tickets, 1):
        avg_score = sum(float(r['rank_score']) for r in ticket) / len(ticket)
        avg_hr    = sum(float(r['line_hit_rate']) for r in ticket if pd.notna(r.get('line_hit_rate'))) / max(1, sum(1 for r in ticket if pd.notna(r.get('line_hit_rate'))))

        # Calculate deviation-aware payouts for both play types
        ticket_dicts = [r._asdict() if hasattr(r, '_asdict') else dict(r) for r in ticket]
        pp_info   = calc_ticket_payout(ticket_dicts, n_legs, 'power')
        flex_info = calc_ticket_payout(ticket_dicts, n_legs, 'flex')
        pp_top    = pp_info['top_payout']
        flex_top  = flex_info['top_payout']
        pp_stake  = pp_info['stake_to_win_100']

        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(cols))
        hcell = ws.cell(row=cur, column=1,
            value=(f"  Ticket #{t_idx}  ·  {n_legs}-Leg {pick_label}"
                   f"  ·  Power: {pp_top}x (${pp_stake:.0f} to win $100)"
                   f"  ·  Flex: {flex_top}x"
                   f"  ·  Avg Hit Rate: {avg_hr:.0%}"
                   f"  ·  Est Win Prob: {avg_hr**n_legs:.0%}"
                   f"  ·  Avg Score: {avg_score:.2f}"))
        hcell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        hcell.fill      = PatternFill('solid', start_color=tab_color)
        hcell.alignment = Alignment(vertical='center', horizontal='left')
        ws.row_dimensions[cur].height = 22
        cur += 1

        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=cur, column=ci, value=h)
            c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=9)
            c.fill      = PatternFill('solid', start_color=HDR_COLOR)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = thin_border()
        ws.row_dimensions[cur].height = 18
        cur += 1

        team_col = 'team_abbr' if 'team_abbr' in ticket[0].index else 'team'
        opp_col  = 'opp_team_abbr' if 'opp_team_abbr' in ticket[0].index else 'opp_team'

        for leg_i, row in enumerate(ticket, 1):
            bg  = 'F8F9FA' if leg_i % 2 == 0 else 'FFFFFF'
            hr  = row.get('line_hit_rate', '')
            hr  = f"{float(hr):.0%}" if pd.notna(hr) and hr != '' else ''
            l5  = round(float(row['stat_last5_avg']), 1) if pd.notna(row.get('stat_last5_avg')) and row.get('stat_last5_avg') != '' else ''
            ssn = round(float(row['stat_season_avg']), 1) if pd.notna(row.get('stat_season_avg')) and row.get('stat_season_avg') != '' else ''
            edg = round(float(row['edge']), 1) if pd.notna(row.get('edge')) and row.get('edge') != '' else ''
            direction = str(row.get('final_bet_direction') or row.get('bet_direction') or 'OVER')

            vals = [leg_i, row['player'], row.get(team_col,''), row.get(opp_col,''),
                    row['prop_type'], row['pick_type'], row['line'],
                    direction, edg, hr, l5, ssn,
                    round(float(row['rank_score']), 2)]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=cur, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = thin_border()
                col_name = cols[ci - 1]
                if col_name == 'Direction':
                    bg2 = DIR_OVER if val == 'OVER' else DIR_UNDER
                    c.fill = PatternFill('solid', start_color=bg2)
                    c.font = Font(bold=True, name='Arial', size=9)
                else:
                    c.fill = PatternFill('solid', start_color=bg)
            cur += 1

        cur += 1  # spacer


def write_summary_sheet(wb, all_ticket_sets):
    ws = wb.create_sheet('SUMMARY', 0)
    headers = ['Sheet', 'Ticket #', 'Legs', 'Type', 'Power Payout',
               '$? to win $100', 'Flex Payout', 'Avg Hit Rate', 'Est Win %', 'Avg Score', 'Players']
    widths  = [20,      10,        6,      12,   13,
               14,              12,           13,            12,          13,          60]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=HDR_COLOR)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    row = 2
    for sheet_name, tickets, n_legs, pick_label in all_ticket_sets:
        for t_idx, ticket in enumerate(tickets, 1):
            avg_score   = sum(float(r['rank_score']) for r in ticket) / len(ticket)
            avg_hr      = sum(float(r['line_hit_rate']) for r in ticket if pd.notna(r.get('line_hit_rate'))) / max(1, sum(1 for r in ticket if pd.notna(r.get('line_hit_rate'))))
            ticket_dicts = [r._asdict() if hasattr(r, '_asdict') else dict(r) for r in ticket]
            pp_info     = calc_ticket_payout(ticket_dicts, n_legs, 'power')
            flex_info   = calc_ticket_payout(ticket_dicts, n_legs, 'flex')
            players     = ' | '.join(f"{r['player']} {r.get('final_bet_direction','OVER')} {r['prop_type']} {r['line']}" for r in ticket)

            vals = [sheet_name, t_idx, n_legs, pick_label,
                    f"{pp_info['top_payout']}x", f"${pp_info['stake_to_win_100']:.0f}",
                    f"{flex_info['top_payout']}x",
                    f"{avg_hr:.0%}", f"{avg_hr**n_legs:.0%}",
                    round(avg_score, 2), players]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                c.border    = thin_border()
                c.fill      = PatternFill('solid', start_color='F8F9FA' if row % 2 == 0 else 'FFFFFF')
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input',        default='step6_ranked_cbb.xlsx')
    ap.add_argument('--sheet',        default='TIER_A')
    ap.add_argument('--output',       default='cbb_tickets.xlsx')
    ap.add_argument('--min_hit_rate', type=float, default=0.7,
                    help='CBB default 0.7 (lower than NBA 0.8 due to smaller sample)')
    ap.add_argument('--max_tickets',  type=int,   default=10)
    ap.add_argument('--legs',         default='3,4,5,6')
    ap.add_argument('--no_fantasy',   action='store_true', default=True)
    args = ap.parse_args()

    leg_counts = [int(x.strip()) for x in args.legs.split(',')]

    print(f"→ Loading: {args.input}")

    # Load full ranked output - we filter per pool type below
    try:
        all_tiers = pd.read_excel(args.input, sheet_name='ALL')
    except Exception:
        all_tiers = pd.read_excel(args.input, sheet_name=None)
        all_tiers = list(all_tiers.values())[0]

    all_tiers['pick_type'] = all_tiers['pick_type'].astype(str).apply(norm_pick_type)

    # Exclude fantasy by default
    if args.no_fantasy and 'prop_norm' in all_tiers.columns:
        all_tiers = all_tiers[~all_tiers['prop_norm'].astype(str).str.lower().eq('fantasy')]
        print(f"  Excluded fantasy props")

    # Standard: include Tier A + B (Standard lines have smaller edges, score lower)
    # Goblin/Demon: Tier A only (soft lines, lots of them)
    std_df = all_tiers[
        (all_tiers['pick_type'] == 'Standard') &
        (all_tiers['tier'].astype(str).isin(['A', 'B']))
    ].copy()

    gob_df = all_tiers[
        (all_tiers['pick_type'] == 'Goblin') &
        (all_tiers['tier'].astype(str).eq('A'))
    ].copy()

    dem_df = all_tiers[
        (all_tiers['pick_type'] == 'Demon') &
        (all_tiers['tier'].astype(str).eq('A'))
    ].copy()

    print(f"  Standard (Tier A+B): {len(std_df)} | Goblin (Tier A): {len(gob_df)} | Demon (Tier A): {len(dem_df)}")

    # Apply hit rate filter — Standard gets a slightly lower threshold
    std_min_hr = max(0.60, args.min_hit_rate - 0.10)
    std_df = std_df[pd.to_numeric(std_df['line_hit_rate'], errors='coerce') >= std_min_hr]
    gob_df = gob_df[pd.to_numeric(gob_df['line_hit_rate'], errors='coerce') >= args.min_hit_rate]
    dem_df = dem_df[pd.to_numeric(dem_df['line_hit_rate'], errors='coerce') >= args.min_hit_rate]

    print(f"  After hit rate filter — Standard >= {std_min_hr}: {len(std_df)} | Goblin/Demon >= {args.min_hit_rate}: {len(gob_df)}/{len(dem_df)}")

    def dedup(df):
        key = ['player', 'prop_type', 'line', 'final_bet_direction'] if 'final_bet_direction' in df.columns else ['player', 'prop_type', 'line']
        return df.sort_values('rank_score', ascending=False).drop_duplicates(subset=key)

    std_df = dedup(std_df)
    gob_df = dedup(gob_df)
    dem_df = dedup(dem_df)

    # Best Mix: Standard + Goblin/Demon Tier A, with Standard given a 0.25 bonus
    # to compensate for naturally lower scores vs Goblin lines
    mix_df = pd.concat([std_df, gob_df, dem_df]).copy()
    mix_df['rank_score_mix'] = mix_df.apply(
        lambda r: float(r['rank_score']) + (0.25 if r['pick_type'] == 'Standard' else 0.0), axis=1)
    mix_df = mix_df.sort_values('rank_score_mix', ascending=False)

    std_pool = std_df.drop_duplicates(subset=['player'])
    gob_pool = gob_df.drop_duplicates(subset=['player'])
    dem_pool = dem_df.drop_duplicates(subset=['player'])
    all_pool = mix_df.drop_duplicates(subset=['player'])

    # For ticket building rank_score, use rank_score_mix for Best Mix
    all_pool = all_pool.copy()
    all_pool['rank_score'] = all_pool['rank_score_mix']

    pools = [('Standard', std_pool), ('Goblin', gob_pool),
             ('Demon', dem_pool),    ('Best Mix', all_pool)]

    wb = Workbook()
    wb.remove(wb.active)
    all_ticket_sets = []

    for pick_label, pool in pools:
        if len(pool) < 2:
            print(f"  ⚠ Skipping {pick_label} — not enough props ({len(pool)})")
            continue
        used_legs: set = set()
        for n_legs in leg_counts:
            if len(pool) < n_legs: continue
            tickets   = build_tickets(pool.copy(), n_legs, args.max_tickets, used_legs)
            sheet_name= f"{pick_label} {n_legs}-Leg"
            write_tickets_sheet(wb, sheet_name, tickets, n_legs, pick_label)
            all_ticket_sets.append((sheet_name, tickets, n_legs, pick_label))
            print(f"  {sheet_name}: {len(tickets)} tickets")

    write_summary_sheet(wb, all_ticket_sets)
    wb.save(args.output)
    print(f"\n✅ Saved → {args.output}")
    print(f"\nPayout reference (to win $100):")
    for label in ('Standard','Goblin','Demon'):
        stakes = [f"{n}-leg: ${stake_to_win(100,label,n):.0f}" for n in leg_counts]
        print(f"  {label:<10} {' | '.join(stakes)}")


if __name__ == '__main__':
    main()
