#!/usr/bin/env python3
"""
step7_rank_props.py  (VECTORIZED 2026-03-01)

PERF: All 12 .apply() calls and 2 row-by-row list comprehensions replaced with
      vectorized pandas/NumPy operations. Estimated 3-4x faster on 8,000+ row slates.
      Excel write engine switched from openpyxl → xlsxwriter (~5x faster write).

PATCH (2026-02-26):
- Fix edge_adj_dr to be direction-aware: UNDERs now get a positive edge
  contribution when projection < line.
- Support projection building for volume props (2PTA/2PTM, 3PTA/3PTM, FTA/FTM, FGA/FGM).
- Adds prop_norm aliases for volume props.

PATCH (2026-02-23):
- grading-informed reweight of scoring components.
"""

from __future__ import annotations

import argparse
import numpy as np
import pandas as pd

# UTF-8 safe Excel export
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

# -------------------- helpers --------------------

def _to_num(s):
    return pd.to_numeric(s, errors="coerce")

def _norm_pick_type_series(s: pd.Series) -> pd.Series:
    t = s.astype(str).str.strip().str.lower()
    return np.where(t.str.contains("gob"), "Goblin",
           np.where(t.str.contains("dem"), "Demon", "Standard"))

# -------------------- weights --------------------

_PROP_WEIGHTS = {
    "pts": 1.03, "reb": 1.06, "ast": 1.05, "stl": 1.08, "blk": 1.02,
    "stocks": 1.04, "3ptmade": 1.03, "3ptattempted": 0.88,
    "fg3m": 1.03, "fg3a": 0.88, "fg2m": 1.01, "fg2a": 0.92,
    "twopointersmade": 1.01, "twopointersattempted": 0.92,
    "fgm": 0.99, "fga": 0.99,
    "freethrowsmade": 1.01, "freethrowsattempted": 0.98,
    "ftm": 1.01, "fta": 0.98,
    "tov": 0.94, "pf": 0.85, "personalfouls": 0.85,
    "pr": 1.01, "pa": 1.01, "ra": 1.02, "pra": 0.99, "fantasy": 1.00,
}

_PROP_HR_PRIOR_OVER = {
    "stl": 0.697, "fantasy": 0.674, "3ptmade": 0.623, "fg3m": 0.623,
    "reb": 0.617, "ra": 0.600, "ast": 0.593, "freethrowsmade": 0.583,
    "ftm": 0.583, "pr": 0.568, "pts": 0.566, "stocks": 0.547, "blk": 0.545,
    "pra": 0.545, "fga": 0.558, "pa": 0.557, "fgm": 0.519, "fg2m": 0.528,
    "twopointersmade": 0.528, "fg2a": 0.463, "twopointersattempted": 0.463,
    "tov": 0.484, "3ptattempted": 0.444, "fg3a": 0.444,
    "pf": 0.424, "personalfouls": 0.424, "fta": 0.545,
    "freethrowsattempted": 0.545,
}

# Under overrides where the under rate != 1 - over_rate
_PROP_HR_PRIOR_UNDER_OVERRIDE = {
    "fantasy": 0.371,
    "fga": 0.645, "fg2a": 0.645, "twopointersattempted": 0.645,
    "reb": 0.591, "pa": 0.590,
    "pts": 0.540, "pr": 0.540, "pra": 0.540,
}

_RELIABILITY_MAP = {"Standard": 1.00, "Goblin": 1.06, "Demon": 0.75}

# -------------------- projection fallback --------------------

_PLAYER_PREFIX_BY_PROP = {
    "fga": "fga", "fgm": "fgm", "fg2a": "fg2a", "fg2m": "fg2m",
    "fg3a": "fg3a", "fg3m": "fg3m", "fta": "fta", "ftm": "ftm",
}

_COMBO_CORRECTIONS = {"pr": 1.05, "pa": 1.06, "ra": 1.08, "pra": 1.07, "fantasy": 1.15}

def _edge_transform_series(edge: pd.Series, cap: float = 3.0, power: float = 0.85) -> pd.Series:
    """Vectorized power-transform with sign preservation."""
    sign = np.sign(edge)
    clipped = np.clip(edge.abs(), 0, cap)
    return sign * (clipped ** power)

def _tier_from_score_series(score: pd.Series) -> pd.Series:
    return np.where(score >= 2.50, "A",
           np.where(score >= 1.75, "B",
           np.where(score >= 1.10, "C", "D")))

def _write_xlsx_openpyxl(output_path: str, out: pd.DataFrame, elig_mask: pd.Series) -> None:
    """Write XLSX with explicit UTF-8 encoding using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create both sheets with UTF-8 safe values
    for sheet_name, df_sheet in [("ALL", out), ("ELIGIBLE", out.loc[elig_mask])]:
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Ensure value is properly UTF-8 encoded (especially for player names)
                if isinstance(value, str):
                    # Force string through UTF-8 encode/decode to ensure proper handling
                    value = value.encode('utf-8').decode('utf-8')
                elif pd.isna(value):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Set encoding in workbook properties
    wb.properties.encoding = 'UTF-8'
    wb.save(output_path)
    print(f"✅ Saved → {output_path} (openpyxl, UTF-8 encoded)")

# -------------------- main --------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="step6_with_team_role_context.csv")
    ap.add_argument("--output", default="step7_ranked_props.xlsx")
    args = ap.parse_args()

    print(f"→ Loading: {args.input}")
    df = pd.read_csv(args.input, dtype=str, encoding="utf-8-sig", 
                     engine='python').fillna("")
    
    # Explicitly ensure all string columns are str type (not object with mixed types)
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str)
    out = df.copy()

    for col, default in [("line", ""), ("pick_type", "Standard"), ("prop_norm", "")]:
        if col not in out.columns:
            out[col] = default

    if "prop_norm" not in out.columns or out["prop_norm"].eq("").all():
        if "prop_type" in out.columns:
            out["prop_norm"] = out["prop_type"].astype(str).str.lower()

    # Normalize prop names
    _PROP_NORM_MAP = {
        "3-pt made": "fg3m", "3-pt attempted": "fg3a",
        "3pt made": "fg3m", "3pt attempted": "fg3a",
        "three pointers made": "fg3m", "three pointers attempted": "fg3a",
        "3-ptm": "fg3m", "3-pta": "fg3a", "3ptm": "fg3m", "3pta": "fg3a",
        "two pointers made": "fg2m", "two pointers attempted": "fg2a",
        "2 pointers made": "fg2m", "2 pointers attempted": "fg2a",
        "2pt made": "fg2m", "2pt attempted": "fg2a",
        "2-pt made": "fg2m", "2-pt attempted": "fg2a",
        "2-ptm": "fg2m", "2-pta": "fg2a", "2ptm": "fg2m", "2pta": "fg2a",
        "free throws made": "ftm", "free throws attempted": "fta",
        "freethrowsmade": "ftm", "freethrowsattempted": "fta",
        "ft made": "ftm", "ft attempted": "fta", "ftm": "ftm", "fta": "fta",
        "fg attempted": "fga", "fg made": "fgm",
        "field goals attempted": "fga", "field goals made": "fgm",
        "fga": "fga", "fgm": "fgm",
        "fg3a": "fg3a", "fg3m": "fg3m", "fg2a": "fg2a", "fg2m": "fg2m",
    }
    out["prop_norm"] = (out["prop_norm"].astype(str).str.lower().str.strip()
                        .map(lambda x: _PROP_NORM_MAP.get(x, x)))

    prop_norm_s = out["prop_norm"].astype(str).str.lower().str.strip()
    line_num    = _to_num(out["line"])
    pick_type_s = pd.Series(_norm_pick_type_series(out["pick_type"]), index=out.index)

    # ── VECTORIZED PROJECTION ─────────────────────────────────────────────────
    v5  = _to_num(out.get("stat_last5_avg",  ""))
    v10 = _to_num(out.get("stat_last10_avg", ""))
    vs  = _to_num(out.get("stat_season_avg", ""))

    # Weighted blend (50/30/20) with partial weight normalization
    w5 = np.where(v5.notna(),  0.50, 0.0)
    w10= np.where(v10.notna(), 0.30, 0.0)
    ws = np.where(vs.notna(),  0.20, 0.0)
    total_w = w5 + w10 + ws
    proj_raw = (
        v5.fillna(0)  * w5 +
        v10.fillna(0) * w10 +
        vs.fillna(0)  * ws
    )
    proj_raw = np.where(total_w > 0.1, proj_raw / total_w, np.nan)

    # Fallback for volume props: look for {prefix}_player_last5_avg etc.
    missing_proj = np.isnan(proj_raw)
    if missing_proj.any():
        for prop_key, prefix in _PLAYER_PREFIX_BY_PROP.items():
            mask = missing_proj & (prop_norm_s == prop_key)
            if not mask.any():
                continue
            for col_cand in [f"{prefix}_player_last5_avg", f"{prefix}_last5_avg"]:
                if col_cand in out.columns:
                    fb = _to_num(out[col_cand])
                    proj_raw = np.where(mask & fb.notna(), fb, proj_raw)
                    break

    # Combo/fantasy correction
    corr = prop_norm_s.map(lambda x: _COMBO_CORRECTIONS.get(x, 1.0)).values
    proj = pd.Series(proj_raw * corr, index=out.index)
    out["projection"] = proj

    out["edge"]     = proj - line_num
    out["abs_edge"] = out["edge"].abs()

    # ── FORCED OVER / BET DIRECTION ───────────────────────────────────────────
    forced = pick_type_s.isin(["Goblin", "Demon"]).astype(int)
    out["forced_over_only"] = forced

    bet_dir = np.where(forced.eq(1), "OVER",
              np.where(_to_num(out["edge"]) >= 0, "OVER", "UNDER"))
    out["bet_direction"] = bet_dir

    # ── ELIGIBILITY ───────────────────────────────────────────────────────────
    miss        = line_num.isna() | proj.isna()
    neg_forced  = forced.eq(1) & (_to_num(out["edge"]) < 0)
    eligible    = (~miss & ~neg_forced).astype(int)
    void_reason = pd.Series("", index=out.index)
    void_reason = void_reason.where(~miss,       "NO_PROJECTION_OR_LINE")
    void_reason = void_reason.where(~neg_forced,  "FORCED_OVER_NEG_EDGE")
    out["eligible"]    = eligible
    out["void_reason"] = void_reason

    elig_mask = eligible.eq(1)

    # ── VECTORIZED EDGE TRANSFORM ─────────────────────────────────────────────
    out["edge_dr"] = _edge_transform_series(_to_num(out["edge"]))

    # ── VECTORIZED LINE HIT RATE ──────────────────────────────────────────────
    # Direction-aware: pick the right column priority
    bet_is_under = pd.Series(bet_dir, index=out.index) == "UNDER"

    def _pick_first_valid(*col_names) -> pd.Series:
        result = pd.Series(np.nan, index=out.index)
        for col in col_names:
            if col in out.columns:
                v = _to_num(out[col])
                result = result.where(result.notna(), v)
        return result

    hr5_over  = _pick_first_valid("line_hit_rate_over_ou_5",  "line_hit_rate_over_5",  "last5_hit_rate")
    hr10_over = _pick_first_valid("line_hit_rate_over_ou_10", "line_hit_rate_over_10")
    hr5_under = _pick_first_valid("line_hit_rate_under_ou_5", "line_hit_rate_under_5")
    hr10_under= _pick_first_valid("line_hit_rate_under_ou_10","line_hit_rate_under_10")

    # Derived under from counts if direct column missing
    l5o = _to_num(out.get("last5_over",  ""))
    l5u = _to_num(out.get("last5_under", ""))
    denom_ou = (l5o + l5u).replace(0, np.nan)
    derived_under = l5u / denom_ou
    hr5_under = hr5_under.where(hr5_under.notna(), derived_under)

    # No push fallback (1 - over) when push==0
    l5p = _to_num(out.get("last5_push", ""))
    hr5_under = hr5_under.where(hr5_under.notna(),
        np.where(l5p.fillna(0) == 0, 1.0 - hr5_over, np.nan))

    hr5  = np.where(bet_is_under, hr5_under, hr5_over)
    hr10 = np.where(bet_is_under, hr10_under, hr10_over)
    hr5  = pd.Series(hr5,  index=out.index)
    hr10 = pd.Series(hr10, index=out.index)

    # Blend 5 and 10 game windows
    line_hit_rate = (
        np.where(hr5.notna() & hr10.notna(), hr5 * 0.50 + hr10 * 0.50,
        np.where(hr5.notna(),  hr5,
        np.where(hr10.notna(), hr10, np.nan)))
    )
    out["line_hit_rate"] = pd.Series(line_hit_rate, index=out.index)

    # ── VECTORIZED MINUTES CERTAINTY ──────────────────────────────────────────
    _MIN_TIER_MAP = {"HIGH": 1.00, "MEDIUM": 0.90, "LOW": 0.75}
    out["minutes_certainty"] = (
        out.get("minutes_tier", pd.Series("", index=out.index))
        .astype(str).str.upper()
        .map(lambda x: _MIN_TIER_MAP.get(x, 0.80))
    )

    # ── VECTORIZED PROP WEIGHT / RELIABILITY ─────────────────────────────────
    out["prop_weight"]      = prop_norm_s.map(lambda x: _PROP_WEIGHTS.get(x, 0.93))
    out["reliability_mult"] = pick_type_s.map(lambda x: _RELIABILITY_MAP.get(x, 0.97))

    # ── VECTORIZED DEF ADJUSTMENT ─────────────────────────────────────────────
    def_rank = _to_num(out.get("OVERALL_DEF_RANK", ""))
    def_adj  = ((def_rank - 15.0) / 15.0 * 0.06).fillna(0.0)
    out["def_adj"] = def_adj

    # ── GAME CONTEXT ADJUSTMENT (Step 6b: Vegas lines) ────────────────────────
    # ctx_adj: -0.08 low total on combo prop, -0.05 blowout risk, -0.15 both
    ctx_adj  = _to_num(out["ctx_adj"]).fillna(0.0)  if "ctx_adj"  in out.columns else pd.Series(0.0, index=out.index)
    out["ctx_adj"] = ctx_adj

    # ── SCHEDULE / REST ADJUSTMENT (Step 6c: B2B, rest days) ─────────────────
    # rest_adj: -0.10 B2B, 0.00 baseline (1-day rest), +0.02 two days, +0.04 three+
    rest_adj = _to_num(out["rest_adj"]).fillna(0.0) if "rest_adj" in out.columns else pd.Series(0.0, index=out.index)
    out["rest_adj"] = rest_adj

    proj_base = _to_num(out["projection"])
    out["projection_adj"] = proj_base * (1.0 + def_adj + ctx_adj + rest_adj)
    out["edge_adj"]       = out["projection_adj"] - line_num

    # ── VECTORIZED EDGE_ADJ_DR (direction-aware) ──────────────────────────────
    edge_adj_signed = np.where(bet_is_under, -_to_num(out["edge_adj"]), _to_num(out["edge_adj"]))
    out["edge_adj_dr"] = _edge_transform_series(pd.Series(edge_adj_signed, index=out.index))

    # ── VECTORIZED DEF RANK SIGNAL ────────────────────────────────────────────
    signal_raw = ((def_rank - 1.0) / 29.0 * 2.0 - 1.0)
    def_signal = np.where(bet_is_under, -signal_raw, signal_raw)
    out["def_rank_signal"] = pd.Series(def_signal, index=out.index)

    # ── VECTORIZED PROP HIT RATE PRIOR ───────────────────────────────────────
    base_prior = prop_norm_s.map(lambda x: _PROP_HR_PRIOR_OVER.get(x, 0.545))
    under_prior = prop_norm_s.map(
        lambda x: _PROP_HR_PRIOR_UNDER_OVERRIDE.get(x, 1.0 - _PROP_HR_PRIOR_OVER.get(x, 0.545))
    )
    out["prop_hr_prior"] = np.where(bet_is_under, under_prior, base_prior)

    # ── VECTORIZED AVG VS LINE ────────────────────────────────────────────────
    for col in ("stat_last5_avg", "stat_last10_avg", "stat_season_avg"):
        out[col + "_num"] = _to_num(out[col]) if col in out.columns else pd.Series(np.nan, index=out.index)

    line_safe = line_num.replace(0, np.nan)

    def _avg_vs_line_vec(avg_col: str, w: float) -> pd.Series:
        v = _to_num(out[avg_col + "_num"]) if (avg_col + "_num") in out.columns else pd.Series(np.nan, index=out.index)
        raw = np.clip((v - line_safe) / line_safe, -1.0, 1.0)
        raw = np.where(bet_is_under, -raw, raw)
        return pd.Series(np.where(v.notna() & line_safe.notna(), raw * w, np.nan), index=out.index)

    avl5  = _avg_vs_line_vec("stat_last5_avg",  0.50)
    avl10 = _avg_vs_line_vec("stat_last10_avg", 0.30)
    avls  = _avg_vs_line_vec("stat_season_avg", 0.20)

    wt5  = np.where(_to_num(out.get("stat_last5_avg_num",  "")).notna() & line_safe.notna(), 0.50, 0.0)
    wt10 = np.where(_to_num(out.get("stat_last10_avg_num", "")).notna() & line_safe.notna(), 0.30, 0.0)
    wts  = np.where(_to_num(out.get("stat_season_avg_num", "")).notna() & line_safe.notna(), 0.20, 0.0)
    total_avl_w = pd.Series(wt5 + wt10 + wts, index=out.index)

    avg_vs_line = (avl5.fillna(0) + avl10.fillna(0) + avls.fillna(0))
    avg_vs_line = avg_vs_line.where(total_avl_w > 0.1, 0.0)
    out["avg_vs_line"] = avg_vs_line

    # ── Z-SCORE (direction-aware) ─────────────────────────────────────────────
    def zcol(s: pd.Series, direction_aware: bool = False) -> pd.Series:
        x = pd.to_numeric(s, errors="coerce")
        result = pd.Series(0.0, index=x.index)
        if direction_aware and "bet_direction" in out.columns:
            for direction in ("OVER", "UNDER"):
                dir_mask = elig_mask & (out["bet_direction"].astype(str).str.upper() == direction)
                if dir_mask.sum() < 2:
                    continue
                mu = x[dir_mask].mean()
                sd = x[dir_mask].std()
                if sd and not np.isnan(sd) and sd > 1e-9:
                    z_vals = (x[dir_mask] - mu) / sd
                    result.loc[dir_mask.index[dir_mask]] = z_vals.values
            return result
        mu = x[elig_mask].mean()
        sd = x[elig_mask].std()
        if sd and not np.isnan(sd) and sd > 1e-9:
            return (x - mu) / sd
        return result

    out["edge_z"]        = zcol(out["edge"],           direction_aware=True)
    out["line_hit_z"]    = zcol(out["line_hit_rate"],   direction_aware=True)
    out["min_z"]         = zcol(out["minutes_certainty"])
    out["def_rank_z"]    = zcol(out["def_rank_signal"],  direction_aware=True)
    out["avg_vs_line_z"] = zcol(out["avg_vs_line"],      direction_aware=True)
    out["prop_hr_z"]     = zcol(out["prop_hr_prior"],    direction_aware=True)

    # ── FINAL SCORE ───────────────────────────────────────────────────────────
    # ctx_adj and rest_adj already baked into projection_adj/edge_adj_dr.
    # We also apply a small direct score bonus/penalty so B2B and blowout risk
    # visibly shift rank even if edge is positive.
    b2b_penalty    = np.where(out.get("b2b_flag",    pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.20, 0.0)
    blowout_penalty= np.where(out.get("blowout_risk", pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.10, 0.0)
    low_total_pen  = np.where(out.get("low_total_flag", pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.10, 0.0)

    score = (
        _to_num(out["edge_adj_dr"]).fillna(0.0)      * 0.85
        + _to_num(out["line_hit_z"]).fillna(0.0)     * 0.85
        + _to_num(out["avg_vs_line_z"]).fillna(0.0)  * 0.75
        + _to_num(out["def_rank_z"]).fillna(0.0)     * 0.80
        + _to_num(out["prop_hr_z"]).fillna(0.0)      * 0.50
        + _to_num(out["min_z"]).fillna(0.0)          * 0.25
        + pd.Series(b2b_penalty,     index=out.index)  # B2B fatigue
        + pd.Series(blowout_penalty, index=out.index)  # blowout bench risk
        + pd.Series(low_total_pen,   index=out.index)  # low O/U game
    )
    score = (
        score
        * _to_num(out["prop_weight"]).fillna(1.0)
        * _to_num(out["reliability_mult"]).fillna(1.0)
    )
    score = score.where(elig_mask, np.nan)

    out["rank_score"] = score
    out["tier"] = pd.Series(
        _tier_from_score_series(_to_num(out["rank_score"])), index=out.index
    )
    out.loc[~elig_mask, "tier"] = "D"

    # ── WRITE XLSX (with explicit UTF-8 handling) ──────────────────────────────
    if HAS_XLSXWRITER:
        try:
            with pd.ExcelWriter(args.output, engine="xlsxwriter", 
                               engine_kwargs={'options': {'strings_to_urls': False}}) as w:
                out.to_excel(w, sheet_name="ALL", index=False)
                out.loc[elig_mask].to_excel(w, sheet_name="ELIGIBLE", index=False)
            print(f"✅ Saved → {args.output} (xlsxwriter, UTF-8 encoded)")
        except Exception as e:
            print(f"⚠️  xlsxwriter failed: {e}")
            _write_xlsx_openpyxl(args.output, out, elig_mask)
    else:
        _write_xlsx_openpyxl(args.output, out, elig_mask)

    print(f"✅ Saved → {args.output}")
    print(f"ALL rows      : {len(out)}")
    std = pick_type_s.eq("Standard")
    print(f"STANDARD rows : {int(std.sum())}")
    print(f"GOB_DEM rows  : {int((~std).sum())}")
    print()
    print("Tier counts (ALL):")
    print(out["tier"].value_counts().to_string())
    print()
    print("Ineligible reason breakdown:")
    vr = out.loc[~elig_mask, "void_reason"].value_counts()
    print(vr.to_string() if len(vr) else "(none)")
    print()
    print("Score percentiles (eligible):")
    rs = _to_num(out.loc[elig_mask, "rank_score"])
    print(rs.quantile([0.50, 0.70, 0.80, 0.85, 0.90, 0.95]).round(3).to_string())


if __name__ == "__main__":
    main()
